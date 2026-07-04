"""if_report — figures, shared-display micrographs/montages, and the Excel
workbook for the ``if_intensity`` mode.

Ported verbatim from the LOCKED panQKI standalone ``03_figures.py`` (SuperPlot
Prism cosmetics + Okabe-Ito + LUT-by-wavelength micrographs + "Sam's floor"
shared display ranges) and ``04_build_excel.py`` (explorable workbook). All
cosmetics are locked; only parameters are sourced from ``cfg`` (fig_seed,
pixel_size, ceiling/floor percentiles, scalebar). Imported by ``if_intensity``.
Human / Homo sapiens.
"""
from __future__ import annotations

import re
from pathlib import Path

import numpy as np
import pandas as pd


# ── LOCKED Prism cosmetics (from make_figures_v5_prism.py) ────────────────────
BAR_WIDTH = 0.55; BORDER_LW = 2.8
ELINEWIDTH = 2.0; CAPSIZE = 6; CAPTHICK = 2.0
SPINE_LW = 1.4; TICK_LEN = 5; TICK_W = 1.4
TITLE_FS = 14; YLABEL_FS = 10; TICKLBL_FS = 12
DOT_S = 72; DOT_ALPHA = 0.45; DOT_EDGE_LW = 0.4; DOT_JITTER = 0.10
FOV_S = 15; FOV_EDGE_LW = 0.25; FOV_JITTER = 0.08

# ── LOCKED Okabe-Ito (WT / KO family — identical across all figures) ──────────
WT_COLOR = "#E69F00"; KO_COLOR = "#009E73"; SEC_COLOR = "#999999"

Z_HLINE, Z_BASE, Z_BORDER, Z_FOV, Z_DOT, Z_ERR = 1, 2, 3, 4, 5, 6


def _genotype_label(g: str) -> str:
    return "WT" if str(g).upper() == "WT" else "KO"


def _slug(s: str) -> str:
    return re.sub(r"[^A-Za-z0-9]+", "_", str(s)).strip("_") or "IF"


# ═══════════════════════════════════════════════════════════════════════════
# SuperPlots
# ═══════════════════════════════════════════════════════════════════════════
def _face(hex_col, alpha=DOT_ALPHA):
    import matplotlib.colors as mcolors
    r, g, b = mcolors.to_rgb(hex_col)
    return (r, g, b, alpha)


def _well_tint(hex_col, rank, n):
    import matplotlib.colors as mcolors
    blend = 0.0 if n <= 1 else 0.6 * rank / (n - 1)
    r, g, b = mcolors.to_rgb(hex_col)
    return (r + (1 - r) * blend, g + (1 - g) * blend, b + (1 - b) * blend, 0.55)


def _stars(p):
    if p is None or (isinstance(p, float) and np.isnan(p)):
        return "n/a"
    if p < 0.001:
        return "***"
    if p < 0.01:
        return "**"
    if p < 0.05:
        return "*"
    return "ns"


def _fmt_p(p):
    if p is None or (isinstance(p, float) and np.isnan(p)):
        return "p=n/a"
    if p >= 0.01:
        return f"p={p:.3f}"
    if p >= 1e-4:
        return f"p={p:.4f}"
    return f"p={p:.2e}"


def _style_axes(ax, yticks_lo, yticks_hi):
    ax.spines["top"].set_visible(False); ax.spines["right"].set_visible(False)
    ax.spines["left"].set_linewidth(SPINE_LW); ax.spines["bottom"].set_linewidth(SPINE_LW)
    ax.spines["left"].set_bounds(yticks_lo, yticks_hi)
    ax.grid(False)
    ax.tick_params(axis="both", which="major", length=TICK_LEN, width=TICK_W,
                   direction="out", color="black", labelsize=TICKLBL_FS)
    for tl in ax.get_yticklabels():
        tl.set_fontweight("bold")


def _superplot_panel(ax, groups, y_label, title, fig_seed, p=None, floor=None):
    """bar = mean of per-well means; per-nucleus cloud shaded by well; per-well
    mean dots over the bar; optional sec-only floor + optional sig bracket."""
    rng = np.random.default_rng(fig_seed)
    xs = list(range(len(groups)))
    all_top = []
    for x, gp in zip(xs, groups):
        wm = np.asarray(gp["well_means"], float); wm = wm[~np.isnan(wm)]
        mean = float(np.mean(wm)) if len(wm) else np.nan
        sd = float(np.std(wm, ddof=1)) if len(wm) > 1 else np.nan
        ax.bar(x, mean, width=BAR_WIDTH, facecolor=gp["color"], edgecolor="none", zorder=Z_BASE)
        ax.bar(x, mean, width=BAR_WIDTH, facecolor="none", edgecolor="black",
               linewidth=BORDER_LW, zorder=Z_BORDER)
        if not np.isnan(sd):
            ax.errorbar(x, mean, yerr=sd, fmt="none", ecolor="black",
                        elinewidth=ELINEWIDTH, capsize=CAPSIZE, capthick=CAPTHICK, zorder=Z_ERR)
        cbw = gp.get("cloud_by_well") or {}
        for wi, (well, arr) in enumerate(sorted(cbw.items())):
            arr = np.asarray(arr, float); arr = arr[~np.isnan(arr)]
            if not len(arr):
                continue
            jit = rng.uniform(-FOV_JITTER, FOV_JITTER, len(arr))
            ax.scatter(x + jit, arr, s=FOV_S,
                       facecolors=[_well_tint(gp["color"], wi, len(cbw))] * len(arr),
                       edgecolors="black", linewidths=FOV_EDGE_LW, zorder=Z_FOV)
        if len(wm):
            jit = rng.uniform(-DOT_JITTER, DOT_JITTER, len(wm))
            ax.scatter(x + jit, wm, s=DOT_S, facecolors=[_face(gp["color"])] * len(wm),
                       edgecolors="black", linewidths=DOT_EDGE_LW, zorder=Z_DOT)
            all_top.append(np.nanmax(wm))
        if not np.isnan(mean + (sd if not np.isnan(sd) else 0)):
            all_top.append(mean + (sd if not np.isnan(sd) else 0))
    if floor is not None and not np.isnan(floor.get("value", np.nan)):
        ax.axhline(floor["value"], color=SEC_COLOR, lw=1.4, ls="--", alpha=0.9, zorder=Z_HLINE)
        all_top.append(floor["value"])
        ax.text(xs[-1] + 0.05, floor["value"], "  " + floor.get("label", "sec-only floor"),
                va="center", ha="left", fontsize=6.5, color=SEC_COLOR)

    data_top = max([t for t in all_top if t is not None and not np.isnan(t)] + [1e-9])
    y_lo = min(0.0, min([np.nanmin(np.asarray(g["well_means"], float))
                         for g in groups if len(g["well_means"])] + [0.0]))
    y_hi = data_top * 1.30
    ax.set_ylim(y_lo, y_hi)
    if p is not None and len(groups) >= 2:
        by = data_top * 1.12; th = data_top * 0.03
        ax.plot([0, 0, 1, 1], [by - th, by, by, by - th], color="black", lw=1.0,
                zorder=Z_ERR + 1, clip_on=False)
        ax.text(0.5, by + th * 0.3, f"{_stars(p)}  ({_fmt_p(p)})", ha="center", va="bottom",
                fontsize=8, fontweight="bold")
    _style_axes(ax, y_lo, data_top)
    ax.set_xticks(xs)
    ax.set_xticklabels([g["label"] for g in groups], fontsize=TICKLBL_FS,
                       fontweight="bold", rotation=0)
    ax.set_xlim(-0.6, len(groups) - 0.4)
    ax.set_ylabel(y_label, fontsize=YLABEL_FS, fontweight="bold", labelpad=4)
    ax.set_title(title, fontsize=TITLE_FS, fontweight="bold", pad=8)


def _save_single(out_dir, groups, y_label, title, fname, fig_seed,
                 p=None, floor=None, footnote=""):
    import matplotlib.pyplot as plt
    fig, ax = plt.subplots(figsize=(3.4, 4.8))
    fig.subplots_adjust(left=0.26, right=0.9, top=0.88, bottom=0.20)
    _superplot_panel(ax, groups, y_label, title, fig_seed, p=p, floor=floor)
    if footnote:
        fig.text(0.04, 0.005, footnote, ha="left", va="bottom", fontsize=5.5,
                 color="#555555", style="italic")
    out_dir.mkdir(parents=True, exist_ok=True)
    path = out_dir / fname
    fig.savefig(path, dpi=600, bbox_inches="tight", facecolor="white")
    plt.close(fig)
    print(f"  [OK] {path}")


def _cloud(per_nucleus, wells, col="nuc_mean_qki"):
    out = {}
    if per_nucleus is None or per_nucleus.empty or "well" not in per_nucleus.columns:
        return out
    sub = per_nucleus[per_nucleus["well"].isin(wells)]
    for well, grp in sub.groupby("well"):
        out[int(well)] = grp[col].dropna().values
    return out


def _wmeans(per_well, wells, col):
    return per_well[per_well["well"].isin(wells)][col].dropna().values


def _p_for(stats, **kw):
    sub = stats
    for k, v in kw.items():
        sub = sub[sub[k] == v]
    if sub.empty:
        return None
    val = sub.iloc[0].get("welch_p", None)
    return None if (val is None or (isinstance(val, float) and np.isnan(val))) else float(val)


def make_superplots(fig_dir, per_well, per_nucleus, stats, per_secondary, pooled, plate, cfg):
    """Emit per-secondary 2-group (nuc_mean + ratio), 3-group floor, and pooled
    fold n3 SuperPlots at 600 DPI."""
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    plt.rcParams.update({"font.family": "sans-serif",
                         "font.sans-serif": ["Arial", "DejaVu Sans"], "figure.dpi": 150})
    fig_dir = Path(fig_dir)
    fig_seed = int(cfg.if_intensity.fig_seed)

    Y_NUC = "Nuclear signal mean (a.u., within DAPI mask)"
    Y_RATIO = "Signal / DAPI (a.u.)"
    Y_FOLD = "Signal fold over secondary-only"

    for sec, gdef in per_secondary.items():
        wt_w, ko_w = gdef["WT"], gdef["KO"]
        sec_wt = _seconly_wells_for(plate, "WT", sec)
        sec_ko = _seconly_wells_for(plate, "KO", sec)
        n_note = f"n={len(wt_w)} WT / {len(ko_w)} KO wells (secondary {sec})"
        for metric, ylab, tag in [("nuc_mean_qki", Y_NUC, "nuc_mean"),
                                  ("ratio_qki_over_dapi", Y_RATIO, "ratio")]:
            groups = [
                dict(label="WT", color=WT_COLOR, well_means=_wmeans(per_well, wt_w, metric),
                     cloud_by_well=_cloud(per_nucleus, wt_w) if metric == "nuc_mean_qki" else {}),
                dict(label="KO", color=KO_COLOR, well_means=_wmeans(per_well, ko_w, metric),
                     cloud_by_well=_cloud(per_nucleus, ko_w) if metric == "nuc_mean_qki" else {}),
            ]
            floor_vals = (np.concatenate([_wmeans(per_well, sec_wt, metric),
                                          _wmeans(per_well, sec_ko, metric)])
                          if (sec_wt or sec_ko) else np.array([]))
            floor_val = float(np.nanmean(floor_vals)) if floor_vals.size else np.nan
            p = _p_for(stats, analysis="per_secondary_raw", secondary=sec, metric=metric)
            _save_single(fig_dir, groups, ylab, f"signal {sec}  (WT vs KO)",
                         f"sec{sec}_{tag}_2group.png", fig_seed, p=p,
                         floor=dict(value=floor_val, label="secondary-only floor"),
                         footnote=f"{n_note}. Absolute intensity, exposure-matched. "
                                  f"Dashed = secondary-only floor.")

        groups3 = [
            dict(label="WT", color=WT_COLOR, well_means=_wmeans(per_well, wt_w, "nuc_mean_qki"),
                 cloud_by_well=_cloud(per_nucleus, wt_w)),
            dict(label="KO", color=KO_COLOR, well_means=_wmeans(per_well, ko_w, "nuc_mean_qki"),
                 cloud_by_well=_cloud(per_nucleus, ko_w)),
            dict(label="sec-only", color=SEC_COLOR,
                 well_means=_wmeans(per_well, sec_wt + sec_ko, "nuc_mean_qki"),
                 cloud_by_well=_cloud(per_nucleus, sec_wt + sec_ko)),
        ]
        _save_single(fig_dir, groups3, Y_NUC, f"signal {sec}: WT / KO / secondary-only",
                     f"sec{sec}_nuc_mean_3group.png", fig_seed,
                     footnote=f"{n_note} + secondary-only controls. Absolute nuclear signal.")

    for metric, tag in [("wellfold_nuc_mean", "nuc_mean"), ("wellfold_ratio", "ratio")]:
        groups = [
            dict(label="WT", color=WT_COLOR,
                 well_means=_wmeans(per_well, pooled["WT"], metric), cloud_by_well={}),
            dict(label="KO", color=KO_COLOR,
                 well_means=_wmeans(per_well, pooled["KO"], metric), cloud_by_well={}),
        ]
        p = _p_for(stats, analysis="pooled_fold_n3", metric=metric)
        _save_single(fig_dir, groups, Y_FOLD, "pooled (fold / sec-only)",
                     f"pooled_fold_{tag}_n3.png", fig_seed, p=p,
                     floor=dict(value=1.0, label="1x (= secondary-only)"),
                     footnote=f"Pooled n={len(pooled['WT'])} wells/genotype "
                              f"({','.join(map(str, pooled['WT']))} vs "
                              f"{','.join(map(str, pooled['KO']))}); fold-over-sec-only "
                              f"makes secondaries poolable. Welch t.")


# ═══════════════════════════════════════════════════════════════════════════
# Shared-display-range micrographs ("Sam's floor")
# ═══════════════════════════════════════════════════════════════════════════
def _seconly_wells_for(plate, genotype, secondary):
    return sorted(
        w for w, v in plate.items()
        if v["arm"] == "seconly" and v["genotype"] == genotype
        and v["secondary"] == str(secondary)
    )


def _norm(img, lo_p=2, hi_p=99.5):
    lo, hi = np.percentile(img, [lo_p, hi_p])
    return np.clip((img - lo) / max(hi - lo, 1e-6), 0, 1)


def _lut(img01, wavelength):
    h, w = img01.shape
    rgb = np.zeros((h, w, 3), float)
    s = str(wavelength)
    if "647" in s or "640" in s:
        rgb[..., 0] = img01; rgb[..., 1] = img01           # yellow
    elif "568" in s or "565" in s or "561" in s:
        rgb[..., 0] = img01; rgb[..., 2] = img01           # magenta
    else:
        rgb[..., 2] = img01                                # blue (DAPI)
    return rgb


def _focus_scores(stack):
    from scipy.ndimage import laplace
    out = np.empty(stack.shape[0], np.float64)
    for z in range(stack.shape[0]):
        p = stack[z]
        out[z] = float(laplace(p).var() * (p.mean() + 1e-6))
    return out


def _windowed_maxproj(dstack, qstack, frac):
    """Central in-focus windowed max-projection: keep the central `frac` of
    planes centered on the sharpest DAPI plane (searched within the central
    band). Brian's improvement over a muddy full-stack max."""
    dstack = np.asarray(dstack, np.float32)
    qstack = np.asarray(qstack, np.float32)
    if dstack.ndim == 2:
        return dstack, qstack
    nz = dstack.shape[0]
    if nz <= 1:
        return dstack[0], qstack[0]
    frac = float(frac) if frac and frac > 0 else 1.0
    n_keep = max(3, int(round(nz * frac)))
    if n_keep >= nz:
        return dstack.max(axis=0), qstack.max(axis=0)
    lo_z = int(nz * (1.0 - frac) / 2.0)
    hi_z = nz - lo_z
    if hi_z <= lo_z:
        lo_z, hi_z = 0, nz
    scores = _focus_scores(dstack)
    z0 = lo_z + int(np.argmax(scores[lo_z:hi_z]))
    half = n_keep // 2
    a, b = max(0, z0 - half), min(nz, z0 + half + 1)
    return dstack[a:b].max(axis=0), qstack[a:b].max(axis=0)


def _compute_shared_ranges(rep_images, plate, per_secondary, cfg):
    """One SHARED (lo, hi) per secondary: hi = SIGNAL(WT-primary) ceiling pct;
    lo = secondary-only floor pct. Applied identically to WT/KO/sec-only."""
    ceil_pct = float(cfg.if_intensity.display_ceiling_pct)
    floor_pct = float(cfg.if_intensity.display_floor_pct)

    def _pix(wells):
        arrs = [rep_images[w][1].ravel() for w in wells if w in rep_images]
        return np.concatenate(arrs) if arrs else None

    ranges = {}
    for sec, grp in per_secondary.items():
        wt_pix = _pix(grp["WT"])
        sec_pix = _pix(_seconly_wells_for(plate, "WT", sec)
                       + _seconly_wells_for(plate, "KO", sec))
        if wt_pix is None:
            ranges[sec] = None
            continue
        hi = float(np.percentile(wt_pix, ceil_pct))
        lo = (float(np.percentile(sec_pix, floor_pct)) if sec_pix is not None
              else float(np.percentile(wt_pix, 1)))
        if hi <= lo:  # degenerate guard
            lo, hi = float(np.percentile(wt_pix, 2)), float(np.percentile(wt_pix, 99.5))
        ranges[sec] = (lo, hi)
        print(f"  [display] secondary {sec}: SHARED range lo={lo:.1f} hi={hi:.1f} "
              f"(ceiling=WT-primary {ceil_pct}p, floor=sec-only {floor_pct}p)")
    return ranges


def _build_rep_from_zstack(cfg, plate):
    """Build well -> (dapi2d, qki2d, secondary) representative images from a
    separate z-stack folder using a central in-focus windowed max-projection."""
    from bioio import BioImage
    import bioio_bioformats
    from . import if_intensity as _ifi

    zdir = Path(cfg.if_intensity.micrograph_zstack_dir)
    frac = float(cfg.if_intensity.micrograph_z_window_frac)
    dapi_key = cfg.if_intensity.dapi_channel_key
    rep = {}
    wells = _ifi._discover_wells(zdir)
    for w in sorted(wells):
        if w not in plate:
            continue
        qki_key = plate[w]["qki_channel"]
        f = wells[w][0]
        try:
            img = BioImage(str(f), reader=bioio_bioformats.Reader)
            # largest >=3ch scene (allow Z>1)
            best = None
            for sc in img.scenes:
                img.set_scene(sc)
                d = img.dims
                c = d.C if "C" in d.order else 1
                area = d.Y * d.X
                if c >= 3 and (best is None or area > best[1]):
                    best = (sc, area)
            if best is None:
                continue
            img.set_scene(best[0])
            names = [str(n) for n in list(img.channel_names or [])]
            didx = [i for i, n in enumerate(names) if dapi_key in n][0]
            qidx = [i for i, n in enumerate(names) if qki_key in n][0]
            dstack = np.asarray(img.get_image_data("ZYX", T=0, C=didx)).astype(np.float32)
            qstack = np.asarray(img.get_image_data("ZYX", T=0, C=qidx)).astype(np.float32)
            dproj, qproj = _windowed_maxproj(dstack, qstack, frac)
            rep[w] = (dproj.astype(np.float64), qproj.astype(np.float64), plate[w]["secondary"])
        except Exception as exc:
            print(f"  [WARN] zstack micrograph well{w}: {exc}")
    return rep


def _render_micrograph(out_dir, qki, dapi, wavelength, title, fname, px_um,
                       qki_range=None, scalebar_um=20):
    import matplotlib.pyplot as plt
    import matplotlib.patches as mpatches
    if qki_range is not None and qki_range[1] > qki_range[0]:
        lo, hi = qki_range
        qki01 = np.clip((qki - lo) / (hi - lo), 0, 1)
    else:
        qki01 = _norm(qki)
    q = _lut(qki01, wavelength)
    d = _lut(_norm(dapi), "405")
    rgb = np.clip(q + d, 0, 1)
    fig, ax = plt.subplots(figsize=(5, 5), dpi=150)
    ax.imshow(rgb); ax.axis("off"); ax.set_title(title, fontsize=9)
    bar_px = scalebar_um / px_um
    h, w = qki.shape
    x0, y0 = w * 0.70, h * 0.93
    ax.add_patch(mpatches.Rectangle((x0, y0), bar_px, h * 0.012, color="white", ec="none"))
    ax.text(x0 + bar_px / 2, y0 - h * 0.015, f"{scalebar_um:g} um",
            color="white", ha="center", va="bottom", fontsize=8)
    out_dir.mkdir(parents=True, exist_ok=True)
    fig.savefig(out_dir / fname, dpi=600, bbox_inches="tight", facecolor="black")
    plt.close(fig)
    print(f"  [OK] {out_dir / fname}")


def make_micrographs(micro_dir, montage_dir, plate, per_secondary, rep_images, px_um, cfg):
    """Per-well shared-display micrographs + the money montage (WT | KO | sec-only)."""
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    import matplotlib.patches as mpatches

    micro_dir = Path(micro_dir); montage_dir = Path(montage_dir)
    scalebar_um = float(cfg.if_intensity.scalebar_um)
    slug = _slug(cfg.experiment.name)

    if str(cfg.if_intensity.micrograph_source) == "zstack" and cfg.if_intensity.micrograph_zstack_dir:
        rep_images = _build_rep_from_zstack(cfg, plate)
    rep_images = rep_images or {}
    if not rep_images:
        print("  [WARN] no representative micrograph images available; skipping micrographs")
        return

    print("\n=== Micrographs (shared display range, LUT by wavelength) ===")
    ranges = _compute_shared_ranges(rep_images, plate, per_secondary, cfg)
    micro_dir.mkdir(parents=True, exist_ok=True)
    with open(micro_dir / "_display_ranges.txt", "w", encoding="utf-8") as f:
        f.write("# SHARED signal display range per secondary (applied to WT/KO/sec-only)\n")
        f.write(f"# ceiling = WT-primary {cfg.if_intensity.display_ceiling_pct}th pct; "
                f"floor = sec-only {cfg.if_intensity.display_floor_pct}th pct\n")
        for sec, rng in ranges.items():
            f.write(f"secondary {sec}: {rng}\n")

    # per-well micrographs
    for w in sorted(rep_images):
        dapi, qki, sec = rep_images[w]
        if w not in plate:
            continue
        pv = plate[w]
        wav = pv["secondary"]
        qki_range = ranges.get(pv["secondary"])
        gl = _genotype_label(pv["genotype"])
        if pv["arm"] == "primary":
            title = f"{gl} — signal ({sec})"
        else:
            title = f"Secondary-only ({sec}) — {gl}"
        fname = f"well{w}_{pv['genotype']}_{pv['arm']}_{sec}.png"
        try:
            _render_micrograph(micro_dir, qki, dapi, wav, title, fname, px_um,
                               qki_range=qki_range, scalebar_um=scalebar_um)
        except Exception as exc:
            print(f"  [ERROR] micrograph well{w}: {exc}")

    # money montage per secondary: WT | KO | secondary-only
    print("\n=== Publication montages (WT | KO | secondary-only, shared display) ===")
    montage_dir.mkdir(parents=True, exist_ok=True)
    for sec, grp in per_secondary.items():
        rng = ranges.get(sec)
        sec_ws = (_seconly_wells_for(plate, "WT", sec)
                  or _seconly_wells_for(plate, "KO", sec))
        picks = [("WT — signal", grp["WT"][0] if grp["WT"] else None),
                 ("KO — signal", grp["KO"][0] if grp["KO"] else None),
                 ("Secondary-only", sec_ws[0] if sec_ws else None)]
        picks = [(lab, w) for lab, w in picks if w is not None and w in rep_images]
        if len(picks) < 2:
            continue
        fig, axes = plt.subplots(1, len(picks), figsize=(4.2 * len(picks), 4.5), dpi=150)
        if len(picks) == 1:
            axes = [axes]
        last_shape = None
        for ax, (lab, w) in zip(axes, picks):
            dapi, qki, _sec = rep_images[w]
            last_shape = qki.shape
            if rng is not None and rng[1] > rng[0]:
                q01 = np.clip((qki - rng[0]) / (rng[1] - rng[0]), 0, 1)
            else:
                q01 = _norm(qki)
            rgb = np.clip(_lut(q01, sec) + _lut(_norm(dapi), "405"), 0, 1)
            ax.imshow(rgb); ax.axis("off")
            ax.set_title(lab, fontsize=12, fontweight="bold", color="black", pad=6)
        h, w_ = last_shape
        bar_px = scalebar_um / px_um
        x0, y0 = w_ * 0.62, h * 0.93
        axes[-1].add_patch(mpatches.Rectangle((x0, y0), bar_px, h * 0.013, color="white", ec="none"))
        axes[-1].text(x0 + bar_px / 2, y0 - h * 0.02, f"{scalebar_um:g} um", color="white",
                      ha="center", va="bottom", fontsize=9)
        fig.suptitle(f"Immunofluorescence — {sec} secondary", fontsize=13,
                     fontweight="bold", y=1.06)
        fig.subplots_adjust(left=0.01, right=0.99, top=0.88, bottom=0.01, wspace=0.03)
        path = montage_dir / f"montage_{slug}_{sec}_WT-KO-secondaryonly.png"
        fig.savefig(path, dpi=600, facecolor="white", bbox_inches="tight")
        plt.close(fig)
        print(f"  [OK] {path}")


# ═══════════════════════════════════════════════════════════════════════════
# Excel workbook (ported from 04_build_excel.py)
# ═══════════════════════════════════════════════════════════════════════════
def _autowidth(ws, get_column_letter):
    for col in ws.columns:
        maxlen = 0
        for cell in col:
            try:
                maxlen = max(maxlen, len(str(cell.value or "")))
            except Exception:
                pass
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(maxlen + 2, 10), 55)


def _write_df_sheet(wb, name, df, styles):
    Font, HDR_FILL, Alignment, get_column_letter, dataframe_to_rows = styles
    ws = wb.create_sheet(name[:31])
    if df is None or df.empty:
        ws.cell(row=1, column=1, value=f"(no rows for {name})")
        return ws
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    for c in range(1, df.shape[1] + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True); cell.fill = HDR_FILL
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(df.shape[1])}{df.shape[0] + 1}"
    _autowidth(ws, get_column_letter)
    return ws


def _build_how_to_read(ws, plate, cfg, Font, NOTE_FILL):
    overrides = cfg.if_intensity.well_secondary_overrides or {}
    override_note = (", ".join(f"well {k} -> {v}" for k, v in overrides.items())
                     if overrides else "none")
    exp_name = cfg.experiment.name or "IF antibody validation"
    lines = [
        (f"{exp_name} — IF antibody validation (WT vs KO) — HOW TO READ", True),
        ("", False),
        ("PURPOSE: validate the antibody. Does the signal give WT >> KO above the "
         "secondary-only floor, and which secondary is cleaner?", False),
        ("This is antibody VALIDATION, not a powered hypothesis test — small n is expected "
         "and stated on every figure.", False),
        ("", False),
        ("CHANNEL ROUTING (critical): each well's signal is read from ITS OWN secondary's "
         "channel. 647 wells -> 640-CSU channel; 568/'565' wells -> 561-CSU channel. DAPI = 405.", False),
        ("The other signal channel for a well has no fluorophore and is NOT signal.", False),
        ("", False),
        ("METRICS (columns):", True),
        ("  nuc_mean_qki         = per-nucleus NUCLEAR signal mean within the DAPI mask (a.u.)", False),
        ("  nuc_integrated_qki   = per-nucleus integrated signal within the DAPI mask (mean x area)", False),
        ("  qki_mean / total_qki = whole-FOV signal mean / sum (all pixels)", False),
        ("  ratio_qki_over_dapi  = TOTAL signal / TOTAL DAPI  (the 'divide-by-DAPI' metric)", False),
        ("  cyto_mean_qki        = cytoplasmic estimate via a dilated-nucleus RING (extra, not headline)", False),
        ("  nuc_over_cyto        = nuclear / cytoplasmic signal", False),
        ("  M1_pn                = 'Per-nucleus' aux (whole-FOV background-corrected / nucleus)", False),
        ("  wellfold_*           = fold-over-sec-only = well / mean(matched-genotype+channel sec-only wells)", False),
        ("                          -> dimensionless; makes the two secondaries poolable.", False),
        ("", False),
        ("REPLICATE STRUCTURE (design BOTH): replicate unit = WELL (biological). Multiple FOV/well "
         "are technical reps averaged into the well. Per-secondary raw absolute AND pooled fold.", False),
        ("", False),
        (f"CONFLICT WELLS (secondary override applied): {override_note}. See the conflict-check "
         f"sheet for the empirical channel-carrier result; the override does NOT auto-flip.", False),
        ("", False),
        ("EXPOSURE: absolute-intensity comparisons are only valid at matched exposure. See "
         "Exposure_report — the routed signal channel is asserted identical across WT/KO/sec-only.", False),
        ("", False),
        ("Human / Homo sapiens. Colors: WT=#E69F00, KO=#009E73 (Okabe-Ito). Figures = 600 DPI PNG.", False),
    ]
    for i, (txt, bold) in enumerate(lines, start=1):
        cell = ws.cell(row=i, column=1, value=txt)
        if bold:
            cell.font = Font(bold=True, size=12 if i == 1 else 11)
            if i == 1:
                cell.fill = NOTE_FILL
    ws.column_dimensions["A"].width = 110


def build_excel(xlsx_path, plate, per_nucleus, per_fov, per_well, stats,
                seconly, exposure, conflict, cfg):
    """Explorable .xlsx: How_to_read FIRST, then Plate_map, Stats_summary,
    Per_well, Per_FOV, Per_nucleus, Seconly_pooled, Exposure_report, conflict."""
    import openpyxl
    from openpyxl.utils import get_column_letter
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.styles import Font, PatternFill, Alignment

    HDR_FILL = PatternFill("solid", fgColor="BDD7EE")
    NOTE_FILL = PatternFill("solid", fgColor="FFF2CC")
    styles = (Font, HDR_FILL, Alignment, get_column_letter, dataframe_to_rows)

    plate_rows = [dict(well=w, **{k: v for k, v in plate[w].items()}) for w in sorted(plate)]
    plate_df = pd.DataFrame(plate_rows)

    wb = openpyxl.Workbook()
    ws0 = wb.active
    ws0.title = "How_to_read"
    _build_how_to_read(ws0, plate, cfg, Font, NOTE_FILL)

    _write_df_sheet(wb, "Plate_map", plate_df, styles)
    _write_df_sheet(wb, "Stats_summary", stats, styles)
    _write_df_sheet(wb, "Per_well", per_well, styles)
    _write_df_sheet(wb, "Per_FOV", per_fov, styles)
    _write_df_sheet(wb, "Per_nucleus", per_nucleus, styles)
    _write_df_sheet(wb, "Seconly_pooled", seconly, styles)
    _write_df_sheet(wb, "Exposure_report", exposure, styles)
    if conflict is not None and not conflict.empty:
        _write_df_sheet(wb, "Conflict_check", conflict, styles)

    xlsx_path = Path(xlsx_path)
    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(xlsx_path)
    print(f"[write] {xlsx_path}")
