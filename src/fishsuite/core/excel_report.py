"""PI-ready Excel deliverable writer.

Produces TWO companion Excel workbooks per run:

  * ``analysis_summary.xlsx`` — the report. A multi-sheet, formatted
    workbook intended for the PI to read. Includes an Executive_Summary
    (plain-language facts), a Comparison_Table (WT vs KO with
    Mann-Whitney U + Cliff's delta), all per-image / per-nucleus / per-spot
    data, cell morphology, thresholds, and a flattened Run_Config.

  * ``analysis_raw_data.xlsx`` — the companion. Same 4 data sheets with
    condition coloring + a Raw_README pointer back to the report. Intended
    for the PI (or anyone) to slice the data themselves.

Both workbooks share:

  * Fixed condition palette in column ``condition``
        WT       #D7E9F7  (light blue)
        KO       #FCE4CC  (light orange)
        sec-only #E6E6E6  (light gray)
  * Bold header row with #F0F0F0 fill + thin bottom border.
  * Freeze panes at A2.
  * Auto-fit column widths capped at 60 chars.
  * Numeric formatting (4-dp fractions / integer counts / 2-dp rates).
  * Sort order: WT, KO, sec-only (then by image, then by nucleus/spot id).
  * Sec-only metric cells that would be misleading (frac_nuclear with 0
    spots) are written as blank (NaN, which pandas writes as empty).

Brian's standing rules honored:

  * "Self-explanatory Excel": README + glossary covers every column.
  * "Agnostic framing": no "validates", "confirms", "active TSS / mature
    mRNA / nascent" interpretation language. Pipeline reports counts;
    biological interpretation is the reader's.
  * "Perfect Excel files": no biology assumptions, every condition
    clearly marked, every column clearly labeled.
"""
from __future__ import annotations

import json
import math
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

import numpy as np
import pandas as pd


# ---------------------------------------------------------------------------
# Shared style constants
# ---------------------------------------------------------------------------

CONDITION_FILLS: Dict[str, str] = {
    # exact strings tested case-insensitively below
    "WT":       "FFD7E9F7",  # light blue
    "KO":       "FFFCE4CC",  # light orange
    "SEC-ONLY": "FFE6E6E6",  # light gray
    "SEC_ONLY": "FFE6E6E6",
    "SECONDARY-ONLY": "FFE6E6E6",
    "NT":       "FFD7E9F7",
    "KD":       "FFFCE4CC",
}

HEADER_FILL_HEX = "FFF0F0F0"
SECTION_FILL_HEX = "FF2C3E50"  # dark blue
SUBSECTION_FILL_HEX = "FF34495E"

# Numeric format strings
FMT_FRACTION = "0.0000"
FMT_RATE_2DP = "0.00"
FMT_INT = "0"
FMT_FLOAT_GEN = "0.000"
FMT_PVAL = "0.000E+00"

# Columns whose cells should render with the integer format
INT_COLUMNS = {
    "nuclei_analyzed", "total_spots_rna1", "total_spots_rna2",
    "nuclear_spots_rna1", "cytoplasmic_spots_rna1",
    "nuclear_spots_rna2", "cytoplasmic_spots_rna2",
    "paired_count_rna1_at_0p3um", "paired_count_rna2_at_0p3um",
    "n_nuclei_border_excluded", "total_spots",
    "dapi_channel", "rna_channel", "rna2_channel",
    "voxel_xy_nm", "voxel_z_nm", "n_z",
    "nucleus_id", "nucleus_area_px", "rna_spot_count",
    "nuclear_spot_count", "cyto_spot_count",
    "n_spots_rna1", "n_spots_rna2",
    "nuclear_spot_count_rna2", "cyto_spot_count_rna2",
    "n_nuclear_rna1_rna2_overlap_per_nucleus", "n_nuclear_rna2_rna1_overlap_per_nucleus",
    "n_cytoplasmic_rna1_spots_per_cell", "n_cytoplasmic_rna2_spots_per_cell",
    "cyto_area_px", "n_voxels", "n_pix", "n_z_slices",
    "sum_rna_intensity", "sum_rna2_intensity",
    "cell_total_intensity_rna1", "cell_total_intensity_rna2",
    "cell_area_px", "sum_rna_intensity_cyto", "sum_rna2_intensity_cyto",
    "nuclear_total_intensity_rna1", "nuclear_total_intensity_rna2",
    "cytoplasmic_total_intensity_rna1", "cytoplasmic_total_intensity_rna2",
    "paired_spot_count_rna1_at_0p3um", "paired_spot_count_rna2_at_0p3um",
    "spot_id", "x_px", "y_px", "z_slice", "z_position_um",
    "spot_peak_intensity", "quality", "spot_fwhm_px",
    "integrated_intensity_fit", "in_nucleus", "in_cytoplasm",
    "paired_at_0p3um", "cell_id",
}

FRACTION_COLUMNS_PREFIX = ("frac_", "paired_fraction_", "nuclear_spot_fraction")
RATE_COLUMNS_SUFFIX = ("_per_nucleus", "_per_cell", "_per_um2", "_um")


# ---------------------------------------------------------------------------
# User-facing label substitution (2026-05-19 Brian).
#
# Every USER-FACING cell value in the README / Executive_Summary /
# Comparison_Table sheets that mentions a generic "RNA1" / "RNA2" / "DAPI"
# is remapped, post-build, to the user's preset labels (cfg.channels.rna_label
# / rna2_label / dapi_label as written into run_config.json by the runner).
#
# INTERNAL artifacts that are deliberately UNCHANGED:
#   * data sheet COLUMN NAMES (n_active_tss_per_nucleus,
#     n_nuclear_rna1_rna2_overlap_per_nucleus, paired_fraction_rna1_*, ...) —
#     these are the contract with downstream tooling
#   * docstrings / comments / Python identifiers
#
# Pulling the labels at workbook-save time (not at module import) lets a
# live preset edit propagate without reloading fishsuite.
# ---------------------------------------------------------------------------


def _resolve_labels_from_run_cfg(run_cfg_flat: Dict) -> Dict[str, str]:
    """Return ``{'rna_label', 'rna2_label', 'dapi_label'}`` from the flattened
    run_config.json. Defaults to the legacy generic names so a missing key
    renders the same as before this helper existed.
    """
    return {
        "rna_label":  str(run_cfg_flat.get("config_resolved.channels.rna_label",
                          run_cfg_flat.get("CHANNEL_RNA_LABEL", "RNA1"))),
        "rna2_label": str(run_cfg_flat.get("config_resolved.channels.rna2_label",
                          run_cfg_flat.get("CHANNEL_RNA2_LABEL", "RNA2"))),
        "dapi_label": str(run_cfg_flat.get("config_resolved.channels.dapi_label",
                          run_cfg_flat.get("CHANNEL_DAPI_LABEL", "DAPI"))),
    }


def _subst_user_text(s: Any, labels: Dict[str, str]) -> Any:
    """Replace standalone "RNA1" / "RNA2" / "DAPI" tokens in a USER-FACING
    string. Word-boundary matching (negative lookarounds on [A-Za-z0-9_]) so
    internal identifiers like ``frac_nuclear_rna1`` (lowercase) and
    ``n_nuclear_rna1_rna2_overlap_per_nucleus`` are NEVER touched. Non-string
    inputs (numbers, None, bools) are returned unchanged.
    """
    if not isinstance(s, str) or not s:
        return s
    import re as _re
    rna1 = labels["rna_label"]; rna2 = labels["rna2_label"]; dapi = labels["dapi_label"]
    if rna1 == "RNA1" and rna2 == "RNA2" and dapi == "DAPI":
        return s
    out = s
    out = _re.sub(r"(?<![A-Za-z0-9_])RNA2(?![A-Za-z0-9_])", rna2, out)
    out = _re.sub(r"(?<![A-Za-z0-9_])RNA1(?![A-Za-z0-9_])", rna1, out)
    if dapi != "DAPI":
        out = _re.sub(r"(?<![A-Za-z0-9_])DAPI(?![A-Za-z0-9_])", dapi, out)
    return out


def _relabel_worksheet(ws, labels: Dict[str, str]) -> None:
    """Walk every cell in a worksheet and apply ``_subst_user_text`` to its
    string value. Skips numeric / boolean / None cells. No-op when the
    active labels match defaults (legacy runs are bit-for-bit identical).
    """
    if (labels["rna_label"] == "RNA1"
            and labels["rna2_label"] == "RNA2"
            and labels["dapi_label"] == "DAPI"):
        return
    try:
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if isinstance(v, str):
                    new_v = _subst_user_text(v, labels)
                    if new_v != v:
                        cell.value = new_v
    except Exception:
        # Never let a relabel pass crash a workbook save.
        pass


# ---------------------------------------------------------------------------
# Curated column descriptions
# ---------------------------------------------------------------------------

PER_IMAGE_GLOSSARY: Dict[str, Tuple[str, str, str]] = {
    "image": ("str", "—",
              "Source image filename (relative to the input directory)."),
    "condition": ("str", "—",
                  "Biological condition for this image (e.g. WT, KO, Sec-Only). "
                  "Driven by the subfolder-to-condition mapping in the config."),
    "secondary_only": ("bool", "—",
                       "True if this file is a secondary-antibody-only control "
                       "(no primary FISH probe). Expected to show ~0 real RNA spots."),
    "nuclei_analyzed": ("int", "count",
                        "Number of nuclei retained for this image after border "
                        "exclusion + min-area filter."),
    "mean_spots_per_nucleus": ("float", "count",
                               "Average number of RNA1 spots per nucleus (alias of "
                               "mean_spots_per_nucleus_rna1; kept for Fiji-pipeline parity)."),
    "median_spots_per_nucleus": ("int", "count",
                                 "Median RNA1 spots per nucleus."),
    "cv_spots_per_nucleus": ("float", "—",
                             "Coefficient of variation (std/mean) of RNA1 spots per nucleus."),
    "frac_nuclei_with_ge_1_spot": ("fraction 0-1", "—",
                                   "Fraction of nuclei with >= 1 RNA1 spot."),
    "frac_nuclei_with_ge_5_spots": ("fraction 0-1", "—",
                                    "Fraction of nuclei with >= 5 RNA1 spots."),
    "frac_nuclei_with_ge_10_spots": ("fraction 0-1", "—",
                                     "Fraction of nuclei with >= 10 RNA1 spots."),
    "mean_spots_per_nucleus_rna1": ("float", "count",
                                    "Average RNA1 spots per nucleus."),
    "mean_spots_per_nucleus_rna2": ("float", "count",
                                    "Average RNA2 spots per nucleus."),
    "median_spots_per_nucleus_rna1": ("int", "count",
                                      "Median RNA1 spots per nucleus."),
    "median_spots_per_nucleus_rna2": ("float", "count",
                                      "Median RNA2 spots per nucleus."),
    "cv_spots_per_nucleus_rna1": ("float", "—",
                                  "Coefficient of variation of RNA1 spots per nucleus."),
    "cv_spots_per_nucleus_rna2": ("float", "—",
                                  "Coefficient of variation of RNA2 spots per nucleus."),
    "frac_nuclei_with_ge_1_spot_rna1": ("fraction 0-1", "—",
                                        "Fraction of nuclei with >= 1 RNA1 spot."),
    "frac_nuclei_with_ge_5_spots_rna1": ("fraction 0-1", "—",
                                         "Fraction of nuclei with >= 5 RNA1 spots."),
    "frac_nuclei_with_ge_10_spots_rna1": ("fraction 0-1", "—",
                                          "Fraction of nuclei with >= 10 RNA1 spots."),
    "frac_nuclei_with_ge_1_spot_rna2": ("fraction 0-1", "—",
                                        "Fraction of nuclei with >= 1 RNA2 spot."),
    "frac_nuclei_with_ge_5_spots_rna2": ("fraction 0-1", "—",
                                         "Fraction of nuclei with >= 5 RNA2 spots."),
    "frac_nuclei_with_ge_10_spots_rna2": ("fraction 0-1", "—",
                                          "Fraction of nuclei with >= 10 RNA2 spots."),
    "total_spots_rna1": ("int", "count",
                         "Total RNA1 spots detected in this image (all nuclei + cytoplasm)."),
    "total_spots_rna2": ("int", "count",
                         "Total RNA2 spots detected in this image."),
    "nuclear_spots_rna1": ("int", "count",
                           "RNA1 spots whose xy center falls inside any nuclear mask."),
    "cytoplasmic_spots_rna1": ("int", "count",
                               "RNA1 spots whose xy center falls in a Voronoi cytoplasm region."),
    "nuclear_spots_rna2": ("int", "count",
                           "RNA2 spots inside a nuclear mask."),
    "cytoplasmic_spots_rna2": ("int", "count",
                               "RNA2 spots in a Voronoi cytoplasm region."),
    "frac_nuclear_rna1": ("fraction 0-1", "—",
                          "Of all RNA1 spots in the image, the fraction whose xy center "
                          "falls inside a nuclear mask."),
    "frac_nuclear_rna2": ("fraction 0-1", "—",
                          "Of all RNA2 spots in the image, the fraction whose xy center "
                          "falls inside a nuclear mask."),
    "mean_cell_total_intensity_fit_rna1": (
        "float", "AU",
        "Mean per-cell summed RNA1 raw intensity over (nucleus + Voronoi cytoplasm)."),
    "median_cell_total_intensity_fit_rna1": (
        "float", "AU", "Median per-cell summed RNA1 raw intensity."),
    "cv_cell_total_intensity_fit_rna1": (
        "float", "—", "CV of per-cell summed RNA1 raw intensity."),
    "mean_cell_total_intensity_fit_rna2": (
        "float", "AU", "Mean per-cell summed RNA2 raw intensity."),
    "median_cell_total_intensity_fit_rna2": (
        "float", "AU", "Median per-cell summed RNA2 raw intensity."),
    "cv_cell_total_intensity_fit_rna2": (
        "float", "—", "CV of per-cell summed RNA2 raw intensity."),
    "mean_nuc_total_intensity_rna1": (
        "float", "AU", "Mean per-nucleus summed RNA1 raw intensity."),
    "median_nuc_total_intensity_rna1": (
        "float", "AU", "Median per-nucleus summed RNA1 raw intensity."),
    "mean_nuc_total_intensity_rna2": (
        "float", "AU", "Mean per-nucleus summed RNA2 raw intensity."),
    "median_nuc_total_intensity_rna2": (
        "float", "AU", "Median per-nucleus summed RNA2 raw intensity."),
    "mean_cell_total_spot_intensity_fit_rna1": (
        "float", "AU",
        "Mean per-cell summed RNA1 SPOT (BigFISH-fit) intensity. Sum of "
        "integrated_intensity_fit across all RNA1 spots in that cell."),
    "median_cell_total_spot_intensity_fit_rna1": (
        "float", "AU", "Median per-cell summed RNA1 spot intensity."),
    "cv_cell_total_spot_intensity_fit_rna1": (
        "float", "—", "CV of per-cell summed RNA1 spot intensity."),
    "mean_cell_total_spot_intensity_fit_rna2": (
        "float", "AU", "Mean per-cell summed RNA2 spot intensity."),
    "median_cell_total_spot_intensity_fit_rna2": (
        "float", "AU", "Median per-cell summed RNA2 spot intensity."),
    "cv_cell_total_spot_intensity_fit_rna2": (
        "float", "—", "CV of per-cell summed RNA2 spot intensity."),
    "mean_nc_ratio_total_intensity_rna1": (
        "float", "—",
        "Mean per-cell nucleus/cytoplasm intensity ratio for RNA1 raw signal."),
    "median_nc_ratio_total_intensity_rna1": (
        "float", "—", "Median N/C intensity ratio, RNA1."),
    "mean_nc_ratio_total_intensity_rna2": (
        "float", "—", "Mean N/C intensity ratio, RNA2."),
    "median_nc_ratio_total_intensity_rna2": (
        "float", "—", "Median N/C intensity ratio, RNA2."),
    # ---- Above-floor intensity image-level rollups (Brian/Sam 2026-05-20) -
    # Image-level mean/median over the per-nucleus above-floor columns.
    # Only present when output.apply_pub_contrast_floor_to_analysis is True.
    "mean_nuclear_above_floor_intensity_rna1": (
        "float", "AU",
        "Mean per-nucleus RNA1 above-floor nuclear intensity. Pub-image "
        "floor is used as a hard threshold on pixel quantification — "
        "pixels below the floor (cytoplasmic noise) are excluded from the "
        "nuclear sum. Set output.apply_pub_contrast_floor_to_analysis to "
        "enable."),
    "median_nuclear_above_floor_intensity_rna1": (
        "float", "AU",
        "Median per-nucleus RNA1 above-floor nuclear intensity."),
    "mean_nuclear_above_floor_intensity_rna2": (
        "float", "AU",
        "Mean per-nucleus RNA2 above-floor nuclear intensity."),
    "median_nuclear_above_floor_intensity_rna2": (
        "float", "AU",
        "Median per-nucleus RNA2 above-floor nuclear intensity."),
    "mean_cytoplasmic_above_floor_intensity_rna1": (
        "float", "AU",
        "Mean per-nucleus RNA1 above-floor cytoplasmic intensity."),
    "median_cytoplasmic_above_floor_intensity_rna1": (
        "float", "AU",
        "Median per-nucleus RNA1 above-floor cytoplasmic intensity."),
    "mean_cytoplasmic_above_floor_intensity_rna2": (
        "float", "AU",
        "Mean per-nucleus RNA2 above-floor cytoplasmic intensity."),
    "median_cytoplasmic_above_floor_intensity_rna2": (
        "float", "AU",
        "Median per-nucleus RNA2 above-floor cytoplasmic intensity."),
    "mean_nc_ratio_above_floor_intensity_rna1": (
        "float", "—",
        "Mean per-nucleus N/C ratio for RNA1 above-floor intensity."),
    "median_nc_ratio_above_floor_intensity_rna1": (
        "float", "—",
        "Median per-nucleus N/C ratio for RNA1 above-floor intensity."),
    "mean_nc_ratio_above_floor_intensity_rna2": (
        "float", "—",
        "Mean per-nucleus N/C ratio for RNA2 above-floor intensity."),
    "median_nc_ratio_above_floor_intensity_rna2": (
        "float", "—",
        "Median per-nucleus N/C ratio for RNA2 above-floor intensity."),
    "mean_frac_nuclear_above_floor_intensity_rna1": (
        "fraction 0-1", "—",
        "Mean per-nucleus fraction of RNA1 above-floor intensity that is "
        "nuclear."),
    "median_frac_nuclear_above_floor_intensity_rna1": (
        "fraction 0-1", "—",
        "Median per-nucleus fraction of RNA1 above-floor intensity that "
        "is nuclear."),
    "mean_frac_nuclear_above_floor_intensity_rna2": (
        "fraction 0-1", "—",
        "Mean per-nucleus fraction of RNA2 above-floor intensity that is "
        "nuclear."),
    "median_frac_nuclear_above_floor_intensity_rna2": (
        "fraction 0-1", "—",
        "Median per-nucleus fraction of RNA2 above-floor intensity that "
        "is nuclear."),
    "mean_n_nuclear_rna1_rna2_overlap_per_nucleus": (
        "float", "count",
        "Mean per-nucleus count of (nuclear RNA1 spot whose center is within "
        "0.3 um xy of an RNA2 spot center). Pipeline reports the count; "
        "biological interpretation is left to the reader."),
    "median_n_nuclear_rna1_rna2_overlap_per_nucleus": (
        "int", "count",
        "Median per-nucleus count of nuclear RNA1-RNA2 overlap events (0.3 um xy)."),
    "mean_n_cytoplasmic_rna1_spots_per_cell": (
        "float", "count",
        "Mean per-nucleus count of cytoplasmic RNA1 spots. Pipeline reports "
        "the count only — biological interpretation is left to the reader."),
    "median_n_cytoplasmic_rna1_spots_per_cell": (
        "float", "count",
        "Median per-nucleus count of cytoplasmic RNA1 spots."),
    "mean_n_cytoplasmic_rna2_spots_per_cell": (
        "float", "count",
        "Mean per-nucleus count of cytoplasmic RNA2 spots."),
    "median_n_cytoplasmic_rna2_spots_per_cell": (
        "int", "count",
        "Median per-nucleus count of cytoplasmic RNA2 spots."),
    "paired_fraction_rna1_at_0p3um": (
        "fraction 0-1", "—",
        "Fraction of RNA1 spots that have an RNA2 spot center within 0.3 um "
        "(xy distance). Asymmetric: RNA1 spots are the denominator."),
    "paired_fraction_rna2_at_0p3um": (
        "fraction 0-1", "—",
        "Fraction of RNA2 spots that have an RNA1 spot center within 0.3 um. "
        "Asymmetric: RNA2 spots are the denominator."),
    "paired_count_rna1_at_0p3um": (
        "int", "count", "Count of RNA1 spots with a paired RNA2 within 0.3 um."),
    "paired_count_rna2_at_0p3um": (
        "int", "count", "Count of RNA2 spots with a paired RNA1 within 0.3 um."),
    "median_nn_distance_rna1_um": (
        "float", "um",
        "Median nearest-neighbor distance from each RNA1 spot to the next "
        "closest RNA1 spot (xy plane)."),
    "median_nn_distance_rna2_um": (
        "float", "um",
        "Median nearest-neighbor distance for RNA2 spots."),
    "rna_threshold_value": (
        "float", "AU",
        "Per-image pixel-coloc MAD threshold value for RNA1 (background-derived "
        "intensity above which a pixel is considered above-noise). Distinct from "
        "the BigFISH LoG spot detection threshold."),
    "rna2_threshold_value": (
        "float", "AU", "Per-image pixel-coloc MAD threshold for RNA2."),
    "rna_bigfish_log_threshold": (
        "float", "AU",
        "Per-image BigFISH Laplacian-of-Gaussian threshold used for RNA1 spot "
        "detection. Lower = more permissive."),
    "rna2_bigfish_log_threshold": (
        "float", "AU",
        "Per-image BigFISH LoG threshold for RNA2 spot detection."),
    "n_nuclei_border_excluded": (
        "int", "count",
        "Number of nuclei dropped because they touched the image border "
        "within border_margin_px."),
    "total_spots": (
        "int", "count",
        "Total spots detected across both channels in this image."),
    "runtime_s": (
        "float", "seconds",
        "Wall-clock processing time for this single image."),
    "dapi_channel": ("int", "—", "Source channel index for DAPI (e.g. 3 = 405 nm)."),
    "rna_channel": ("int", "—", "Source channel index for RNA1 (e.g. 0 = 640 nm)."),
    "rna2_channel": ("int", "—", "Source channel index for RNA2 (e.g. 1 = 561 nm)."),
    "voxel_xy_nm": ("int", "nm",
                    "Physical pixel size in xy (image metadata)."),
    "voxel_z_nm": ("int", "nm", "Physical z-step (image metadata)."),
    "n_z": ("int", "count",
            "Number of z-slices in the source image (total stack, not just z window)."),
}

PER_NUCLEUS_GLOSSARY: Dict[str, Tuple[str, str, str]] = {
    "image": ("str", "—", "Source image filename. Joins to Per_Image_Summary.image."),
    "condition": ("str", "—", "Biological condition (WT / KO / Sec-Only)."),
    "secondary_only": ("bool", "—", "True if this nucleus came from a sec-only control image."),
    "experiment_id": ("float", "—",
                      "Free-form experiment identifier (NaN if not assigned). "
                      "Reserved for batching multi-experiment runs."),
    "nucleus_id": ("int", "—",
                   "Per-image integer label index. Stable within one image; not unique across images."),
    "nucleus_area_px": ("int", "px^2",
                       "Nucleus mask area in pixels (2D, on the autofocused DAPI plane)."),
    "rna_mean_in_nucleus": ("float", "AU",
                           "Mean RNA1 pixel intensity inside the nucleus mask."),
    "rna_nuclear_mean": ("float", "AU",
                        "Alias for rna_mean_in_nucleus (Fiji-pipeline parity)."),
    "rna_cytoplasmic_mean": ("float", "AU",
                            "Mean RNA1 pixel intensity inside the Voronoi cytoplasm region."),
    "rna_nc_ratio": ("float", "—",
                    "Per-nucleus nuclear/cytoplasmic mean intensity ratio for RNA1."),
    "rna_spot_count": ("int", "count",
                      "Number of RNA1 spots assigned to this nucleus (nuclear + cytoplasmic) "
                      "via nucleus_id."),
    "nuclear_spot_count": ("int", "count",
                          "RNA1 spots inside this nucleus's nuclear mask."),
    "cyto_spot_count": ("int", "count",
                       "RNA1 spots inside this nucleus's Voronoi cytoplasm."),
    "nuclear_spot_fraction": ("fraction 0-1", "—",
                             "Of RNA1 spots assigned to this nucleus, the fraction inside the "
                             "nuclear mask."),
    "nuclear_spot_density_per_um2": ("float", "count/um^2",
                                    "RNA1 nuclear spots / nucleus area in um^2."),
    "rna_spot_mean_intensity_bgc_blend": ("float", "AU",
                                         "Mean RNA1 spot intensity (background-corrected, blend method)."),
    "rna_spot_total_intensity_bgc_blend": ("float", "AU",
                                          "Summed RNA1 spot intensity (background-corrected, blend method)."),
    "rna_spot_median_intensity_bgc_blend": ("float", "AU",
                                           "Median RNA1 spot intensity (background-corrected, blend method)."),
    "rna_spot_mean_intensity_fit": ("float", "AU",
                                   "Mean RNA1 spot intensity (BigFISH 2D Gaussian fit method)."),
    "rna_spot_total_intensity_fit": ("float", "AU",
                                    "Summed RNA1 spot intensity (BigFISH fit method)."),
    "rna_spot_median_intensity_fit": ("float", "AU",
                                     "Median RNA1 spot intensity (fit method)."),
    "rna_spot_intensity_cv_fit": ("float", "—",
                                 "CV of RNA1 spot intensities (fit method) within this nucleus."),
    "sum_rna_intensity": ("int", "AU",
                         "Total RNA1 raw pixel intensity summed over the nuclear mask."),
    "rna2_mean_in_nucleus": ("float", "AU",
                            "Mean RNA2 pixel intensity inside the nucleus mask."),
    "rna2_nuclear_mean": ("float", "AU", "Alias for rna2_mean_in_nucleus."),
    "rna2_cytoplasmic_mean": ("float", "AU",
                             "Mean RNA2 pixel intensity in the Voronoi cytoplasm."),
    "rna2_nc_ratio": ("float", "—",
                    "Per-nucleus nuclear/cytoplasmic mean intensity ratio for RNA2."),
    "n_spots_rna1": ("int", "count", "RNA1 spots assigned to this nucleus (alias of rna_spot_count)."),
    "n_spots_rna2": ("int", "count", "RNA2 spots assigned to this nucleus."),
    "nuclear_spot_count_rna2": ("int", "count",
                               "RNA2 spots inside this nucleus's nuclear mask."),
    "cyto_spot_count_rna2": ("int", "count",
                            "RNA2 spots inside this nucleus's Voronoi cytoplasm."),
    "nuclear_spot_fraction_rna2": ("fraction 0-1", "—",
                                  "Of RNA2 spots assigned to this nucleus, fraction inside the "
                                  "nuclear mask."),
    "nuclear_spot_density_per_um2_rna2": ("float", "count/um^2",
                                         "RNA2 nuclear spots / nucleus area in um^2."),
    "rna2_spot_mean_intensity_fit": ("float", "AU",
                                    "Mean RNA2 spot intensity (BigFISH fit method)."),
    "rna2_spot_total_intensity_fit": ("float", "AU",
                                     "Summed RNA2 spot intensity (fit method)."),
    "rna2_spot_median_intensity_fit": ("float", "AU",
                                      "Median RNA2 spot intensity (fit method)."),
    "rna2_spot_intensity_cv_fit": ("float", "—",
                                  "CV of RNA2 spot intensities (fit method) within this nucleus."),
    "sum_rna2_intensity": ("int", "AU",
                          "Total RNA2 raw pixel intensity summed over the nuclear mask."),
    "median_nn_distance_rna1_um": ("float", "um",
                                  "Median nearest-neighbor distance among the RNA1 spots in this nucleus."),
    "median_nn_distance_rna2_um": ("float", "um",
                                  "Median nearest-neighbor distance among the RNA2 spots in this nucleus."),
    "paired_fraction_rna1_at_0p3um": (
        "fraction 0-1", "—",
        "Per-nucleus paired fraction (RNA1 side): of this nucleus's RNA1 spots, "
        "fraction within 0.3 um xy of any RNA2 spot."),
    "paired_fraction_rna2_at_0p3um": (
        "fraction 0-1", "—",
        "Per-nucleus paired fraction (RNA2 side): of this nucleus's RNA2 spots, "
        "fraction within 0.3 um xy of any RNA1 spot."),
    "paired_spot_count_rna1_at_0p3um": ("int", "count",
                                       "Count of this nucleus's RNA1 spots paired to an RNA2 within 0.3 um."),
    "paired_spot_count_rna2_at_0p3um": ("int", "count",
                                       "Count of this nucleus's RNA2 spots paired to an RNA1 within 0.3 um."),
    "cell_total_intensity_rna1": ("int", "AU",
                                 "Total RNA1 raw intensity over (nucleus + Voronoi cytoplasm)."),
    "cell_total_intensity_rna2": ("int", "AU",
                                 "Total RNA2 raw intensity over (nucleus + Voronoi cytoplasm)."),
    "cell_area_px": ("int", "px^2",
                    "Area in pixels of the full cell footprint (nucleus + Voronoi cytoplasm)."),
    "sum_rna_intensity_cyto": ("int", "AU",
                              "RNA1 raw intensity summed over the Voronoi cytoplasm region only."),
    "sum_rna2_intensity_cyto": ("int", "AU",
                               "RNA2 raw intensity summed over the Voronoi cytoplasm region only."),
    "nuclear_total_intensity_rna1": ("int", "AU",
                                    "RNA1 raw intensity summed over the nuclear mask. "
                                    "Equivalent to sum_rna_intensity (kept for clarity)."),
    "nuclear_total_intensity_rna2": ("int", "AU", "RNA2 raw intensity summed over the nuclear mask."),
    "cytoplasmic_total_intensity_rna1": ("int", "AU",
                                        "Alias for sum_rna_intensity_cyto."),
    "cytoplasmic_total_intensity_rna2": ("int", "AU", "Alias for sum_rna2_intensity_cyto."),
    "nc_ratio_total_intensity_rna1": ("float", "—",
                                     "Per-nucleus N/C ratio computed from total (summed) RNA1 intensity."),
    "nc_ratio_total_intensity_rna2": ("float", "—",
                                     "Per-nucleus N/C ratio computed from total RNA2 intensity."),
    # ---- Above-floor intensity variants (Brian/Sam 2026-05-20) ------------
    # The "above-floor" intensity columns clip pixel values below the
    # publication-image contrast floor (per channel) to zero before summing.
    # Same threshold the eye uses on the rendered pub PNG. Only present when
    # output.apply_pub_contrast_floor_to_analysis is True. NaN when no
    # global floor is available (e.g. pub_contrast_mode = auto_per_image).
    # Floor value per channel is recorded under batch_contrast in
    # run_config.json.
    "nuclear_above_floor_intensity_rna1": (
        "float", "AU",
        "RNA1 raw intensity summed over the nuclear mask, with pixel values "
        "below the pub-image floor (channel rna1) excluded. Floor value is "
        "recorded under batch_contrast in run_config.json. Set "
        "output.apply_pub_contrast_floor_to_analysis to enable this column."),
    "nuclear_above_floor_intensity_rna2": (
        "float", "AU",
        "RNA2 raw intensity summed over the nuclear mask, with pixel values "
        "below the pub-image floor (channel rna2) excluded. Floor value is "
        "recorded under batch_contrast in run_config.json. Set "
        "output.apply_pub_contrast_floor_to_analysis to enable this column."),
    "cytoplasmic_above_floor_intensity_rna1": (
        "float", "AU",
        "RNA1 raw intensity summed over the Voronoi cytoplasm region, with "
        "pixel values below the pub-image floor (channel rna1) excluded. "
        "Floor value recorded under batch_contrast in run_config.json."),
    "cytoplasmic_above_floor_intensity_rna2": (
        "float", "AU",
        "RNA2 raw intensity summed over the Voronoi cytoplasm region, with "
        "pixel values below the pub-image floor (channel rna2) excluded."),
    "nc_ratio_above_floor_intensity_rna1": (
        "float", "—",
        "Per-nucleus N/C ratio for RNA1 above-floor intensity "
        "(nuclear_above_floor / cytoplasmic_above_floor). Same display "
        "floor as the publication PNG."),
    "nc_ratio_above_floor_intensity_rna2": (
        "float", "—",
        "Per-nucleus N/C ratio for RNA2 above-floor intensity."),
    "frac_nuclear_above_floor_intensity_rna1": (
        "fraction 0-1", "—",
        "Per-nucleus fraction of RNA1 above-floor intensity that is nuclear: "
        "nuclear_above_floor / (nuclear_above_floor + cytoplasmic_above_floor)."),
    "frac_nuclear_above_floor_intensity_rna2": (
        "fraction 0-1", "—",
        "Per-nucleus fraction of RNA2 above-floor intensity that is nuclear."),
    "n_nuclear_rna1_rna2_overlap_per_nucleus": (
        "int", "count",
        "Per-nucleus count of nuclear RNA1 spots within 0.3 um xy of an RNA2 spot. "
        "Pipeline reports the count; biological interpretation is the reader's."),
    "n_nuclear_rna2_rna1_overlap_per_nucleus": (
        "int", "count",
        "Per-nucleus count of nuclear RNA2 spots within 0.3 um xy of an RNA1 spot."),
    "n_cytoplasmic_rna1_spots_per_cell": (
        "int", "count",
        "Per-nucleus count of cytoplasmic RNA1 spots. Pipeline reports the count; "
        "biological interpretation is the reader's."),
    "n_cytoplasmic_rna2_spots_per_cell": (
        "int", "count",
        "Per-nucleus count of cytoplasmic RNA2 spots."),
    "cyto_area_px": ("int", "px^2",
                    "Voronoi cytoplasm area in pixels for this nucleus."),
    "cyto_estimation_method": ("str", "—",
                              "Cytoplasm-region estimation method (typically 'voronoi')."),
    "n_voxels": ("int", "px^2",
                "Mask voxel count. For 2D analysis equals nucleus_area_px."),
    "n_pix": ("int", "px^2", "Alias of n_voxels."),
    "n_z_slices": ("int", "count",
                  "Total z-slices in the source stack (not the z-window subset)."),
    "z_mode": ("str", "—",
              "Z-projection mode used to derive the 2D plane "
              "(e.g. 'autofocus', 'max', 'sum')."),
    "z_range": ("str", "—",
               "Z window applied before autofocus, formatted 'start-end' (0-indexed)."),
    "voxel_xy_um": ("float", "um/px", "Physical pixel size, xy."),
    "voxel_z_um": ("float", "um", "Physical z-step."),
    "rna_threshold_value": ("float", "AU",
                           "Per-image RNA1 pixel-coloc MAD threshold (see Per_Image_Summary glossary)."),
    "rna2_threshold_value": ("float", "AU", "Per-image RNA2 pixel-coloc MAD threshold."),
    "rna_frac_above_thr": ("fraction 0-1", "—",
                          "Fraction of nuclear pixels with RNA1 intensity above rna_threshold_value."),
    "rna2_frac_above_thr": ("fraction 0-1", "—",
                           "Fraction of nuclear pixels with RNA2 intensity above rna2_threshold_value."),
    "dapi_mean_in_nucleus": ("float", "AU",
                            "Mean DAPI pixel intensity inside the nuclear mask "
                            "(useful for normalization / DNA content QC)."),
}

PER_SPOT_GLOSSARY: Dict[str, Tuple[str, str, str]] = {
    "image": ("str", "—", "Source image filename. Joins to Per_Image_Summary.image."),
    "condition": ("str", "—", "Biological condition (WT / KO / Sec-Only)."),
    "secondary_only": ("bool", "—", "True if from a sec-only control image."),
    "experiment_id": ("float", "—", "Optional experiment identifier (NaN if not assigned)."),
    "channel": ("str", "—", "Which probe this spot belongs to: 'rna1' or 'rna2'."),
    "spot_id": ("int", "—", "Per-image, per-channel integer spot index."),
    "nucleus_id": ("int", "—",
                  "Nucleus this spot is assigned to (label index). 0 = no nucleus "
                  "containing this spot (free in cytoplasm or outside any cell)."),
    "in_nucleus": ("int (0/1)", "—",
                  "1 if the spot center xy falls inside any nuclear mask, else 0."),
    "in_cytoplasm": ("int (0/1)", "—",
                    "1 if the spot center xy falls inside a Voronoi cytoplasm region, else 0."),
    "x_px": ("int", "px", "Spot center, x pixel coordinate (0-indexed, image frame)."),
    "y_px": ("int", "px", "Spot center, y pixel coordinate."),
    "z_slice": ("int", "—",
               "Z-slice index of the spot center. 0 in autofocus 2D mode."),
    "z_position_um": ("int", "um",
                     "Z-position in physical units. 0 in autofocus 2D mode."),
    "spot_peak_intensity": ("int", "AU",
                           "Raw peak intensity at the BigFISH-fitted spot center."),
    "quality": ("int", "AU",
               "BigFISH spot quality score (currently set equal to spot_peak_intensity)."),
    "spot_fwhm_px": ("int", "px",
                    "Full-width-half-max of the fitted 2D Gaussian, in pixels."),
    "spot_diameter_um": ("float", "um",
                        "Spot diameter in um (FWHM x voxel_xy_um)."),
    "spot_area_px": ("float", "px^2",
                    "Spot area in pixels (pi x (FWHM/2)^2)."),
    "integrated_intensity_fit": ("int", "AU",
                                "BigFISH 2D Gaussian fit integrated intensity for this spot."),
    "nn_distance_um": ("float", "um",
                      "Nearest-neighbor distance from this spot to the next closest spot in the "
                      "SAME channel (xy plane)."),
    "paired_at_0p3um": ("int (0/1)", "—",
                       "1 if this spot has a spot from the OTHER channel within 0.3 um xy, else 0."),
}

CELL_MORPH_GLOSSARY: Dict[str, Tuple[str, str, str]] = {
    "image": ("str", "—", "Source image filename."),
    "condition": ("str", "—", "Biological condition."),
    "experiment_id": ("float", "—", "Optional experiment identifier."),
    "cell_id": ("int", "—", "Per-image integer cell index (matches nucleus_id by construction)."),
    "nucleus_id": ("int", "—", "Per-image integer nucleus index."),
    "segmentation_mode": ("str", "—", "Nucleus segmentation backend used ('stardist' or 'cellpose')."),
    "area_um2": ("float", "um^2", "Nucleus mask area in physical units."),
    "perimeter_um": ("float", "um", "Nucleus perimeter in physical units."),
    "circularity": ("float", "—",
                   "4*pi*area / perimeter^2. 1.0 = perfect circle; lower = more irregular."),
    "aspect_ratio": ("float", "—",
                    "Major axis / minor axis of the fitted ellipse. 1.0 = round."),
    "roundness": ("float", "—",
                 "4*area / (pi*major_axis^2). 1.0 = perfect circle, lower = elongated."),
    "elongation": ("float", "—", "Alias of aspect_ratio."),
    "solidity": ("float", "—",
                "Mask area / convex hull area. 1.0 = convex; lower = lobed/concave."),
    "feret_max_um": ("float", "um",
                    "Maximum Feret diameter (longest caliper distance across the mask)."),
    "feret_min_um": ("float", "um",
                    "Minimum Feret diameter (shortest caliper distance across the mask)."),
}

THRESHOLDS_GLOSSARY: Dict[str, Tuple[str, str, str]] = {
    "image": ("str", "—", "Source image filename."),
    "rna_threshold_used": ("float", "AU",
                          "RNA1 BigFISH LoG threshold actually used for spot detection in this image."),
    "rna_threshold_value": ("float", "AU",
                           "RNA1 pixel-coloc MAD threshold value (batch- or image-scope; see scope)."),
    "rna_threshold_method": ("str", "—",
                            "Method used to derive the pixel-coloc threshold (e.g. 'pixel_coloc_mad')."),
    "rna_threshold_mode": ("str", "—",
                         "Threshold mode for pixel-coloc (e.g. 'mad', 'percentile')."),
    "rna_threshold_k_mad": ("float", "—",
                           "MAD multiplier 'k' used in threshold = median + k*MAD."),
    "rna_threshold_scope": ("str", "—",
                           "Threshold scope: 'batch' (one value used for all images) or 'image' (per-image)."),
    "rna_bigfish_log_threshold": ("float", "AU",
                                 "BigFISH Laplacian-of-Gaussian threshold actually applied for RNA1 spot detection."),
    "rna2_threshold_used": ("float", "AU", "RNA2 BigFISH LoG threshold used."),
    "rna2_threshold_value": ("float", "AU", "RNA2 pixel-coloc MAD threshold value."),
    "rna2_threshold_method": ("str", "—", "RNA2 pixel-coloc threshold method."),
    "rna2_threshold_mode": ("str", "—", "RNA2 pixel-coloc threshold mode."),
    "rna2_threshold_k_mad": ("float", "—", "RNA2 MAD multiplier."),
    "rna2_threshold_scope": ("str", "—", "RNA2 threshold scope ('batch' / 'image')."),
    "rna2_bigfish_log_threshold": ("float", "AU", "BigFISH LoG threshold applied for RNA2 spot detection."),
    "dapi_threshold_method": ("str", "—",
                             "DAPI thresholding method for the nuclei mask "
                             "(e.g. 'Otsu dark', 'Triangle')."),
    "dapi_threshold_value": ("int", "AU", "DAPI mask threshold intensity value."),
    "spot_coloc_pair_distance_um": ("float", "um",
                                   "Maximum xy center-to-center distance used to call RNA1-RNA2 'paired'."),
    "watershed": ("bool", "—",
                 "Whether watershed post-processing was applied to the nucleus segmentation."),
    "nuc_min_area_px": ("int", "px^2",
                       "Minimum nucleus area filter applied after segmentation."),
    "exclude_border_nuclei": ("bool", "—",
                             "Whether nuclei touching the image border were excluded."),
    "z_mode": ("str", "—",
              "Z-projection mode (autofocus / max / sum)."),
    "z_start": ("int", "—", "Inclusive start slice of the z window applied before autofocus."),
    "z_end": ("int", "—", "Inclusive end slice of the z window."),
    "segmentation_backend": ("str", "—", "Nucleus segmentation backend ('stardist' / 'cellpose')."),
    "stardist_prob_threshold": ("float", "—",
                               "StarDist object probability threshold for nucleus calls."),
    "spot_backend": ("str", "—", "Spot detection backend ('bigfish')."),
    "bigfish_spot_radius_nm": ("int", "nm",
                              "Expected spot radius in xy used to size BigFISH's LoG filter."),
    "bigfish_voxel_size_nm": ("int", "nm", "Voxel size in xy passed to BigFISH."),
    "bigfish_voxel_z_nm": ("int", "nm", "Voxel z-step passed to BigFISH."),
    "rna_bigfish_spot_radius_nm": ("int", "nm", "Per-channel xy spot radius for RNA1."),
    "rna2_bigfish_spot_radius_nm": ("int", "nm", "Per-channel xy spot radius for RNA2."),
    "rna_bigfish_spot_radius_z_nm": ("int", "nm", "Per-channel z spot radius for RNA1."),
    "rna2_bigfish_spot_radius_z_nm": ("int", "nm", "Per-channel z spot radius for RNA2."),
    "rna_threshold_multiplier": ("float", "—",
                                "Scaling applied to the pixel-coloc threshold to produce the "
                                "BigFISH LoG threshold for RNA1."),
    "rna2_threshold_multiplier": ("float", "—", "Same multiplier for RNA2."),
    "rna_only_nuclear_spots": ("bool", "—",
                              "If True, only nuclear RNA1 spots are retained."),
    "rna2_only_nuclear_spots": ("bool", "—",
                               "If True, only nuclear RNA2 spots are retained."),
    "rna_min_sep_px": ("int", "px", "Minimum xy separation between RNA1 spots in BigFISH detection."),
    "rna2_min_sep_px": ("int", "px", "Minimum xy separation between RNA2 spots in BigFISH detection."),
    "dapi_label": ("str", "—", "Human-readable label for the DAPI channel."),
    "rna_label": ("str", "—", "Human-readable label for the RNA1 channel."),
    "rna2_label": ("str", "—", "Human-readable label for the RNA2 channel."),
}


GLOSSARIES: Dict[str, Dict[str, Tuple[str, str, str]]] = {
    "Per_Image_Summary": PER_IMAGE_GLOSSARY,
    "Per_Nucleus_Metrics": PER_NUCLEUS_GLOSSARY,
    "Per_Spot_Metrics": PER_SPOT_GLOSSARY,
    "Cell_Morphology": CELL_MORPH_GLOSSARY,
    "Thresholds": THRESHOLDS_GLOSSARY,
}


# ---------------------------------------------------------------------------
# Stats helpers (Mann-Whitney U + Cliff's delta)
# ---------------------------------------------------------------------------

def _mannwhitney_and_cliffs_delta(
    a: Sequence[float], b: Sequence[float],
) -> Tuple[Optional[float], Optional[float]]:
    """Return (Mann-Whitney p-value, Cliff's delta) for samples a vs b.

    Returns (None, None) if either sample is too small or all NaN.
    Uses two-sided MWU. Cliff's delta in [-1, +1]; sign means a-vs-b
    (positive = a tends to be greater than b)."""
    aa = pd.Series(a, dtype="float64").dropna().to_numpy()
    bb = pd.Series(b, dtype="float64").dropna().to_numpy()
    if len(aa) < 2 or len(bb) < 2:
        return None, None
    try:
        from scipy.stats import mannwhitneyu  # type: ignore
        # alternative='two-sided' is the default but make it explicit
        res = mannwhitneyu(aa, bb, alternative="two-sided")
        p = float(res.pvalue)
    except Exception:
        p = None
    # Cliff's delta computed exactly from the MWU statistic relationship:
    #   delta = (2 * U / (n1 * n2)) - 1
    # We compute it manually via outer comparison for clarity / robustness.
    try:
        # vectorized: count(>) and count(<) across all pairs
        gt = (aa[:, None] > bb[None, :]).sum()
        lt = (aa[:, None] < bb[None, :]).sum()
        n_pairs = len(aa) * len(bb)
        delta = float(gt - lt) / float(n_pairs) if n_pairs else None
    except Exception:
        delta = None
    return p, delta


def _mean_sem_n(values: Sequence[float]) -> Tuple[Optional[float], Optional[float], int]:
    s = pd.Series(values, dtype="float64").dropna()
    if not len(s):
        return None, None, 0
    mean = float(s.mean())
    sem = float(s.sem()) if len(s) > 1 else float("nan")
    return mean, sem, int(len(s))


# ---------------------------------------------------------------------------
# Generic helpers
# ---------------------------------------------------------------------------

def _condition_sort_key(condition: Any, sec: Any) -> Tuple[int, str]:
    """Sort order: WT (0) -> KO (1) -> sec-only (2) -> other (3)."""
    if bool(sec):
        return (2, str(condition))
    c = str(condition).upper()
    if c == "WT" or c == "NT":
        return (0, c)
    if c == "KO" or c == "KD":
        return (1, c)
    return (3, c)


def _sort_df_by_condition(df: pd.DataFrame, extra_sort_cols: Sequence[str] = ()) -> pd.DataFrame:
    if not len(df) or "condition" not in df.columns:
        return df.copy()
    work = df.copy()
    sec_col = work["secondary_only"] if "secondary_only" in work.columns else pd.Series(False, index=work.index)
    # 2026-05-19 Brian: remap sec-only rows' condition label to 'sec-only' so
    # the displayed condition matches the gray fill that _condition_fill_hex
    # applies. Previously the cell value still showed the parent subfolder
    # condition (WT/KO) while the fill was gray, which made PI reviewers
    # ask "is this row KO or sec-only?". secondary_only column is preserved
    # untouched for downstream tools that filter on the bool.
    is_sec_mask = sec_col.astype(bool)
    if is_sec_mask.any():
        work.loc[is_sec_mask, "condition"] = "sec-only"
    work["__sort_key"] = [
        _condition_sort_key(c, s) for c, s in zip(work["condition"], sec_col)
    ]
    sort_cols = ["__sort_key"] + [c for c in extra_sort_cols if c in work.columns]
    work = work.sort_values(sort_cols, kind="stable").drop(columns=["__sort_key"])
    return work.reset_index(drop=True)


def _condition_fill_hex(condition: Any, sec: Any) -> Optional[str]:
    """Look up the cell fill for a (condition, sec_only) pair."""
    if bool(sec):
        return CONDITION_FILLS["SEC-ONLY"]
    key = str(condition).strip().upper()
    return CONDITION_FILLS.get(key)


def _numfmt_for(colname: str) -> Optional[str]:
    if colname in INT_COLUMNS:
        return FMT_INT
    if any(colname.startswith(p) for p in FRACTION_COLUMNS_PREFIX):
        return FMT_FRACTION
    if colname.startswith("paired_count"):
        return FMT_INT
    if colname.startswith("paired_fraction"):
        return FMT_FRACTION
    if colname.endswith("_um") or colname.endswith("_um2"):
        return FMT_RATE_2DP
    if colname.startswith("mean_") or colname.startswith("median_") or colname.startswith("cv_"):
        return FMT_FLOAT_GEN
    return None


def _format_data_sheet(
    wb,
    sheet_name: str,
    df: pd.DataFrame,
):
    """Apply uniform formatting to a data sheet:
      * bold + light-gray header with bottom border
      * freeze A2
      * auto-fit column widths (cap 60)
      * numeric format per column
      * condition column color fill
    """
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    if sheet_name not in wb.sheetnames:
        return
    ws = wb[sheet_name]
    BOLD = Font(bold=True, size=11)
    THIN = Side(border_style="thin", color="FF888888")
    HDR_BORDER = Border(bottom=Side(border_style="medium", color="FF333333"))
    HDR_FILL = PatternFill("solid", fgColor=HEADER_FILL_HEX)

    # Header row formatting
    if df.shape[0] >= 0 and df.shape[1] > 0:
        for cell in ws[1]:
            cell.font = BOLD
            cell.fill = HDR_FILL
            cell.border = HDR_BORDER
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
    ws.freeze_panes = "A2"

    # Build {col_letter: col_name} lookup
    col_names: List[str] = []
    for j in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=j).value
        col_names.append(str(v) if v is not None else "")

    # Find the condition column index (1-based)
    cond_idx = None
    sec_idx = None
    for j, name in enumerate(col_names, start=1):
        if name == "condition":
            cond_idx = j
        elif name == "secondary_only":
            sec_idx = j

    # Apply numeric formats per column
    for j, name in enumerate(col_names, start=1):
        fmt = _numfmt_for(name)
        if fmt is None:
            continue
        letter = get_column_letter(j)
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=j)
            if cell.value is None or cell.value == "":
                continue
            cell.number_format = fmt

    # Color the condition column cells
    if cond_idx is not None:
        for row in range(2, ws.max_row + 1):
            cond_v = ws.cell(row=row, column=cond_idx).value
            sec_v = (
                ws.cell(row=row, column=sec_idx).value
                if sec_idx is not None else None
            )
            hex_color = _condition_fill_hex(cond_v, sec_v)
            if hex_color:
                ws.cell(row=row, column=cond_idx).fill = PatternFill(
                    "solid", fgColor=hex_color,
                )

    # Auto-fit widths capped at 60
    for j, name in enumerate(col_names, start=1):
        letter = get_column_letter(j)
        max_len = len(name)
        # Sample first ~500 rows for speed
        sample_rows = min(ws.max_row, 500)
        for row in range(2, sample_rows + 1):
            v = ws.cell(row=row, column=j).value
            if v is None:
                continue
            s = str(v)
            if len(s) > max_len:
                max_len = len(s)
            if max_len >= 60:
                break
        ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 60)


# ---------------------------------------------------------------------------
# Section builders for analysis_summary.xlsx
# ---------------------------------------------------------------------------

def _build_readme(
    ws,
    *,
    fishsuite_version: str,
    run_start_utc: str,
    config_path: Path,
    input_dir: Path,
    output_dir: Path,
    z_mode: str,
    z_start: int,
    z_end: int,
    images: list,
    n_workers: int,
    nuclei_df: pd.DataFrame,
    per_image_df: pd.DataFrame,
    spots_df: pd.DataFrame,
    morph_df: pd.DataFrame,
    thr_df: pd.DataFrame,
    run_cfg_flat: Dict[str, Any],
) -> List[str]:
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    BOLD = Font(bold=True, size=11)
    H1 = Font(bold=True, size=14, color="FFFFFFFF")
    H2 = Font(bold=True, size=12, color="FFFFFFFF")
    SECTION_FILL = PatternFill("solid", fgColor=SECTION_FILL_HEX)
    SUBSECTION_FILL = PatternFill("solid", fgColor=SUBSECTION_FILL_HEX)
    TABLE_HDR_FILL = PatternFill("solid", fgColor=HEADER_FILL_HEX)
    THIN = Side(border_style="thin", color="FF999999")
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    WRAP = Alignment(wrap_text=True, vertical="top")
    LEFT_TOP = Alignment(horizontal="left", vertical="top", wrap_text=True)

    fallback_cols: List[str] = []
    row = 1

    def section_header(title: str) -> None:
        nonlocal row
        c = ws.cell(row=row, column=1, value=title)
        c.font = H1
        c.fill = SECTION_FILL
        for cc in range(2, 5):
            ws.cell(row=row, column=cc).fill = SECTION_FILL
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        ws.cell(row=row, column=1).alignment = LEFT_TOP
        ws.row_dimensions[row].height = 22
        row += 1

    def subsection_header(title: str) -> None:
        nonlocal row
        c = ws.cell(row=row, column=1, value=title)
        c.font = H2
        c.fill = SUBSECTION_FILL
        for cc in range(2, 5):
            ws.cell(row=row, column=cc).fill = SUBSECTION_FILL
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        ws.cell(row=row, column=1).alignment = LEFT_TOP
        row += 1

    def kv(key: str, value: Any) -> None:
        nonlocal row
        c1 = ws.cell(row=row, column=1, value=key)
        c1.font = BOLD
        c2 = ws.cell(row=row, column=2, value=str(value))
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        c2.alignment = LEFT_TOP
        row += 1

    def table_header(headers: Iterable[str]) -> None:
        nonlocal row
        for i, h in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=i, value=h)
            cell.font = BOLD
            cell.fill = TABLE_HDR_FILL
            cell.border = BORDER
            cell.alignment = LEFT_TOP
        row += 1

    def table_row(values: Iterable[Any]) -> None:
        nonlocal row
        for i, v in enumerate(values, start=1):
            cell = ws.cell(row=row, column=i, value=v)
            cell.border = BORDER
            cell.alignment = WRAP
        row += 1

    def blank() -> None:
        nonlocal row
        row += 1

    # ── A. RUN HEADER ────────────────────────────────────────────────────
    section_header("A. RUN HEADER")
    kv("Title", "fishsuite RNA-FISH analysis report")
    # 2026-05-19 Brian: build subtitle from the actual config so the labels
    # the user chose (e.g. "Exons" / "Introns") propagate, and we don't
    # hardcode dye names (Cy5/Cy3) that may not match the actual probes.
    _ch = run_cfg_flat
    _rna_lbl = str(_ch.get("config_resolved.channels.rna_label", "RNA1"))
    _rna2_lbl = str(_ch.get("config_resolved.channels.rna2_label", "RNA2"))
    _dapi_lbl = str(_ch.get("config_resolved.channels.dapi_label", "DAPI"))
    _ch_rna = _ch.get("config_resolved.channels.rna", 0)
    _ch_rna2 = _ch.get("config_resolved.channels.rna2", 1)
    _ch_dapi = _ch.get("config_resolved.channels.dapi", 3)
    kv(
        "Subtitle",
        f"Two-channel FISH: ch{_ch_rna} (647 nm) = {_rna_lbl}; "
        f"ch{_ch_rna2} (561 nm) = {_rna2_lbl}; "
        f"ch{_ch_dapi} (405 nm) = {_dapi_lbl}.",
    )
    kv("Run start (UTC)", run_start_utc)
    kv("fishsuite version", fishsuite_version)
    kv("Config file", str(config_path))
    kv("Input directory", str(input_dir))
    kv("Output directory", str(output_dir))
    kv("Z window (start-end, inclusive)", f"{z_start}-{z_end}")
    kv("Z mode", z_mode)
    kv("Parallel workers", n_workers)

    n_images = len(images)
    n_real = sum(1 for im in images if not im.sec_only)
    n_sec = sum(1 for im in images if im.sec_only)
    by_cond: Dict[str, int] = {}
    for im in images:
        key = f"{im.condition}{' (sec-only)' if im.sec_only else ''}"
        by_cond[key] = by_cond.get(key, 0) + 1
    cond_str = ", ".join(f"{k}: {v}" for k, v in by_cond.items()) or "—"
    kv("Number of images", f"{n_images} ({n_real} real, {n_sec} sec-only)")
    kv("Breakdown by condition", cond_str)
    if len(nuclei_df) and "condition" in nuclei_df.columns:
        nuc_counts = nuclei_df.groupby("condition").size().to_dict()
        nuc_str = ", ".join(f"{k}: {v}" for k, v in nuc_counts.items())
    else:
        nuc_str = "0"
    kv("Nuclei segmented per condition (after filters)", nuc_str)
    blank()

    # ── B. SHEET INDEX ───────────────────────────────────────────────────
    section_header("B. SHEET INDEX")
    sheet_descriptions = [
        ("README",
         "This sheet. Run header, sheet index, methods, glossary, interpretation, "
         "figure index."),
        ("Executive_Summary",
         "One-page plain-language summary the PI reads first. Total counts, per-condition "
         "averages, sec-only control check, descriptive paragraph."),
        ("PI_Focus",
         "Curated PI deliverable — every metric the PI cares about (spot counts, above-floor "
         "intensity, spot brightness, spot size) compared WT vs KO across nuclear and "
         "cytoplasmic compartments. One-stop glance."),
        ("Comparison_Table",
         "WT vs KO side-by-side comparison for ~15 high-priority metrics, with "
         "Mann-Whitney U p-values and Cliff's delta effect sizes (computed per-nucleus)."),
        ("Per_Image_Summary",
         "One row per image. Aggregate spot counts, nuclear fractions, "
         "intensities, paired (overlap) fractions, thresholds, provenance per FOV."),
        ("Per_Nucleus_Metrics",
         "One row per nucleus. Per-cell spot counts (RNA1 / RNA2, nuclear / cytoplasmic), "
         "per-cell intensities, paired counts, nearest-neighbor distances."),
        ("Per_Spot_Metrics",
         "One row per detected FISH spot. xy positions, intensities, sizes, "
         "nucleus assignment, in_nucleus / in_cytoplasm / paired flags."),
        ("Cell_Morphology",
         "One row per nucleus. Nuclear morphology (area, perimeter, "
         "circularity, Feret diameters, solidity)."),
        ("Thresholds",
         "Per-image detection thresholds: pixel-coloc MAD thresholds, "
         "BigFISH LoG thresholds, segmentation parameters, channel labels, z window."),
        ("Run_Config",
         "Flattened run_config.json: every config parameter as a key | value row, "
         "with critical params (contrast mode, z window, BigFISH multipliers) bolded."),
    ]
    table_header(["Sheet", "Description"])
    for name, desc in sheet_descriptions:
        c1 = ws.cell(row=row, column=1, value=name)
        c1.font = BOLD
        c1.border = BORDER
        c1.alignment = WRAP
        c2 = ws.cell(row=row, column=2, value=desc)
        c2.alignment = WRAP
        c2.border = BORDER
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        row += 1
    blank()

    # ── C. METHODS SUMMARY ───────────────────────────────────────────────
    section_header("C. METHODS SUMMARY")
    # Pull a few notable items from the flattened config for visibility here
    contrast_mode = run_cfg_flat.get("PUB_CONTRAST_MODE", run_cfg_flat.get("pub_contrast_mode", "(not set)"))
    seg_backend = run_cfg_flat.get("SEGMENTATION_BACKEND", run_cfg_flat.get("segmentation_backend", "stardist"))
    spot_backend = run_cfg_flat.get("SPOT_BACKEND", run_cfg_flat.get("spot_backend", "bigfish"))
    rna_mult = run_cfg_flat.get("rna_threshold_multiplier",
                                run_cfg_flat.get("foci.rna_threshold_multiplier", "(see Thresholds)"))
    rna2_mult = run_cfg_flat.get("rna2_threshold_multiplier",
                                 run_cfg_flat.get("foci.rna2_threshold_multiplier", "(see Thresholds)"))
    methods_rows = [
        ("Nucleus segmentation",
         f"{seg_backend} on the DAPI channel; min-area + border-exclusion filters applied."),
        ("Z handling",
         f"Z mode='{z_mode}'; window slices [{z_start}-{z_end}] (inclusive). "
         "Autofocus picks DAPI's best in-focus plane within the window; all other channels "
         "are extracted at that same z plane so masks + spots are spatially consistent."),
        ("Spot detection",
         f"Backend='{spot_backend}'. BigFISH Laplacian-of-Gaussian (LoG) per channel; "
         "per-image LoG thresholds derived by scaling the pixel-coloc MAD threshold by "
         f"the per-channel multiplier (RNA1={rna_mult}, RNA2={rna2_mult}). See Thresholds sheet."),
        ("Pixel-coloc threshold scope",
         f"Default scope: batch. One MAD-based threshold per channel applied across all images "
         "for cross-image comparability. BigFISH LoG spot detection stays per-image."),
        ("Spot pairing (overlap)",
         "RNA1 spot and RNA2 spot are 'paired' if their xy center-to-center distance is "
         "<= 0.3 um. Asymmetric definition: paired_fraction_rna1 uses RNA1 spots as the "
         "denominator; paired_fraction_rna2 uses RNA2 spots."),
        ("Publication-image contrast",
         f"Mode='{contrast_mode}'. PNGs in publication_images/ use one uniform (min, max) per "
         "channel across the whole batch when contrast_mode='auto_batch'. See run_config.json -> "
         "batch_contrast for exact floor/ceil values."),
        ("Spot detection floor",
         "When apply_pub_contrast_floor_to_spots = True, spots with peak intensity below "
         "the channel's resolved contrast floor (see `batch_contrast` block) are dropped after "
         "BigFISH LoG detection. The filter runs BEFORE per-nucleus stratification + pairing, so "
         "all downstream counts (nuclear / cytoplasmic, paired_fraction, NN distances) reflect "
         "the filtered spot set. BigFISH LoG detection itself is unchanged. Independent of the "
         "apply_pub_contrast_floor_to_analysis flag (which only affects pixel-intensity "
         "quantification columns, not spot counts)."),
        ("Statistical comparison",
         "WT vs KO computed per-nucleus (larger n than per-image) with two-sided Mann-Whitney "
         "U test. Cliff's delta reported as a nonparametric effect size in [-1, +1]; "
         "magnitude > 0.474 typically called 'large', 0.33-0.474 'medium', 0.147-0.33 'small'. "
         "Sec-only excluded from significance testing."),
        ("Sec-only handling",
         "Files flagged secondary_only=True have no primary FISH probe; expected ~0 detected "
         "RNA1 / RNA2 spots. Reported for completeness; not included in WT vs KO statistics. "
         "Cells where 0 spots would make a metric undefined (e.g. frac_nuclear with 0 spots) are "
         "left blank rather than written as 0 or NaN."),
        ("Intensities are arbitrary units (AU)",
         "Intensity columns are raw 16-bit pixel values from the source microscope file. No "
         "absolute calibration is applied; compare within a run, not across different sessions."),
        ("Counts are pipeline outputs only",
         "Columns named n_nuclear_rna1_rna2_overlap_per_nucleus, n_cytoplasmic_*_spots_per_cell, etc. are reported as "
         "the literal counts the pipeline measured (nuclear RNA1<->RNA2 overlap counts, "
         "cytoplasmic RNA1 counts). Biological interpretation is left to the reader."),
    ]
    table_header(["Topic", "Notes"])
    for topic, note in methods_rows:
        c1 = ws.cell(row=row, column=1, value=topic)
        c1.font = BOLD
        c1.border = BORDER
        c1.alignment = WRAP
        c2 = ws.cell(row=row, column=2, value=note)
        c2.border = BORDER
        c2.alignment = WRAP
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        row += 1
    blank()

    # ── D. GLOSSARY (split by sheet) ─────────────────────────────────────
    section_header("D. GLOSSARY (column-by-column, grouped by sheet)")
    glossary_targets = [
        ("D1. Per_Image_Summary", "Per_Image_Summary", per_image_df),
        ("D2. Per_Nucleus_Metrics", "Per_Nucleus_Metrics", nuclei_df),
        ("D3. Per_Spot_Metrics", "Per_Spot_Metrics", spots_df),
        ("D4. Cell_Morphology", "Cell_Morphology", morph_df),
        ("D5. Thresholds", "Thresholds", thr_df),
    ]
    for sub_title, gloss_key, df in glossary_targets:
        if not len(df):
            continue
        subsection_header(sub_title)
        gloss = GLOSSARIES.get(gloss_key, {})
        table_header(["Column", "Sheet", "Type", "Units", "Description"])
        # Stretch: 5 columns this time. Adjust merged width below.
        for col in df.columns:
            entry = gloss.get(col)
            if entry is None:
                fallback_cols.append(f"{gloss_key}:{col}")
                t, u, d = ("?", "?", "(no description — see column name)")
            else:
                t, u, d = entry
            r1 = ws.cell(row=row, column=1, value=col)
            r1.border = BORDER
            r1.alignment = WRAP
            r2 = ws.cell(row=row, column=2, value=gloss_key)
            r2.border = BORDER
            r2.alignment = WRAP
            r3 = ws.cell(row=row, column=3, value=t)
            r3.border = BORDER
            r3.alignment = WRAP
            r4 = ws.cell(row=row, column=4, value=u)
            r4.border = BORDER
            r4.alignment = WRAP
            # We have 4 visible columns A-D. Type+Units share column 3+4. The
            # description goes in column 4 as well — overwrite with the longer
            # text in column 4 (units already shown), or merge column 5 into 4.
            # Simpler: collapse to 4 columns by putting type/units inline in
            # the description.
            r4.value = f"{u}    |    {d}" if u not in ("—", "?") else d
            row += 1
        blank()

    # ── E. INTERPRETATION GUIDE ──────────────────────────────────────────
    section_header("E. INTERPRETATION GUIDE")
    interp_rows = [
        ("Sec-only controls",
         "Sec-only rows are useful as a noise floor. Expected: ~0 RNA1 / RNA2 spots; real DAPI; "
         "real autofluorescence. Real cells, no probe. Anything > 0 is a detector hallucination."),
        ("Direction of WT-vs-KO statistics",
         "Cliff's delta is computed as (P(WT > KO) - P(WT < KO)). Positive delta means WT values "
         "tend to be greater than KO values for that metric. Magnitude rules of thumb: |delta| "
         "< 0.147 negligible, 0.147-0.33 small, 0.33-0.474 medium, > 0.474 large."),
        ("Overlap / paired definition",
         "An RNA1 spot and an RNA2 spot are 'paired' if their xy center-to-center distance "
         "is <= 0.3 um. Asymmetric: paired_fraction_rna1 and paired_fraction_rna2 can differ "
         "because the denominators differ. Pairing is xy-only (autofocus 2D, all spots coplanar)."),
        ("Counts vs interpretation",
         "Column names like n_nuclear_rna1_rna2_overlap_per_nucleus, n_cytoplasmic_rna1_spots_per_cell are the "
         "literal counts the pipeline measured (nuclear RNA1-RNA2 overlap count, cytoplasmic "
         "RNA1 count). They are NOT biological calls. The neutral wording in Comparison_Table "
         "('mean nuclear RNA1-RNA2 overlap per nucleus', 'mean cytoplasmic RNA1 spots per cell') "
         "states what was measured without claiming what it means."),
        ("Where to look next",
         "Executive_Summary -> PI_Focus -> Comparison_Table for the headline result, then the publication "
         "figures (figures/97_CORE_overview_panel.png and similar) for the same numbers "
         "visualized, then Per_Image_Summary if you want to inspect FOV-by-FOV variability."),
        ("Above-floor intensity",
         "Above-floor intensity = sum of (pixel_value - channel_floor) over the masked region, "
         "clipping at zero. Quantifies signal above the user-set noise threshold (matches what "
         "is visible in publication images). See batch_contrast block in run_config.json for "
         "the per-channel floor values."),
    ]
    table_header(["Topic", "Notes"])
    for topic, note in interp_rows:
        c1 = ws.cell(row=row, column=1, value=topic)
        c1.font = BOLD
        c1.border = BORDER
        c1.alignment = WRAP
        c2 = ws.cell(row=row, column=2, value=note)
        c2.border = BORDER
        c2.alignment = WRAP
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        row += 1
    blank()

    # ── F. WHERE TO FIND FIGURES ─────────────────────────────────────────
    section_header("F. WHERE TO FIND THE FIGURES")
    kv("Figures directory", str(Path(output_dir) / "figures"))
    blank()
    figs = [
        ("Headline overview panel (curated 6-figure summary)",
         "figures/97_CORE_overview_panel.png"),
        ("Overlap / colocalization overview panel",
         "figures/98_COLOC_overview_panel.png"),
        ("Composition: RNA1 vs RNA2 localization across conditions",
         "figures/57_localization_composition_both_channels.png"),
        ("Per-channel spot counts per cell — RNA1 by condition",
         "figures/17_box_spots_per_cell_rna1_by_condition.png"),
        ("Per-channel spot counts per cell — RNA2 by condition",
         "figures/18_box_spots_per_cell_rna2_by_condition.png"),
        ("Spot localization composition stacked bars",
         "figures/52, 53, 54, 55"),
        ("Per-cell nuclear-fraction boxplots",
         "figures/15b (RNA1), 16b (RNA2)"),
        ("Per-image QC overlays",
         "qc_overlays/<image>__qc_*.png"),
        ("Per-nucleus popouts",
         "nuclei_popouts/<image>__nuc*.png"),
        ("Pipeline walkthrough panels",
         "pipeline_walkthrough/<image>__walkthrough_*.png"),
    ]
    table_header(["Concept", "Path (relative to output dir)"])
    for concept, path in figs:
        c1 = ws.cell(row=row, column=1, value=concept)
        c1.border = BORDER
        c1.alignment = WRAP
        c2 = ws.cell(row=row, column=2, value=path)
        c2.border = BORDER
        c2.alignment = WRAP
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        row += 1

    # Column widths + freeze
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 100
    ws.freeze_panes = "A2"

    return fallback_cols


def _build_executive_summary(
    ws,
    *,
    per_image_df: pd.DataFrame,
    nuclei_df: pd.DataFrame,
    spots_df: pd.DataFrame,
) -> None:
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    BOLD = Font(bold=True, size=11)
    H1 = Font(bold=True, size=14, color="FFFFFFFF")
    SECTION_FILL = PatternFill("solid", fgColor=SECTION_FILL_HEX)
    TABLE_HDR_FILL = PatternFill("solid", fgColor=HEADER_FILL_HEX)
    THIN = Side(border_style="thin", color="FF999999")
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    WRAP = Alignment(wrap_text=True, vertical="top")
    LEFT_TOP = Alignment(horizontal="left", vertical="top", wrap_text=True)

    row = 1

    def section_header(title: str) -> None:
        nonlocal row
        c = ws.cell(row=row, column=1, value=title)
        c.font = H1
        c.fill = SECTION_FILL
        for cc in range(2, 7):
            ws.cell(row=row, column=cc).fill = SECTION_FILL
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        ws.cell(row=row, column=1).alignment = LEFT_TOP
        ws.row_dimensions[row].height = 22
        row += 1

    def table_header(headers: Iterable[str]) -> None:
        nonlocal row
        for i, h in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=i, value=h)
            cell.font = BOLD
            cell.fill = TABLE_HDR_FILL
            cell.border = BORDER
            cell.alignment = LEFT_TOP
        row += 1

    def table_row(values: Iterable[Any]) -> None:
        nonlocal row
        for i, v in enumerate(values, start=1):
            cell = ws.cell(row=row, column=i, value=v)
            cell.border = BORDER
            cell.alignment = WRAP
        row += 1

    def blank() -> None:
        nonlocal row
        row += 1

    # Build ordered condition list: WT, KO, sec-only at the end
    conds_real: List[str] = []
    conds_sec: List[str] = []
    if len(per_image_df) and "condition" in per_image_df.columns:
        for cond in per_image_df["condition"].dropna().unique():
            is_sec = bool(
                per_image_df.loc[per_image_df["condition"] == cond, "secondary_only"].any()
                if "secondary_only" in per_image_df.columns else False
            )
            # If a condition has mixed sec/real, treat real first
            has_real = bool(
                ~per_image_df.loc[per_image_df["condition"] == cond, "secondary_only"].all()
                if "secondary_only" in per_image_df.columns else True
            )
            if has_real and cond not in conds_real:
                conds_real.append(cond)
        # sec-only entries
        if "secondary_only" in per_image_df.columns:
            sec_imgs = per_image_df.loc[per_image_df["secondary_only"], :]
            for cond in sec_imgs["condition"].dropna().unique():
                label = f"sec-only ({cond})"
                if label not in conds_sec:
                    conds_sec.append(label)

    # Sort real conds: WT, NT first, then KO/KD, then others
    def _real_rank(c: str) -> int:
        u = str(c).upper()
        if u in ("WT", "NT"):
            return 0
        if u in ("KO", "KD"):
            return 1
        return 2
    conds_real.sort(key=_real_rank)

    all_conds = conds_real + conds_sec

    # ── Top-line numbers ─────────────────────────────────────────────────
    section_header("EXECUTIVE SUMMARY")
    ws.cell(row=row, column=1, value="Total images analyzed").font = BOLD
    ws.cell(row=row, column=2, value=int(len(per_image_df)))
    row += 1
    if len(nuclei_df) and "condition" in nuclei_df.columns:
        total_nuc = int(len(nuclei_df))
    else:
        total_nuc = int(per_image_df["nuclei_analyzed"].sum()) if "nuclei_analyzed" in per_image_df.columns else 0
    ws.cell(row=row, column=1, value="Total nuclei (after filters)").font = BOLD
    ws.cell(row=row, column=2, value=total_nuc)
    row += 1
    if len(spots_df) and "channel" in spots_df.columns:
        n_rna1 = int((spots_df["channel"] == "rna1").sum())
        n_rna2 = int((spots_df["channel"] == "rna2").sum())
    else:
        n_rna1 = int(per_image_df.get("total_spots_rna1", pd.Series(dtype=float)).sum())
        n_rna2 = int(per_image_df.get("total_spots_rna2", pd.Series(dtype=float)).sum())
    ws.cell(row=row, column=1, value="Total RNA1 spots detected").font = BOLD
    ws.cell(row=row, column=2, value=n_rna1)
    row += 1
    ws.cell(row=row, column=1, value="Total RNA2 spots detected").font = BOLD
    ws.cell(row=row, column=2, value=n_rna2)
    row += 1
    blank()

    # ── Per-condition table ──────────────────────────────────────────────
    table_header([
        "Condition", "n images", "n nuclei",
        "mean RNA1 spots/nuc", "mean RNA2 spots/nuc",
        "mean nuc-frac RNA1",
    ])
    for cond_label in all_conds:
        if cond_label.startswith("sec-only ("):
            base_cond = cond_label[len("sec-only ("):-1]
            mask_img = (per_image_df["condition"] == base_cond) & per_image_df.get("secondary_only", False)
            mask_nuc = (
                (nuclei_df["condition"] == base_cond)
                & nuclei_df.get("secondary_only", pd.Series(False, index=nuclei_df.index))
            ) if len(nuclei_df) else pd.Series(dtype=bool)
        else:
            mask_img = (per_image_df["condition"] == cond_label) & (~per_image_df.get(
                "secondary_only", pd.Series(False, index=per_image_df.index)
            ))
            mask_nuc = (
                (nuclei_df["condition"] == cond_label)
                & (~nuclei_df.get("secondary_only", pd.Series(False, index=nuclei_df.index)))
            ) if len(nuclei_df) else pd.Series(dtype=bool)
        n_img = int(mask_img.sum()) if len(per_image_df) else 0
        n_nuc = int(mask_nuc.sum()) if len(nuclei_df) else 0
        # 2026-05-27: column resolution is mode-dependent. rna_rna mode emits
        # *_rna1 / *_rna2 suffixed columns (per-nucleus n_spots_rna1 and
        # per-image mean_spots_per_nucleus_rna1). rna_only mode emits the
        # UNSUFFIXED rna_spot_count (per-nucleus) and mean_spots_per_nucleus
        # (per-image), with no RNA2 channel. The previous code unconditionally
        # fell back to per_image_df["mean_spots_per_nucleus_rna1"] when
        # "n_spots_rna1" was absent, which raised KeyError in rna_only mode
        # (the non-fatal crash at the end of the CPU run). Resolve each value
        # against whichever column exists, defaulting to NaN otherwise so
        # neither mode can KeyError.
        def _mean_nuc(col: str) -> float:
            if len(nuclei_df) and col in nuclei_df.columns and n_nuc:
                return float(nuclei_df.loc[mask_nuc, col].mean())
            return float("nan")

        def _mean_img(col: str) -> float:
            if len(per_image_df) and col in per_image_df.columns and n_img:
                return float(per_image_df.loc[mask_img, col].mean())
            return float("nan")

        if len(nuclei_df) and "n_spots_rna1" in nuclei_df.columns:
            # rna_rna (two-channel) per-nucleus columns.
            m_r1 = _mean_nuc("n_spots_rna1")
            m_r2 = _mean_nuc("n_spots_rna2")
        elif len(nuclei_df) and "rna_spot_count" in nuclei_df.columns:
            # rna_only per-nucleus column (single channel; no RNA2).
            m_r1 = _mean_nuc("rna_spot_count")
            m_r2 = float("nan")
        else:
            # No per-nucleus table — fall back to per-image means (handles
            # both suffixed and unsuffixed schemas without KeyError).
            m_r1 = _mean_img("mean_spots_per_nucleus_rna1")
            if math.isnan(m_r1):
                m_r1 = _mean_img("mean_spots_per_nucleus")
            m_r2 = _mean_img("mean_spots_per_nucleus_rna2")
        # Nuclear fraction: per-nucleus nuclear_spot_fraction exists in BOTH
        # modes (rna_only + rna_rna). Fall back to the per-image suffixed
        # column only if the per-nucleus column is missing.
        if len(nuclei_df) and "nuclear_spot_fraction" in nuclei_df.columns:
            frac_r1 = _mean_nuc("nuclear_spot_fraction")
        else:
            frac_r1 = _mean_img("frac_nuclear_rna1")

        def _fmt_f(x: float) -> Any:
            return "" if math.isnan(x) else round(x, 4)

        table_row([
            cond_label, n_img, n_nuc,
            _fmt_f(m_r1), _fmt_f(m_r2), _fmt_f(frac_r1),
        ])
    blank()

    # ── Sec-only control check ───────────────────────────────────────────
    ws.cell(row=row, column=1, value="Sec-only control check").font = Font(bold=True, size=12)
    row += 1
    if "secondary_only" in per_image_df.columns:
        sec = per_image_df.loc[per_image_df["secondary_only"], :]
        if len(sec):
            total_sec_r1 = int(sec.get("total_spots_rna1", pd.Series([0]*len(sec))).sum())
            total_sec_r2 = int(sec.get("total_spots_rna2", pd.Series([0]*len(sec))).sum())
            status = "PASS (0 spots in both channels)" if total_sec_r1 == 0 and total_sec_r2 == 0 \
                else f"REVIEW: sec-only spots detected (RNA1={total_sec_r1}, RNA2={total_sec_r2})"
            ws.cell(row=row, column=1, value=f"Status: {status}").alignment = WRAP
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
            row += 1
            ws.cell(row=row, column=1, value=
                f"Sec-only images: {len(sec)} | total RNA1 spots: {total_sec_r1} | "
                f"total RNA2 spots: {total_sec_r2}").alignment = WRAP
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
            row += 1
        else:
            ws.cell(row=row, column=1, value="No sec-only control images in this run.").alignment = WRAP
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
            row += 1
    blank()

    # ── Descriptive paragraph (FACTS ONLY) ───────────────────────────────
    ws.cell(row=row, column=1, value="Descriptive summary (facts only — interpretation left to the reader)").font = Font(bold=True, size=12)
    row += 1
    paragraph = _build_descriptive_paragraph(per_image_df, nuclei_df, conds_real)
    p_cell = ws.cell(row=row, column=1, value=paragraph)
    p_cell.alignment = WRAP
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    ws.row_dimensions[row].height = 90
    row += 1

    # Column widths
    from openpyxl.utils import get_column_letter
    for j, w in enumerate([28, 12, 12, 22, 22, 22], start=1):
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = "A2"


def _build_descriptive_paragraph(
    per_image_df: pd.DataFrame,
    nuclei_df: pd.DataFrame,
    conds_real: List[str],
) -> str:
    """Plain-language facts-only paragraph for the Executive_Summary."""
    if not conds_real or not len(nuclei_df):
        return "No data."

    parts: List[str] = []
    for cond in conds_real:
        if "secondary_only" in nuclei_df.columns:
            mask = (nuclei_df["condition"] == cond) & (~nuclei_df["secondary_only"])
        else:
            mask = (nuclei_df["condition"] == cond)
        n_nuc = int(mask.sum())
        if not n_nuc:
            continue
        m_r1 = float(nuclei_df.loc[mask, "n_spots_rna1"].mean()) if "n_spots_rna1" in nuclei_df.columns else float("nan")
        m_r2 = float(nuclei_df.loc[mask, "n_spots_rna2"].mean()) if "n_spots_rna2" in nuclei_df.columns else float("nan")
        parts.append(
            f"{cond} (n={n_nuc} nuclei): mean RNA1 spots/nucleus = "
            f"{m_r1:.2f}; mean RNA2 spots/nucleus = {m_r2:.2f}."
        )
    return (
        "Descriptive measurements for this run: " + " ".join(parts) +
        " See Comparison_Table for WT vs KO statistical comparison "
        "(Mann-Whitney U + Cliff's delta) and Per_Nucleus_Metrics for the underlying "
        "per-cell values. Biological interpretation is the reader's."
    )


# ---------------------------------------------------------------------------
# Comparison table
# ---------------------------------------------------------------------------

# (metric_label, per_image column, per_nucleus column for stats, category)
# per_image column is used for mean/SEM display; per_nucleus column for MWU.
# If per_nucleus column is None, stats are skipped (image-only metric).
COMPARISON_METRICS: List[Tuple[str, str, Optional[str], str]] = [
    ("Nuclei analyzed (sum)", "nuclei_analyzed", None, "Counts"),
    ("Mean RNA1 spots per nucleus", "mean_spots_per_nucleus_rna1", "n_spots_rna1", "Counts"),
    ("Mean RNA2 spots per nucleus", "mean_spots_per_nucleus_rna2", "n_spots_rna2", "Counts"),
    ("Mean nuclear fraction RNA1", "frac_nuclear_rna1", "nuclear_spot_fraction", "Localization fractions"),
    ("Mean nuclear fraction RNA2", "frac_nuclear_rna2", "nuclear_spot_fraction_rna2", "Localization fractions"),
    ("Paired fraction RNA1 (within 0.3 um)", "paired_fraction_rna1_at_0p3um",
     "paired_fraction_rna1_at_0p3um", "Overlap fractions"),
    ("Paired fraction RNA2 (within 0.3 um)", "paired_fraction_rna2_at_0p3um",
     "paired_fraction_rna2_at_0p3um", "Overlap fractions"),
    ("Mean nuclear RNA1<->RNA2 overlap count per nucleus", "mean_n_nuclear_rna1_rna2_overlap_per_nucleus",
     "n_nuclear_rna1_rna2_overlap_per_nucleus", "Overlap fractions"),
    ("Mean cytoplasmic RNA1 spots per cell", "mean_n_cytoplasmic_rna1_spots_per_cell",
     "n_cytoplasmic_rna1_spots_per_cell", "Counts"),
    ("Mean cytoplasmic RNA2 spots per cell", "mean_n_cytoplasmic_rna2_spots_per_cell",
     "n_cytoplasmic_rna2_spots_per_cell", "Counts"),
    ("Mean RNA1 per-cell summed spot intensity (BigFISH fit)",
     "mean_cell_total_spot_intensity_fit_rna1", "rna_spot_total_intensity_fit", "Intensities"),
    ("Mean RNA2 per-cell summed spot intensity (BigFISH fit)",
     "mean_cell_total_spot_intensity_fit_rna2", "rna2_spot_total_intensity_fit", "Intensities"),
    ("Mean N/C ratio RNA1 (raw intensity)", "mean_nc_ratio_total_intensity_rna1",
     "nc_ratio_total_intensity_rna1", "Intensities"),
    ("Mean N/C ratio RNA2 (raw intensity)", "mean_nc_ratio_total_intensity_rna2",
     "nc_ratio_total_intensity_rna2", "Intensities"),
    # ---- Above-floor intensity metrics (Brian/Sam 2026-05-20) -----------
    # These rows only render when the per-image / per-nucleus tables actually
    # contain the column — _build_comparison_table skips metrics whose
    # per_image column is absent. So when apply_pub_contrast_floor_to_analysis
    # is False, these rows are silently dropped from the table.
    ("Mean nuclear above-floor intensity RNA1",
     "mean_nuclear_above_floor_intensity_rna1",
     "nuclear_above_floor_intensity_rna1", "Intensities"),
    ("Mean nuclear above-floor intensity RNA2",
     "mean_nuclear_above_floor_intensity_rna2",
     "nuclear_above_floor_intensity_rna2", "Intensities"),
    ("Mean cytoplasmic above-floor intensity RNA1",
     "mean_cytoplasmic_above_floor_intensity_rna1",
     "cytoplasmic_above_floor_intensity_rna1", "Intensities"),
    ("Mean cytoplasmic above-floor intensity RNA2",
     "mean_cytoplasmic_above_floor_intensity_rna2",
     "cytoplasmic_above_floor_intensity_rna2", "Intensities"),
    ("Mean N/C ratio RNA1 (above-floor intensity)",
     "mean_nc_ratio_above_floor_intensity_rna1",
     "nc_ratio_above_floor_intensity_rna1", "Intensities"),
    ("Mean N/C ratio RNA2 (above-floor intensity)",
     "mean_nc_ratio_above_floor_intensity_rna2",
     "nc_ratio_above_floor_intensity_rna2", "Intensities"),
    ("Mean nuclear fraction RNA1 (above-floor intensity)",
     "mean_frac_nuclear_above_floor_intensity_rna1",
     "frac_nuclear_above_floor_intensity_rna1", "Localization fractions"),
    ("Mean nuclear fraction RNA2 (above-floor intensity)",
     "mean_frac_nuclear_above_floor_intensity_rna2",
     "frac_nuclear_above_floor_intensity_rna2", "Localization fractions"),
    ("Median RNA1 spot NN distance (xy)", "median_nn_distance_rna1_um",
     "median_nn_distance_rna1_um", "Sizes"),
    ("Median RNA2 spot NN distance (xy)", "median_nn_distance_rna2_um",
     "median_nn_distance_rna2_um", "Sizes"),
]

CATEGORY_FILLS: Dict[str, str] = {
    "Counts":                  "FFE8EFF7",
    "Localization fractions":  "FFE7F4E0",
    "Overlap fractions":       "FFFFF4D1",
    "Intensities":             "FFF7E0E8",
    "Sizes":                   "FFE9E1F2",
}


def _build_comparison_table(
    ws,
    *,
    per_image_df: pd.DataFrame,
    nuclei_df: pd.DataFrame,
) -> None:
    from openpyxl.formatting.rule import ColorScaleRule
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    BOLD = Font(bold=True, size=11)
    H1 = Font(bold=True, size=14, color="FFFFFFFF")
    SECTION_FILL = PatternFill("solid", fgColor=SECTION_FILL_HEX)
    TABLE_HDR_FILL = PatternFill("solid", fgColor=HEADER_FILL_HEX)
    THIN = Side(border_style="thin", color="FF999999")
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    WRAP = Alignment(wrap_text=True, vertical="top")
    LEFT_TOP = Alignment(horizontal="left", vertical="top", wrap_text=True)

    row = 1

    # Title strip
    ws.cell(row=row, column=1, value="COMPARISON TABLE — WT vs KO (Mann-Whitney U + Cliff's delta)").font = H1
    for cc in range(1, 10):
        ws.cell(row=row, column=cc).fill = SECTION_FILL
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
    ws.cell(row=row, column=1).alignment = LEFT_TOP
    ws.row_dimensions[row].height = 24
    row += 1

    # Sub-note
    note_cell = ws.cell(row=row, column=1, value=
        "Statistics computed per-nucleus (larger n than per-image). Means + SEMs shown for "
        "WT, KO, and sec-only. p-values are two-sided Mann-Whitney U. Cliff's delta is a "
        "nonparametric effect size in [-1, +1] (positive = WT > KO). Sec-only excluded from tests."
    )
    note_cell.alignment = WRAP
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
    ws.row_dimensions[row].height = 32
    row += 1
    row += 1  # spacer

    # Identify which condition is WT and which is KO (case-insensitive)
    if not len(nuclei_df) or "condition" not in nuclei_df.columns:
        ws.cell(row=row, column=1, value="(no nucleus-level data)").font = BOLD
        return

    if "secondary_only" in nuclei_df.columns:
        real_nuc = nuclei_df.loc[~nuclei_df["secondary_only"], :]
        sec_nuc = nuclei_df.loc[nuclei_df["secondary_only"], :]
    else:
        real_nuc = nuclei_df
        sec_nuc = nuclei_df.iloc[0:0]
    if "secondary_only" in per_image_df.columns:
        real_img = per_image_df.loc[~per_image_df["secondary_only"], :]
        sec_img = per_image_df.loc[per_image_df["secondary_only"], :]
    else:
        real_img = per_image_df
        sec_img = per_image_df.iloc[0:0]

    real_conds = list(dict.fromkeys(real_nuc["condition"].dropna().tolist()))
    real_conds.sort(key=lambda c: 0 if str(c).upper() in ("WT", "NT") else
                                  1 if str(c).upper() in ("KO", "KD") else 2)

    wt_cond = next((c for c in real_conds if str(c).upper() in ("WT", "NT")), None)
    ko_cond = next((c for c in real_conds if str(c).upper() in ("KO", "KD")), None)
    # Fallback to the first two if we don't see canonical names
    if wt_cond is None and real_conds:
        wt_cond = real_conds[0]
    if ko_cond is None and len(real_conds) >= 2:
        ko_cond = real_conds[1]

    # Header
    headers = [
        "Metric", "Category",
        f"{wt_cond}_mean", f"{wt_cond}_SEM", f"{wt_cond}_n",
        f"{ko_cond}_mean", f"{ko_cond}_SEM", f"{ko_cond}_n",
        "MWU p-value", "Cliff's delta",
        "sec-only_mean", "sec-only_n",
    ]
    for i, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=i, value=h)
        cell.font = BOLD
        cell.fill = TABLE_HDR_FILL
        cell.border = BORDER
        cell.alignment = LEFT_TOP
    pval_col_idx = headers.index("MWU p-value") + 1
    row += 1
    p_data_start = row

    def _row_values_for(metric_label: str,
                        img_col: str,
                        nuc_col: Optional[str],
                        category: str) -> List[Any]:
        # Prefer per-nucleus values for mean/SEM/n display when available
        # (larger n than per-image). Fall back to per-image otherwise.
        # Stats always use per-nucleus when nuc_col is supplied.
        use_nuc = nuc_col is not None and len(real_nuc) and nuc_col in real_nuc.columns

        # WT
        if wt_cond is not None:
            if use_nuc:
                wt_vals = real_nuc.loc[real_nuc["condition"] == wt_cond, nuc_col]
            elif img_col in real_img.columns:
                wt_vals = real_img.loc[real_img["condition"] == wt_cond, img_col]
            else:
                wt_vals = pd.Series(dtype=float)
            wt_mean, wt_sem, wt_n = _mean_sem_n(wt_vals)
        else:
            wt_mean, wt_sem, wt_n = None, None, 0
        # KO
        if ko_cond is not None:
            if use_nuc:
                ko_vals = real_nuc.loc[real_nuc["condition"] == ko_cond, nuc_col]
            elif img_col in real_img.columns:
                ko_vals = real_img.loc[real_img["condition"] == ko_cond, img_col]
            else:
                ko_vals = pd.Series(dtype=float)
            ko_mean, ko_sem, ko_n = _mean_sem_n(ko_vals)
        else:
            ko_mean, ko_sem, ko_n = None, None, 0

        # Stats: per-nucleus only
        pval, delta = None, None
        if use_nuc and wt_cond and ko_cond:
            a = real_nuc.loc[real_nuc["condition"] == wt_cond, nuc_col]
            b = real_nuc.loc[real_nuc["condition"] == ko_cond, nuc_col]
            pval, delta = _mannwhitney_and_cliffs_delta(a.values, b.values)

        # Sec-only: use per-image (sec-only nuclei have 0 spots → most
        # nuc-level metrics are undefined; per-image total/mean is informative)
        sec_mean = None
        sec_n = 0
        if len(sec_img) and img_col in sec_img.columns:
            sub = sec_img[img_col].dropna()
            sec_n = int(len(sub))
            sec_mean = float(sub.mean()) if sec_n else None

        return [
            metric_label, category,
            wt_mean, wt_sem, wt_n,
            ko_mean, ko_sem, ko_n,
            pval, delta,
            sec_mean, sec_n,
        ]

    for metric_label, img_col, nuc_col, category in COMPARISON_METRICS:
        # 2026-05-20 Brian/Sam: skip metric rows whose underlying columns
        # are absent from BOTH the per-image and per-nucleus tables. This
        # keeps the above-floor rows from polluting the table with blank
        # cells when output.apply_pub_contrast_floor_to_analysis = False.
        _img_present = img_col in per_image_df.columns
        _nuc_present = (nuc_col is not None) and (nuc_col in nuclei_df.columns)
        if not (_img_present or _nuc_present):
            continue
        vals = _row_values_for(metric_label, img_col, nuc_col, category)
        for i, v in enumerate(vals, start=1):
            cell = ws.cell(row=row, column=i, value=v)
            cell.border = BORDER
            cell.alignment = WRAP if i <= 2 else LEFT_TOP
        # Color the metric name and category cells by category
        cat_hex = CATEGORY_FILLS.get(category)
        if cat_hex:
            ws.cell(row=row, column=1).fill = PatternFill("solid", fgColor=cat_hex)
            ws.cell(row=row, column=2).fill = PatternFill("solid", fgColor=cat_hex)
        # Number formats: means / SEMs / sec-only_mean -> general float; p -> sci; delta -> 3dp; n -> int
        for i in (3, 4, 6, 7, 11):
            c = ws.cell(row=row, column=i)
            if c.value is not None:
                c.number_format = "0.0000"
        for i in (5, 8, 12):
            c = ws.cell(row=row, column=i)
            if c.value is not None:
                c.number_format = "0"
        # p-value
        c = ws.cell(row=row, column=9)
        if c.value is not None:
            c.number_format = "0.0000E+00"
        # delta
        c = ws.cell(row=row, column=10)
        if c.value is not None:
            c.number_format = "0.000"
        row += 1
    p_data_end = row - 1

    # Column widths
    from openpyxl.utils import get_column_letter
    widths = [44, 22, 14, 14, 10, 14, 14, 10, 14, 14, 14, 12]
    for j, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = "C4"

    # Conditional formatting on p-value column (green for low p)
    if p_data_end >= p_data_start:
        rng = f"{get_column_letter(pval_col_idx)}{p_data_start}:{get_column_letter(pval_col_idx)}{p_data_end}"
        rule = ColorScaleRule(
            start_type="num", start_value=0, start_color="FF1B5E20",       # dark green at p=0
            mid_type="num",   mid_value=0.05, mid_color="FF81C784",        # light green at p=0.05
            end_type="num",   end_value=1.0,  end_color="FFFFFFFF",        # white at p=1
        )
        ws.conditional_formatting.add(rng, rule)


# ---------------------------------------------------------------------------
# PI_Focus sheet (2026-05-20 Brian/Sam)
#
# One curated sheet the PI opens first. Every metric they care about,
# organized by metric × condition × compartment × channel, with
# Mann-Whitney p + Cliff's delta + direction arrow.
#
# Layout:
#   Section A — SPOT COUNTS PER CELL
#   Section B — SIGNAL ABOVE FLOOR (pixel intensity, floor-subtracted)
#   Section C — SPOT-INTRINSIC BRIGHTNESS (peak intensity, per-compartment)
#   Section D — SPOT SIZE (diameter um, per-compartment)
#   Section E — SIGNAL LOCALIZATION (nuclear fractions, overlap fractions)
#
# Stats:
#   - Per-nucleus columns (Sections A/B + nuclear-fraction rows in E) →
#     per-nucleus Mann-Whitney + Cliff's delta (matches Comparison_Table).
#   - Per-spot columns (Sections C/D) → per-image MEAN across each
#     image's filtered spot population, then Mann-Whitney across images.
#     This keeps the test n at per-image (n=1-2 per condition for the
#     BIN1 4-image preset — usually too few to be significant, will
#     just render p/delta = NaN; the means + n still show).
#   - Overlap-fraction rows (Section E) → per-nucleus when available.
# ---------------------------------------------------------------------------


PI_FOCUS_CATEGORY_FILLS: Dict[str, str] = {
    "A. SPOT COUNTS PER CELL":          "FFE8EFF7",  # light blue (Counts)
    "B. SIGNAL ABOVE FLOOR":            "FFF7E0E8",  # light pink (Intensities)
    "C. SPOT-INTRINSIC BRIGHTNESS":     "FFF7E0E8",
    "D. SPOT SIZE":                     "FFE9E1F2",  # lavender (Sizes)
    "E. SIGNAL LOCALIZATION":           "FFE7F4E0",  # light green (Localization)
    "F. COUNT x BRIGHTNESS — PER-NUCLEUS": "FFFFEEB3",  # warm amber (headline)
}


def _direction_arrow(
    wt_mean: Optional[float],
    ko_mean: Optional[float],
    delta: Optional[float],
) -> str:
    """Return an arrow + condition label for the comparison direction.

    delta = P(WT > KO) - P(WT < KO). Positive delta → WT tends higher.
    """
    if delta is None or not (delta == delta):  # NaN
        # Fall back to mean comparison if we have it
        if wt_mean is None or ko_mean is None:
            return ""
        if wt_mean > ko_mean:
            return "up WT"
        if ko_mean > wt_mean:
            return "up KO"
        return "="
    if abs(delta) < 1e-9:
        return "="
    return "up WT" if delta > 0 else "up KO"


def _per_nucleus_means_from_spots(
    spots_df: pd.DataFrame,
    channel: str,
    compartment: Optional[str],
    value_col: str,
    real_only_mask: Optional[pd.Series] = None,
) -> pd.DataFrame:
    """For per-spot quantities, compute the mean per (image, nucleus_id,
    condition) for a given channel/compartment subset. This gives much
    higher statistical N than the per-image aggregation (Section C/D),
    one row per nucleus that detected >0 spots in the subset.

    Returns a DataFrame with columns ['image', 'nucleus_id', 'condition',
    'mean']. Cells with no spots in the subset drop out (so n = cells with
    at least one spot of this type, not all cells).
    """
    if not len(spots_df) or "channel" not in spots_df.columns:
        return pd.DataFrame(columns=["image", "nucleus_id", "condition", "mean"])

    sub = spots_df.copy()
    if real_only_mask is not None:
        sub = sub.loc[real_only_mask, :]
    sub = sub.loc[sub["channel"].astype(str) == channel]
    if compartment == "nuclear" and "in_nucleus" in sub.columns:
        sub = sub.loc[sub["in_nucleus"].astype(bool) == True]
    elif compartment == "cytoplasmic" and "in_cytoplasm" in sub.columns:
        sub = sub.loc[sub["in_cytoplasm"].astype(bool) == True]
    if not len(sub) or value_col not in sub.columns:
        return pd.DataFrame(columns=["image", "nucleus_id", "condition", "mean"])
    if "nucleus_id" not in sub.columns:
        return pd.DataFrame(columns=["image", "nucleus_id", "condition", "mean"])
    sub[value_col] = pd.to_numeric(sub[value_col], errors="coerce")
    # For cytoplasmic, nucleus_id is still set to the parent cell — but
    # only-cytoplasmic spots have in_nucleus=0; we want to attribute them
    # to the *cell* they belong to. Drop spots with nucleus_id == 0/NaN
    # (orphan cytoplasmic spots not assigned to any cell).
    sub = sub.dropna(subset=[value_col, "nucleus_id"])
    sub = sub.loc[sub["nucleus_id"] > 0]
    if not len(sub):
        return pd.DataFrame(columns=["image", "nucleus_id", "condition", "mean"])
    grouped = (
        sub.groupby(["image", "nucleus_id", "condition"], dropna=False)[value_col]
        .mean()
        .reset_index()
        .rename(columns={value_col: "mean"})
    )
    return grouped


def _per_image_means_from_spots(
    spots_df: pd.DataFrame,
    channel: str,
    compartment: Optional[str],
    value_col: str,
    real_only_mask: Optional[pd.Series] = None,
) -> pd.DataFrame:
    """For per-spot quantities (peak intensity, diameter), compute the
    mean per (image, condition) for a given channel/compartment subset.

    Returns a DataFrame with columns ['image', 'condition', 'mean'].
    Empty subsets (no spots for an image+channel+compartment) drop out.
    """
    if not len(spots_df) or "channel" not in spots_df.columns:
        return pd.DataFrame(columns=["image", "condition", "mean"])

    sub = spots_df.copy()
    if real_only_mask is not None:
        sub = sub.loc[real_only_mask, :]
    sub = sub.loc[sub["channel"].astype(str) == channel]
    if compartment == "nuclear" and "in_nucleus" in sub.columns:
        sub = sub.loc[sub["in_nucleus"].astype(bool) == True]
    elif compartment == "cytoplasmic" and "in_cytoplasm" in sub.columns:
        sub = sub.loc[sub["in_cytoplasm"].astype(bool) == True]
    if not len(sub) or value_col not in sub.columns:
        return pd.DataFrame(columns=["image", "condition", "mean"])
    sub[value_col] = pd.to_numeric(sub[value_col], errors="coerce")
    sub = sub.dropna(subset=[value_col])
    if not len(sub):
        return pd.DataFrame(columns=["image", "condition", "mean"])
    grouped = (
        sub.groupby(["image", "condition"], dropna=False)[value_col]
        .mean()
        .reset_index()
        .rename(columns={value_col: "mean"})
    )
    return grouped


def _build_pi_focus(
    ws,
    *,
    per_image_df: pd.DataFrame,
    nuclei_df: pd.DataFrame,
    spots_df: pd.DataFrame,
) -> None:
    """Curated PI deliverable. See top-of-block docstring.

    Sheet layout: 6 columns —
        Metric | WT_mean ± SEM (n) | KO_mean ± SEM (n) | MWU p | Cliff's delta | Direction
    Section header rows merge across all 6 columns and use the dark-blue
    fill from the rest of the workbook.
    """
    from openpyxl.formatting.rule import ColorScaleRule
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    BOLD = Font(bold=True, size=11)
    BOLD_WHITE = Font(bold=True, size=12, color="FFFFFFFF")
    H1 = Font(bold=True, size=14, color="FFFFFFFF")
    SECTION_FILL = PatternFill("solid", fgColor=SECTION_FILL_HEX)
    SUBSECTION_FILL = PatternFill("solid", fgColor=SUBSECTION_FILL_HEX)
    TABLE_HDR_FILL = PatternFill("solid", fgColor=HEADER_FILL_HEX)
    THIN = Side(border_style="thin", color="FF999999")
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    WRAP = Alignment(wrap_text=True, vertical="top")
    LEFT_TOP = Alignment(horizontal="left", vertical="top", wrap_text=True)

    N_COLS = 6

    row = 1

    # ── Title strip ───────────────────────────────────────────────────────
    title_cell = ws.cell(
        row=row, column=1,
        value="PI-FOCUS METRICS — WT vs KO per channel per compartment",
    )
    title_cell.font = H1
    for cc in range(1, N_COLS + 1):
        ws.cell(row=row, column=cc).fill = SECTION_FILL
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=N_COLS)
    ws.cell(row=row, column=1).alignment = LEFT_TOP
    ws.row_dimensions[row].height = 26
    row += 1

    note = ws.cell(
        row=row, column=1,
        value=(
            "One-stop deliverable. Sections A/B/E compute statistics PER-NUCLEUS "
            "(n = nuclei per condition). Sections C/D compute PER-SPOT means "
            "first, then run the test across PER-IMAGE means (test n = images "
            "per condition). Direction column: arrow + condition that the metric "
            "trends toward (sign of Cliff's delta; mean fallback if delta is NaN). "
            "Stat tests are skipped when either condition has < 3 samples (cell "
            "renders blank). Sec-only images are excluded from tests."
        ),
    )
    note.alignment = WRAP
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=N_COLS)
    ws.row_dimensions[row].height = 56
    row += 1
    row += 1  # spacer

    # ── Identify WT / KO real conditions ──────────────────────────────────
    if not len(nuclei_df) or "condition" not in nuclei_df.columns:
        ws.cell(row=row, column=1, value="(no nucleus-level data)").font = BOLD
        return

    if "secondary_only" in nuclei_df.columns:
        real_nuc = nuclei_df.loc[~nuclei_df["secondary_only"].astype(bool), :]
    else:
        real_nuc = nuclei_df
    if "secondary_only" in per_image_df.columns:
        real_img_mask = ~per_image_df["secondary_only"].astype(bool)
        real_img = per_image_df.loc[real_img_mask, :]
    else:
        real_img = per_image_df
        real_img_mask = pd.Series(True, index=per_image_df.index)
    # Spot-level real-only mask: filter by image membership in real_img
    if len(spots_df) and "image" in spots_df.columns and "image" in real_img.columns:
        real_images = set(real_img["image"].astype(str).tolist())
        spot_real_mask = spots_df["image"].astype(str).isin(real_images)
    else:
        spot_real_mask = None

    real_conds = list(dict.fromkeys(real_nuc["condition"].dropna().tolist()))
    real_conds.sort(key=lambda c: 0 if str(c).upper() in ("WT", "NT") else
                                  1 if str(c).upper() in ("KO", "KD") else 2)
    wt_cond = next((c for c in real_conds if str(c).upper() in ("WT", "NT")), None)
    ko_cond = next((c for c in real_conds if str(c).upper() in ("KO", "KD")), None)
    if wt_cond is None and real_conds:
        wt_cond = real_conds[0]
    if ko_cond is None and len(real_conds) >= 2:
        ko_cond = real_conds[1]

    # ── Header row ────────────────────────────────────────────────────────
    headers = [
        "Metric",
        f"{wt_cond} mean +/- SEM (n)",
        f"{ko_cond} mean +/- SEM (n)",
        "Mann-Whitney p",
        "Cliff's delta",
        "Direction",
    ]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = BOLD
        c.fill = TABLE_HDR_FILL
        c.border = BORDER
        c.alignment = LEFT_TOP
    pval_col_idx = headers.index("Mann-Whitney p") + 1
    row += 1
    p_data_start = row

    # ── Helpers ───────────────────────────────────────────────────────────
    def write_section_header(title: str) -> None:
        nonlocal row
        c = ws.cell(row=row, column=1, value=title)
        c.font = BOLD_WHITE
        for cc in range(1, N_COLS + 1):
            ws.cell(row=row, column=cc).fill = SUBSECTION_FILL
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=N_COLS)
        ws.cell(row=row, column=1).alignment = LEFT_TOP
        ws.row_dimensions[row].height = 22
        row += 1

    def _fmt_mean_sem_n(mean: Optional[float], sem: Optional[float], n: int) -> str:
        if mean is None or n == 0:
            return f"n={n}"
        if sem is None or not (sem == sem):  # NaN
            return f"{mean:.3g} (n={n})"
        return f"{mean:.3g} +/- {sem:.3g} (n={n})"

    def write_metric_row(
        metric_label: str,
        wt_vals: Sequence[float],
        ko_vals: Sequence[float],
        section_title: str,
    ) -> None:
        nonlocal row
        wt_mean, wt_sem, wt_n = _mean_sem_n(wt_vals)
        ko_mean, ko_sem, ko_n = _mean_sem_n(ko_vals)
        # Stat test only when n >= 3 on BOTH sides — skip silently otherwise
        # (per requirements: stat tests skip silently when sample size < 3).
        if wt_n >= 3 and ko_n >= 3:
            pval, delta = _mannwhitney_and_cliffs_delta(
                list(wt_vals), list(ko_vals),
            )
        else:
            pval, delta = None, None
        direction = _direction_arrow(wt_mean, ko_mean, delta)

        ws.cell(row=row, column=1, value=metric_label)
        ws.cell(row=row, column=2, value=_fmt_mean_sem_n(wt_mean, wt_sem, wt_n))
        ws.cell(row=row, column=3, value=_fmt_mean_sem_n(ko_mean, ko_sem, ko_n))
        ws.cell(row=row, column=4, value=pval if pval is not None else None)
        ws.cell(row=row, column=5, value=delta if delta is not None else None)
        ws.cell(row=row, column=6, value=direction)
        # Borders + alignment + per-section fill on metric name column
        for cc in range(1, N_COLS + 1):
            cell = ws.cell(row=row, column=cc)
            cell.border = BORDER
            cell.alignment = LEFT_TOP
        cat_fill_hex = PI_FOCUS_CATEGORY_FILLS.get(section_title)
        if cat_fill_hex:
            ws.cell(row=row, column=1).fill = PatternFill("solid", fgColor=cat_fill_hex)
        # Number formats
        if pval is not None:
            ws.cell(row=row, column=4).number_format = "0.0000E+00"
        if delta is not None:
            ws.cell(row=row, column=5).number_format = "0.000"
        row += 1

    def get_nuc_vals(col: str, cond: Any) -> Sequence[float]:
        if cond is None or col not in real_nuc.columns:
            return []
        s = real_nuc.loc[real_nuc["condition"] == cond, col]
        return pd.to_numeric(s, errors="coerce").dropna().tolist()

    def get_img_vals(col: str, cond: Any) -> Sequence[float]:
        if cond is None or col not in real_img.columns:
            return []
        s = real_img.loc[real_img["condition"] == cond, col]
        return pd.to_numeric(s, errors="coerce").dropna().tolist()

    # Helper to build per-image mean-of-spot values for Sections C/D.
    def per_image_spot_vals(
        channel: str, compartment: Optional[str], value_col: str, cond: Any,
    ) -> Sequence[float]:
        if cond is None:
            return []
        agg = _per_image_means_from_spots(
            spots_df, channel, compartment, value_col, real_only_mask=spot_real_mask,
        )
        if not len(agg):
            return []
        return agg.loc[agg["condition"] == cond, "mean"].tolist()

    # ── Section A: SPOT COUNTS PER CELL ───────────────────────────────────
    sec_A = "A. SPOT COUNTS PER CELL"
    write_section_header(sec_A)
    # Total / nuclear / cytoplasmic per channel — use per-nucleus columns
    # (n_spots_rna1, nuclear_spot_count, cyto_spot_count, n_spots_rna2,
    # nuclear_spot_count_rna2, cyto_spot_count_rna2,
    # n_nuclear_rna1_rna2_overlap_per_nucleus).
    a_rows = [
        ("Total Introns spots per cell",          "n_spots_rna1"),
        ("Nuclear Introns spots per cell",        "nuclear_spot_count"),
        ("Cytoplasmic Introns spots per cell",    "cyto_spot_count"),
        ("Total Exons spots per cell",            "n_spots_rna2"),
        ("Nuclear Exons spots per cell",          "nuclear_spot_count_rna2"),
        ("Cytoplasmic Exons spots per cell",      "cyto_spot_count_rna2"),
        ("Nuclear Introns+Exons overlap per cell", "n_nuclear_rna1_rna2_overlap_per_nucleus"),
    ]
    for label, col in a_rows:
        write_metric_row(label, get_nuc_vals(col, wt_cond), get_nuc_vals(col, ko_cond), sec_A)

    # spacer
    row += 1

    # ── Section B: SIGNAL ABOVE FLOOR ─────────────────────────────────────
    sec_B = "B. SIGNAL ABOVE FLOOR"
    write_section_header(sec_B)
    b_rows = [
        ("Mean per-pixel above-floor intensity, Introns NUCLEAR",
         "nuclear_above_floor_intensity_rna1"),
        ("Mean per-pixel above-floor intensity, Introns CYTOPLASM",
         "cytoplasmic_above_floor_intensity_rna1"),
        ("N/C ratio above-floor, Introns",
         "nc_ratio_above_floor_intensity_rna1"),
        ("Mean per-pixel above-floor intensity, Exons NUCLEAR",
         "nuclear_above_floor_intensity_rna2"),
        ("Mean per-pixel above-floor intensity, Exons CYTOPLASM",
         "cytoplasmic_above_floor_intensity_rna2"),
        ("N/C ratio above-floor, Exons",
         "nc_ratio_above_floor_intensity_rna2"),
    ]
    for label, col in b_rows:
        write_metric_row(label, get_nuc_vals(col, wt_cond), get_nuc_vals(col, ko_cond), sec_B)

    row += 1

    # ── Section C: SPOT-INTRINSIC BRIGHTNESS ──────────────────────────────
    # Per-spot peak intensity → mean per image → compare across images.
    sec_C = "C. SPOT-INTRINSIC BRIGHTNESS"
    write_section_header(sec_C)
    c_specs = [
        ("Introns spot peak intensity — NUCLEAR spots only",  "rna1",  "nuclear",     "spot_peak_intensity"),
        ("Introns spot peak intensity — CYTOPLASMIC spots only", "rna1", "cytoplasmic", "spot_peak_intensity"),
        ("Exons spot peak intensity — NUCLEAR spots only",    "rna2",  "nuclear",     "spot_peak_intensity"),
        ("Exons spot peak intensity — CYTOPLASMIC spots only", "rna2", "cytoplasmic", "spot_peak_intensity"),
    ]
    # Cache per-image means so we can compute the nuc/cyto ratio per channel.
    c_image_means: Dict[Tuple[str, Optional[str]], pd.DataFrame] = {}
    for label, ch, comp, val_col in c_specs:
        key = (ch, comp)
        if key not in c_image_means:
            c_image_means[key] = _per_image_means_from_spots(
                spots_df, ch, comp, val_col, real_only_mask=spot_real_mask,
            )
        agg = c_image_means[key]
        wt_v = agg.loc[agg["condition"] == wt_cond, "mean"].tolist() if wt_cond is not None else []
        ko_v = agg.loc[agg["condition"] == ko_cond, "mean"].tolist() if ko_cond is not None else []
        write_metric_row(label, wt_v, ko_v, sec_C)

    # Add nuc/cyto ratio rows per channel (per-image: mean_nuc / mean_cyto).
    for ch_label, ch in (("Introns", "rna1"), ("Exons", "rna2")):
        nuc_agg = c_image_means.get((ch, "nuclear"))
        cyt_agg = c_image_means.get((ch, "cytoplasmic"))
        if nuc_agg is None or cyt_agg is None or not len(nuc_agg) or not len(cyt_agg):
            # Render the row anyway with empty values for schema consistency
            write_metric_row(
                f"{ch_label} spot peak intensity ratio (nuc/cyto)", [], [], sec_C,
            )
            continue
        merged = pd.merge(
            nuc_agg.rename(columns={"mean": "nuc"}),
            cyt_agg.rename(columns={"mean": "cyt"}),
            on=["image", "condition"], how="inner",
        )
        # Avoid div-by-zero: drop rows where cyt==0 (treat as NaN ratio)
        merged = merged.loc[merged["cyt"] > 0, :].copy()
        merged["ratio"] = merged["nuc"] / merged["cyt"]
        wt_v = merged.loc[merged["condition"] == wt_cond, "ratio"].tolist() if wt_cond is not None else []
        ko_v = merged.loc[merged["condition"] == ko_cond, "ratio"].tolist() if ko_cond is not None else []
        write_metric_row(
            f"{ch_label} spot peak intensity ratio (nuc/cyto)", wt_v, ko_v, sec_C,
        )

    row += 1

    # ── Section D: SPOT SIZE ──────────────────────────────────────────────
    sec_D = "D. SPOT SIZE"
    write_section_header(sec_D)
    d_specs = [
        ("Introns spot diameter (um) — NUCLEAR",   "rna1", "nuclear",     "spot_diameter_um"),
        ("Introns spot diameter (um) — CYTOPLASMIC", "rna1", "cytoplasmic", "spot_diameter_um"),
        ("Exons spot diameter (um) — NUCLEAR",     "rna2", "nuclear",     "spot_diameter_um"),
        ("Exons spot diameter (um) — CYTOPLASMIC", "rna2", "cytoplasmic", "spot_diameter_um"),
    ]
    for label, ch, comp, val_col in d_specs:
        agg = _per_image_means_from_spots(
            spots_df, ch, comp, val_col, real_only_mask=spot_real_mask,
        )
        wt_v = agg.loc[agg["condition"] == wt_cond, "mean"].tolist() if wt_cond is not None else []
        ko_v = agg.loc[agg["condition"] == ko_cond, "mean"].tolist() if ko_cond is not None else []
        write_metric_row(label, wt_v, ko_v, sec_D)

    row += 1

    # ── Section E: SIGNAL LOCALIZATION ────────────────────────────────────
    sec_E = "E. SIGNAL LOCALIZATION"
    write_section_header(sec_E)
    # Fraction of total spots that are nuclear: prefer per-nucleus
    # "nuclear_spot_fraction" / "nuclear_spot_fraction_rna2" (population
    # of per-cell nuclear fractions over cells that detected >0 spots).
    e_rows_nuc = [
        ("Fraction of total Introns spots that are nuclear",
         "nuclear_spot_fraction"),
        ("Fraction of total Exons spots that are nuclear",
         "nuclear_spot_fraction_rna2"),
    ]
    for label, col in e_rows_nuc:
        write_metric_row(label, get_nuc_vals(col, wt_cond), get_nuc_vals(col, ko_cond), sec_E)

    # Overlap fractions — per-nucleus paired_fraction_*_at_0p3um is the
    # asymmetric per-cell paired fraction; fall back to per-image
    # "paired_fraction_rna1_at_0p3um" if absent at the per-nucleus level.
    overlap_specs = [
        ("Fraction of total Introns spots that overlap with Exons",
         "paired_fraction_rna1_at_0p3um"),
        ("Fraction of total Exons spots that overlap with Introns",
         "paired_fraction_rna2_at_0p3um"),
    ]
    for label, col in overlap_specs:
        if col in real_nuc.columns:
            wt_v = get_nuc_vals(col, wt_cond)
            ko_v = get_nuc_vals(col, ko_cond)
        else:
            wt_v = get_img_vals(col, wt_cond)
            ko_v = get_img_vals(col, ko_cond)
        write_metric_row(label, wt_v, ko_v, sec_E)

    # ── Section F: COUNT × BRIGHTNESS — PER-NUCLEUS ──────────────────────
    # 2026-05-21 Brian: the "fewer-but-brighter" story needs to be
    # surfaced at PER-NUCLEUS granularity so it has real statistical
    # power (n=cells, not n=images). Section C/D test across images,
    # which collapses to n=1-3 for typical small datasets and renders
    # no p-values. Section F pairs spot count with per-cell mean spot
    # peak intensity AND per-cell total spot intensity (= count × peak,
    # the combined headline). All from spots_df aggregated per cell.
    row += 1
    sec_F = "F. COUNT x BRIGHTNESS — PER-NUCLEUS"
    write_section_header(sec_F)

    def per_nuc_spot_vals(
        channel: str, compartment: Optional[str], value_col: str, cond: Any,
    ) -> Sequence[float]:
        if cond is None:
            return []
        agg = _per_nucleus_means_from_spots(
            spots_df, channel, compartment, value_col, real_only_mask=spot_real_mask,
        )
        if not len(agg):
            return []
        return agg.loc[agg["condition"] == cond, "mean"].tolist()

    f_specs_brightness = [
        ("Introns mean spot peak per cell — ANY compartment",   "rna1", None,          "spot_peak_intensity"),
        ("Introns mean spot peak per cell — NUCLEAR spots",     "rna1", "nuclear",     "spot_peak_intensity"),
        ("Introns mean spot peak per cell — CYTOPLASMIC spots", "rna1", "cytoplasmic", "spot_peak_intensity"),
        ("Exons mean spot peak per cell — ANY compartment",     "rna2", None,          "spot_peak_intensity"),
        ("Exons mean spot peak per cell — NUCLEAR spots",       "rna2", "nuclear",     "spot_peak_intensity"),
        ("Exons mean spot peak per cell — CYTOPLASMIC spots",   "rna2", "cytoplasmic", "spot_peak_intensity"),
    ]
    for label, ch, comp, val_col in f_specs_brightness:
        wt_v = per_nuc_spot_vals(ch, comp, val_col, wt_cond)
        ko_v = per_nuc_spot_vals(ch, comp, val_col, ko_cond)
        write_metric_row(label, wt_v, ko_v, sec_F)

    # Total integrated spot intensity per cell — the combined "count x brightness"
    # headline. Already in nuclei_metrics as rna_spot_total_intensity_fit.
    f_specs_totals = [
        ("Total Introns spot intensity per cell (count x peak combined)",
         "rna_spot_total_intensity_fit"),
        ("Total Exons spot intensity per cell (count x peak combined)",
         "rna2_spot_total_intensity_fit"),
    ]
    for label, col in f_specs_totals:
        write_metric_row(label, get_nuc_vals(col, wt_cond), get_nuc_vals(col, ko_cond), sec_F)

    # ── Section G: NUCLEOLUS SUBNUCLEAR BREAKDOWN (only if nucleolus on) ──
    # 2026-05-21 Brian: when cfg.nucleolus.enabled is True, the spot_metrics
    # CSV has an `in_nucleolus` column. Surface nucleolar spots per cell
    # AND fraction-of-nuclear-spots-in-nucleolus per cell, so the WT vs KO
    # comparison can directly answer "is the nuclear retention going INTO
    # the nucleolus, or into chromatin?"
    has_nucleolus_col = "in_nucleolus" in spots_df.columns if len(spots_df) else False
    if has_nucleolus_col:
        row += 1
        sec_G = "G. NUCLEOLUS SUBNUCLEAR BREAKDOWN"
        # Use a fresh fill color
        PI_FOCUS_CATEGORY_FILLS[sec_G] = "FFFFD8B5"  # light orange (nucleolus)
        write_section_header(sec_G)

        # Helper: per-nucleus aggregation of "spots with `mask_col` == True"
        # for a given channel.
        def per_nuc_nucleolus_counts(channel: str, in_nucleolus_only: bool, cond: Any) -> Sequence[float]:
            if cond is None or not len(spots_df):
                return []
            sub = spots_df.copy()
            if spot_real_mask is not None:
                sub = sub.loc[spot_real_mask, :]
            if "channel" in sub.columns:
                sub = sub.loc[sub["channel"].astype(str) == channel]
            if not len(sub) or "nucleus_id" not in sub.columns:
                return []
            # Only count spots inside the nucleus mask
            sub = sub.loc[sub["in_nucleus"].astype(bool) == True]
            if in_nucleolus_only:
                sub = sub.loc[sub["in_nucleolus"].astype(bool) == True]
            sub = sub.loc[sub["nucleus_id"] > 0]
            counts = (
                sub.groupby(["image", "nucleus_id"]).size()
                .reset_index(name="n")
            )
            # Attach condition by joining via nuclei_df
            if cond is not None and len(real_nuc):
                _nuc = real_nuc[["image", "nucleus_id", "condition"]]
                counts = counts.merge(_nuc, on=["image", "nucleus_id"], how="left")
            return counts.loc[counts["condition"] == cond, "n"].tolist()

        # Per-cell nucleolar count rows
        g_count_rows = [
            ("Nucleolar Introns spots per cell", "rna1"),
            ("Nucleolar Exons spots per cell",   "rna2"),
        ]
        for label, ch in g_count_rows:
            wt_v = per_nuc_nucleolus_counts(ch, True, wt_cond)
            ko_v = per_nuc_nucleolus_counts(ch, True, ko_cond)
            write_metric_row(label, wt_v, ko_v, sec_G)

        # Fraction-of-nuclear-spots-in-nucleolus per cell
        def per_nuc_nucleolus_fraction(channel: str, cond: Any) -> Sequence[float]:
            if cond is None or not len(spots_df):
                return []
            sub = spots_df.copy()
            if spot_real_mask is not None:
                sub = sub.loc[spot_real_mask, :]
            if "channel" in sub.columns:
                sub = sub.loc[sub["channel"].astype(str) == channel]
            sub = sub.loc[(sub["in_nucleus"].astype(bool) == True) & (sub["nucleus_id"] > 0)]
            if not len(sub):
                return []
            # For each cell, fraction of nuclear spots that are in nucleolus
            grouped = sub.groupby(["image", "nucleus_id"]).agg(
                n_nuc=("in_nucleus", "sum"),
                n_nucleolus=("in_nucleolus", "sum"),
            ).reset_index()
            grouped["frac"] = grouped["n_nucleolus"] / grouped["n_nuc"].replace(0, np.nan)
            if len(real_nuc):
                _nuc = real_nuc[["image", "nucleus_id", "condition"]]
                grouped = grouped.merge(_nuc, on=["image", "nucleus_id"], how="left")
            return grouped.loc[grouped["condition"] == cond, "frac"].dropna().tolist()

        g_frac_rows = [
            ("Fraction of nuclear Introns spots that are in nucleolus", "rna1"),
            ("Fraction of nuclear Exons spots that are in nucleolus",   "rna2"),
        ]
        for label, ch in g_frac_rows:
            wt_v = per_nuc_nucleolus_fraction(ch, wt_cond)
            ko_v = per_nuc_nucleolus_fraction(ch, ko_cond)
            write_metric_row(label, wt_v, ko_v, sec_G)

        # Per-nucleus nucleolus area (computed by nucleolus.py + merged
        # into nuclei_metrics.csv)
        if "nucleolus_area_px" in real_nuc.columns:
            write_metric_row(
                "Nucleolus area (px)",
                get_nuc_vals("nucleolus_area_px", wt_cond),
                get_nuc_vals("nucleolus_area_px", ko_cond),
                sec_G,
            )
        if "nucleolus_fraction_of_nucleus" in real_nuc.columns:
            write_metric_row(
                "Nucleolus area / nucleus area",
                get_nuc_vals("nucleolus_fraction_of_nucleus", wt_cond),
                get_nuc_vals("nucleolus_fraction_of_nucleus", ko_cond),
                sec_G,
            )
        if "dapi_cv" in real_nuc.columns:
            write_metric_row(
                "DAPI CV per nucleus (chromatin texture)",
                get_nuc_vals("dapi_cv", wt_cond),
                get_nuc_vals("dapi_cv", ko_cond),
                sec_G,
            )
        if "heterochromatin_fraction" in real_nuc.columns:
            write_metric_row(
                "Heterochromatin fraction (DAPI >= 1.5x nuclear median)",
                get_nuc_vals("heterochromatin_fraction", wt_cond),
                get_nuc_vals("heterochromatin_fraction", ko_cond),
                sec_G,
            )

    p_data_end = row - 1

    # ── Column widths + freeze ────────────────────────────────────────────
    widths = [62, 24, 24, 16, 14, 14]
    for j, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = "A5"

    # ── Conditional formatting on p-value column (green for low p) ────────
    if p_data_end >= p_data_start:
        rng = f"{get_column_letter(pval_col_idx)}{p_data_start}:{get_column_letter(pval_col_idx)}{p_data_end}"
        rule = ColorScaleRule(
            start_type="num", start_value=0,    start_color="FF1B5E20",
            mid_type="num",   mid_value=0.05,   mid_color="FF81C784",
            end_type="num",   end_value=1.0,    end_color="FFFFFFFF",
        )
        ws.conditional_formatting.add(rng, rule)


# ---------------------------------------------------------------------------
# Run_Config sheet
# ---------------------------------------------------------------------------

def _flatten_run_config(output_dir: Path) -> Dict[str, Any]:
    """Read run_config.json from output_dir if present and flatten it."""
    candidate = output_dir / "run_config.json"
    if not candidate.exists():
        return {}
    try:
        with open(candidate, "r", encoding="utf-8") as fh:
            data = json.load(fh)
    except Exception:
        return {}
    flat: Dict[str, Any] = {}

    def _walk(prefix: str, node: Any) -> None:
        if isinstance(node, dict):
            for k, v in node.items():
                key = f"{prefix}.{k}" if prefix else str(k)
                _walk(key, v)
        elif isinstance(node, list):
            # Render lists as JSON to keep one row per key
            flat[prefix] = json.dumps(node)
        else:
            flat[prefix] = node

    _walk("", data)
    return flat


CRITICAL_RUN_CONFIG_KEYS = {
    # All-caps top-level keys written by runner.py
    "ANALYSIS_MODE", "Z_MODE", "Z_START", "Z_END", "APPLIED_PROFILE",
    "SEGMENTATION_BACKEND", "SPOT_BACKEND",
    "BIGFISH_VOXEL_SIZE_NM", "BIGFISH_VOXEL_Z_NM",
    "BIGFISH_SPOT_RADIUS_NM", "BIGFISH_SPOT_RADIUS_Z_NM",
    "BIGFISH_THRESHOLD", "NUC_MIN_AREA_PX", "EXCLUDE_BORDER_NUCLEI",
    "DISP_FLOOR_PERCENTILE", "DISP_CEIL_PERCENTILE",
    # Resolved/nested keys that matter
    "config_resolved.z_stack.mode", "config_resolved.z_stack.start_slice",
    "config_resolved.z_stack.end_slice",
    "config_resolved.foci.rna_threshold_multiplier",
    "config_resolved.foci.rna2_threshold_multiplier",
    "config_resolved.foci.bigfish_voxel_size_nm",
    "config_resolved.foci.bigfish_voxel_z_nm",
    "config_resolved.foci.bigfish_spot_radius_nm",
    "config_resolved.foci.bigfish_spot_radius_z_nm",
    "config_resolved.publication.contrast_mode",
    "config_resolved.publication.batch_contrast.rna.floor",
    "config_resolved.publication.batch_contrast.rna.ceil",
    "config_resolved.publication.batch_contrast.rna2.floor",
    "config_resolved.publication.batch_contrast.rna2.ceil",
    "config_resolved.publication.batch_contrast.dapi.floor",
    "config_resolved.publication.batch_contrast.dapi.ceil",
    # Top-level batch_contrast (where runner.py actually writes the computed
    # per-channel floor/ceil — under the canonical run_config "batch_contrast"
    # key, not nested under publication).
    "batch_contrast.rna.floor",
    "batch_contrast.rna.ceil",
    "batch_contrast.rna2.floor",
    "batch_contrast.rna2.ceil",
    "batch_contrast.dapi.floor",
    "batch_contrast.dapi.ceil",
    # 2026-05-20 reference_image mode (Sam-style per-channel tuning).
    "PUB_CONTRAST_MODE",
    "config_resolved.output.pub_contrast_mode",
    "config_resolved.output.manual_rna_reference_image",
    "config_resolved.output.manual_rna_floor_region",
    "config_resolved.output.manual_rna_floor_pct",
    "config_resolved.output.manual_rna_ceil_region",
    "config_resolved.output.manual_rna_ceil_pct",
    "config_resolved.output.manual_rna2_reference_image",
    "config_resolved.output.manual_rna2_floor_region",
    "config_resolved.output.manual_rna2_floor_pct",
    "config_resolved.output.manual_rna2_ceil_region",
    "config_resolved.output.manual_rna2_ceil_pct",
    "config_resolved.nuclei.backend",
    "config_resolved.nuclei.min_area_px",
    "config_resolved.nuclei.exclude_border",
    "config_resolved.pixel_coloc.threshold_scope",
    "config_resolved.pixel_coloc.k_mad",
    # 2026-05-20 pub-contrast-floor-as-analysis-floor + spot-detection floor.
    "config_resolved.output.apply_pub_contrast_floor_to_analysis",
    "config_resolved.output.apply_pub_contrast_floor_to_spots",
    "APPLY_PUB_CONTRAST_FLOOR_TO_ANALYSIS",
    "APPLY_PUB_CONTRAST_FLOOR_TO_SPOTS",
}


def _build_run_config_sheet(ws, flat: Dict[str, Any]) -> None:
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    BOLD = Font(bold=True, size=11)
    H1 = Font(bold=True, size=14, color="FFFFFFFF")
    SECTION_FILL = PatternFill("solid", fgColor=SECTION_FILL_HEX)
    TABLE_HDR_FILL = PatternFill("solid", fgColor=HEADER_FILL_HEX)
    CRITICAL_FILL = PatternFill("solid", fgColor="FFFFF2CC")  # pale yellow
    THIN = Side(border_style="thin", color="FF999999")
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    WRAP = Alignment(wrap_text=True, vertical="top")
    LEFT_TOP = Alignment(horizontal="left", vertical="top", wrap_text=True)

    row = 1
    c = ws.cell(row=row, column=1, value="RUN_CONFIG — flattened parameters")
    c.font = H1
    c.fill = SECTION_FILL
    ws.cell(row=row, column=2).fill = SECTION_FILL
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    ws.row_dimensions[row].height = 22
    row += 1
    # Note row
    note = ws.cell(row=row, column=1, value=
        "Every parameter from run_config.json. Pale-yellow rows = critical params "
        "(contrast mode, z window, BigFISH multipliers, segmentation backend, etc.).")
    note.alignment = WRAP
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    ws.row_dimensions[row].height = 32
    row += 1
    row += 1

    for i, h in enumerate(["Parameter", "Value"], start=1):
        cell = ws.cell(row=row, column=i, value=h)
        cell.font = BOLD
        cell.fill = TABLE_HDR_FILL
        cell.border = BORDER
        cell.alignment = LEFT_TOP
    row += 1

    if not flat:
        ws.cell(row=row, column=1, value="(run_config.json not yet written at workbook build time)")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        return

    for key in sorted(flat.keys()):
        val = flat[key]
        is_crit = key in CRITICAL_RUN_CONFIG_KEYS
        k_cell = ws.cell(row=row, column=1, value=key)
        v_cell = ws.cell(row=row, column=2, value=str(val))
        k_cell.border = BORDER
        v_cell.border = BORDER
        k_cell.alignment = WRAP
        v_cell.alignment = WRAP
        if is_crit:
            k_cell.font = BOLD
            v_cell.font = BOLD
            k_cell.fill = CRITICAL_FILL
            v_cell.fill = CRITICAL_FILL
        row += 1

    ws.column_dimensions["A"].width = 60
    ws.column_dimensions["B"].width = 80
    ws.freeze_panes = "A5"


# ---------------------------------------------------------------------------
# Public entry point — analysis_summary.xlsx
# ---------------------------------------------------------------------------

def write_analysis_summary_workbook(
    *,
    out_path: Path,
    per_image_df: pd.DataFrame,
    nuclei_df: pd.DataFrame,
    spots_df: pd.DataFrame,
    morph_df: pd.DataFrame,
    thr_df: pd.DataFrame,
    fishsuite_version: str,
    run_start_utc: str,
    config_path: Path,
    input_dir: Path,
    output_dir: Path,
    z_mode: str,
    z_start: int,
    z_end: int,
    images: list,
    n_workers: int,
) -> List[str]:
    """Write the full PI-ready ``analysis_summary.xlsx`` workbook.

    Sheets, in order:
        1.  README
        2.  Executive_Summary
        3.  PI_Focus
        4.  Comparison_Table
        5.  Per_Image_Summary
        6.  Per_Nucleus_Metrics
        7.  Per_Spot_Metrics
        8.  Cell_Morphology
        9.  Thresholds
        10. Run_Config

    Returns the list of column names that fell through the glossary
    fallback so callers can log them.
    """
    out_path = Path(out_path)
    output_dir = Path(output_dir)

    # Sort all data sheets by (condition, image, [nucleus/spot id])
    per_image_sorted = _sort_df_by_condition(per_image_df, extra_sort_cols=["image"])
    nuclei_sorted = _sort_df_by_condition(nuclei_df, extra_sort_cols=["image", "nucleus_id"])
    spots_sorted = _sort_df_by_condition(spots_df, extra_sort_cols=["image", "channel", "spot_id"])
    morph_sorted = _sort_df_by_condition(morph_df, extra_sort_cols=["image", "nucleus_id"])
    # Thresholds: sort by joining the per-image condition lookup
    if len(thr_df) and "image" in thr_df.columns and "image" in per_image_sorted.columns:
        cond_lookup = per_image_sorted.set_index("image")["condition"].to_dict()
        sec_lookup = (
            per_image_sorted.set_index("image")["secondary_only"].to_dict()
            if "secondary_only" in per_image_sorted.columns else {}
        )
        thr_with_cond = thr_df.copy()
        thr_with_cond["__cond"] = thr_with_cond["image"].map(cond_lookup)
        thr_with_cond["__sec"] = thr_with_cond["image"].map(sec_lookup)
        thr_with_cond["__sort_key"] = [
            _condition_sort_key(c, s) for c, s in zip(thr_with_cond["__cond"], thr_with_cond["__sec"])
        ]
        thr_sorted = (
            thr_with_cond.sort_values(["__sort_key", "image"], kind="stable")
            .drop(columns=["__cond", "__sec", "__sort_key"])
            .reset_index(drop=True)
        )
    else:
        thr_sorted = thr_df.copy()

    # ---- 1. Stub workbook ---------------------------------------------------
    with pd.ExcelWriter(out_path, engine="openpyxl") as xl:
        pd.DataFrame({"_stub": [""]}).to_excel(xl, sheet_name="README", index=False)
        pd.DataFrame({"_stub": [""]}).to_excel(xl, sheet_name="Executive_Summary", index=False)
        pd.DataFrame({"_stub": [""]}).to_excel(xl, sheet_name="PI_Focus", index=False)
        pd.DataFrame({"_stub": [""]}).to_excel(xl, sheet_name="Comparison_Table", index=False)
        per_image_sorted.to_excel(xl, sheet_name="Per_Image_Summary", index=False)
        if len(nuclei_sorted):
            nuclei_sorted.to_excel(xl, sheet_name="Per_Nucleus_Metrics", index=False)
        if len(spots_sorted):
            spots_sorted.to_excel(xl, sheet_name="Per_Spot_Metrics", index=False)
        if len(morph_sorted):
            morph_sorted.to_excel(xl, sheet_name="Cell_Morphology", index=False)
        if len(thr_sorted):
            thr_sorted.to_excel(xl, sheet_name="Thresholds", index=False)
        pd.DataFrame({"_stub": [""]}).to_excel(xl, sheet_name="Run_Config", index=False)

    # ---- 2. Re-open + rebuild & format -------------------------------------
    from openpyxl import load_workbook

    wb = load_workbook(out_path)

    # Replace stub sheets
    for name in ("README", "Executive_Summary", "PI_Focus", "Comparison_Table", "Run_Config"):
        if name in wb.sheetnames:
            del wb[name]

    # README first
    readme_ws = wb.create_sheet("README", 0)
    run_cfg_flat = _flatten_run_config(output_dir)
    fallback_cols = _build_readme(
        readme_ws,
        fishsuite_version=fishsuite_version,
        run_start_utc=run_start_utc,
        config_path=config_path,
        input_dir=input_dir,
        output_dir=output_dir,
        z_mode=z_mode,
        z_start=z_start,
        z_end=z_end,
        images=images,
        n_workers=n_workers,
        nuclei_df=nuclei_sorted,
        per_image_df=per_image_sorted,
        spots_df=spots_sorted,
        morph_df=morph_sorted,
        thr_df=thr_sorted,
        run_cfg_flat=run_cfg_flat,
    )
    # Executive_Summary second
    exec_ws = wb.create_sheet("Executive_Summary", 1)
    _build_executive_summary(
        exec_ws,
        per_image_df=per_image_sorted,
        nuclei_df=nuclei_sorted,
        spots_df=spots_sorted,
    )
    # PI_Focus third (curated PI deliverable — see _build_pi_focus docstring)
    pi_ws = wb.create_sheet("PI_Focus", 2)
    _build_pi_focus(
        pi_ws,
        per_image_df=per_image_sorted,
        nuclei_df=nuclei_sorted,
        spots_df=spots_sorted,
    )
    # Comparison_Table fourth
    cmp_ws = wb.create_sheet("Comparison_Table", 3)
    _build_comparison_table(
        cmp_ws,
        per_image_df=per_image_sorted,
        nuclei_df=nuclei_sorted,
    )
    # Run_Config last
    rc_ws = wb.create_sheet("Run_Config")
    _build_run_config_sheet(rc_ws, run_cfg_flat)
    # Move Run_Config to the end (after Thresholds)
    sheets = wb.sheetnames
    desired_order = [
        "README", "Executive_Summary", "PI_Focus", "Comparison_Table",
        "Per_Image_Summary", "Per_Nucleus_Metrics", "Per_Spot_Metrics",
        "Cell_Morphology", "Thresholds", "Run_Config",
    ]
    # Reorder
    wb._sheets = [wb[n] for n in desired_order if n in sheets]

    # Format data sheets
    for sheet_name, df in [
        ("Per_Image_Summary", per_image_sorted),
        ("Per_Nucleus_Metrics", nuclei_sorted),
        ("Per_Spot_Metrics", spots_sorted),
        ("Cell_Morphology", morph_sorted),
        ("Thresholds", thr_sorted),
    ]:
        if sheet_name not in wb.sheetnames or not len(df):
            continue
        _format_data_sheet(wb, sheet_name, df)

    # 2026-05-19 Brian: post-build label substitution. Walk the README +
    # Executive_Summary + Comparison_Table cell values and remap
    # "RNA1"/"RNA2"/"DAPI" generic tokens to the preset labels. Data sheets
    # are skipped — their column NAMES are the contract with downstream tooling
    # and must not be renamed. Cell VALUES on data sheets are numeric / image
    # filenames / condition strings and contain no channel tokens.
    _labels = _resolve_labels_from_run_cfg(run_cfg_flat)
    for sheet_name in ("README", "Executive_Summary", "PI_Focus", "Comparison_Table"):
        if sheet_name in wb.sheetnames:
            _relabel_worksheet(wb[sheet_name], _labels)

    wb.save(out_path)
    return fallback_cols


# ---------------------------------------------------------------------------
# Public entry point — analysis_raw_data.xlsx
# ---------------------------------------------------------------------------

def write_raw_data_workbook(
    *,
    out_path: Path,
    per_image_df: pd.DataFrame,
    nuclei_df: pd.DataFrame,
    spots_df: pd.DataFrame,
    morph_df: pd.DataFrame,
    fishsuite_version: str,
    run_start_utc: str,
    output_dir: Path,
) -> None:
    """Write the companion raw-data workbook.

    Same 4 data sheets (Per_Image_Summary, Per_Nucleus_Metrics, Per_Spot_Metrics,
    Cell_Morphology) sorted + condition-colored; plus a single Raw_README sheet
    pointing back to ``analysis_summary.xlsx``.
    """
    out_path = Path(out_path)

    per_image_sorted = _sort_df_by_condition(per_image_df, extra_sort_cols=["image"])
    nuclei_sorted = _sort_df_by_condition(nuclei_df, extra_sort_cols=["image", "nucleus_id"])
    spots_sorted = _sort_df_by_condition(spots_df, extra_sort_cols=["image", "channel", "spot_id"])
    morph_sorted = _sort_df_by_condition(morph_df, extra_sort_cols=["image", "nucleus_id"])

    with pd.ExcelWriter(out_path, engine="openpyxl") as xl:
        pd.DataFrame({"_stub": [""]}).to_excel(xl, sheet_name="Raw_README", index=False)
        per_image_sorted.to_excel(xl, sheet_name="Per_Image_Summary", index=False)
        if len(nuclei_sorted):
            nuclei_sorted.to_excel(xl, sheet_name="Per_Nucleus_Metrics", index=False)
        if len(spots_sorted):
            spots_sorted.to_excel(xl, sheet_name="Per_Spot_Metrics", index=False)
        if len(morph_sorted):
            morph_sorted.to_excel(xl, sheet_name="Cell_Morphology", index=False)

    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    wb = load_workbook(out_path)

    # Rebuild Raw_README
    if "Raw_README" in wb.sheetnames:
        del wb["Raw_README"]
    ws = wb.create_sheet("Raw_README", 0)
    BOLD = Font(bold=True, size=11)
    H1 = Font(bold=True, size=14, color="FFFFFFFF")
    SECTION_FILL = PatternFill("solid", fgColor=SECTION_FILL_HEX)
    TABLE_HDR_FILL = PatternFill("solid", fgColor=HEADER_FILL_HEX)
    THIN = Side(border_style="thin", color="FF999999")
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    WRAP = Alignment(wrap_text=True, vertical="top")
    LEFT_TOP = Alignment(horizontal="left", vertical="top", wrap_text=True)

    row = 1
    c = ws.cell(row=row, column=1, value="RAW DATA WORKBOOK (companion to analysis_summary.xlsx)")
    c.font = H1
    c.fill = SECTION_FILL
    for cc in range(2, 5):
        ws.cell(row=row, column=cc).fill = SECTION_FILL
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    ws.cell(row=row, column=1).alignment = LEFT_TOP
    ws.row_dimensions[row].height = 24
    row += 1

    def kv(k: str, v: Any) -> None:
        nonlocal row
        a = ws.cell(row=row, column=1, value=k)
        a.font = BOLD
        b = ws.cell(row=row, column=2, value=str(v))
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        b.alignment = LEFT_TOP
        row += 1

    kv("Purpose",
       "Sliceable raw data. Same per-image / per-nucleus / per-spot / "
       "cell-morphology data as analysis_summary.xlsx, with condition coloring "
       "and consistent sorting — but no pre-computed analysis. Use this for "
       "your own statistical re-analysis.")
    kv("For interpretation, see", "analysis_summary.xlsx in the same directory")
    kv("Run start (UTC)", run_start_utc)
    kv("fishsuite version", fishsuite_version)
    kv("Output directory", str(output_dir))
    row += 1
    ws.cell(row=row, column=1, value="Condition color key (column 'condition'):").font = BOLD
    row += 1
    for label, hex_color in [
        ("WT (or NT)",      "FFD7E9F7"),
        ("KO (or KD)",      "FFFCE4CC"),
        ("Sec-only / control", "FFE6E6E6"),
    ]:
        ws.cell(row=row, column=1, value=label).font = BOLD
        cell = ws.cell(row=row, column=2, value="(sample fill)")
        cell.fill = PatternFill("solid", fgColor=hex_color)
        cell.alignment = LEFT_TOP
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        row += 1
    row += 1

    ws.cell(row=row, column=1, value="Sheets in this workbook:").font = BOLD
    row += 1
    for i, h in enumerate(["Sheet", "Description"], start=1):
        c2 = ws.cell(row=row, column=i, value=h)
        c2.font = BOLD
        c2.fill = TABLE_HDR_FILL
        c2.border = BORDER
        c2.alignment = LEFT_TOP
    row += 1
    sheet_descs = [
        ("Per_Image_Summary",
         "One row per image. All FOV-level aggregates."),
        ("Per_Nucleus_Metrics",
         "One row per nucleus. Per-cell counts, intensities, paired counts, NN distances."),
        ("Per_Spot_Metrics",
         "One row per detected spot. xy positions, intensities, sizes, flags."),
        ("Cell_Morphology",
         "One row per nucleus. Nuclear shape descriptors."),
    ]
    for nm, desc in sheet_descs:
        a = ws.cell(row=row, column=1, value=nm)
        a.border = BORDER
        a.alignment = WRAP
        a.font = BOLD
        b = ws.cell(row=row, column=2, value=desc)
        b.border = BORDER
        b.alignment = WRAP
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        row += 1

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 24
    ws.column_dimensions["D"].width = 60
    ws.freeze_panes = "A2"

    # Move Raw_README to first position (create_sheet puts it last initially in some cases)
    desired = ["Raw_README", "Per_Image_Summary", "Per_Nucleus_Metrics",
               "Per_Spot_Metrics", "Cell_Morphology"]
    wb._sheets = [wb[n] for n in desired if n in wb.sheetnames]

    # Format data sheets
    for sheet_name, df in [
        ("Per_Image_Summary", per_image_sorted),
        ("Per_Nucleus_Metrics", nuclei_sorted),
        ("Per_Spot_Metrics", spots_sorted),
        ("Cell_Morphology", morph_sorted),
    ]:
        if sheet_name not in wb.sheetnames or not len(df):
            continue
        _format_data_sheet(wb, sheet_name, df)

    wb.save(out_path)
