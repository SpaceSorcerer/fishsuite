"""singlecell -- CPU single-cell (per-nucleus) treatment analysis of a run.

A POST-RUN utility (like ``coloc_backfill`` / ``walkthrough_figure``): it reads a
COMPLETED run's ``nuclei_metrics.csv`` and, for every meaningful per-nucleus
metric, computes the single-cell treatment picture WITHOUT touching the GPU or
re-reading any image:

  1. DOSE-RESPONSE vs a per-nucleus ABUNDANCE axis (default: ``nuclear_spot_count``
     = the primary-FISH / RNA1 nuclear spot count). Spearman rho + p, pooled and
     split per group, plus binned means. "Does the metric track abundance within
     single cells?"
  2. MATCHED-ABUNDANCE comparison of the two main conditions (control vs
     perturbation) WITHIN each abundance bin (Mann-Whitney). "Is a difference
     driven by abundance level, or is there a genuine treatment effect once
     abundance is matched?"
  3. GROUP (condition-depth) HETEROGENEITY: the perturbation group split by its
     OWN abundance tertile (Kruskal-Wallis across tertiles) -- reveals
     subpopulation structure (e.g. strong- vs weak-knockdown nuclei).
  4. DISTRIBUTION shape per condition + a per-REPLICATE (biological-replicate)
     Welch t-test on the group means.

If the run carries the rotation-null association columns (the density-corrected
proper-background coloc metric -- see ``coloc_backfill`` / rna_rna), an extra
SATURATION headline is emitted: whether the associated-partner COUNT scales with
abundance while the associated FRACTION stays flat (proportional association) or
the FRACTION rises (saturation).

GENERIC: nothing is hardcoded to MIAT/QKI. The abundance axis, the two groups to
compare, and how condition labels map to (group, biological-replicate) are all
parameterized with sensible auto-detected defaults. Column labels are prettified
using the run's own channel labels.

Note on the coloc metric: the rotation "proper background" null is the
DOCUMENTED CANONICAL coloc statistic for this pipeline (density-corrected;
retains each nucleus's own spot constellation) -- see POSTRUN_UTILITIES.md.

Outputs (written under ``<run>/deliverables/singlecell/`` by default):
  * ``single_cell_analysis.xlsx``  -- explorable workbook (How_to_read first)
  * ``figures/*.png``              -- locked-style NT-vs-perturbation SuperPlots,
                                      Okabe-Ito dose scatters + matched lines
  * ``SINGLE_CELL_FINDINGS.md``    -- plain-language readout

CLI::

    python -m fishsuite.core.singlecell --run-dir <run> [--abundance-col ...]
        [--group-a NT --group-b KD] [--no-figures] [--no-excel] [--out-subdir ...]
"""
from __future__ import annotations

import argparse
import json
import os
import re
import sys
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import numpy as np
import pandas as pd

# ---------------------------------------------------------------------------
# Column classification
# ---------------------------------------------------------------------------
# Columns that are structural/identifiers, never treated as a "metric".
_STRUCTURAL_COLS = {
    "image", "condition", "secondary_only", "experiment_id", "nucleus_id",
    "z_mode", "z_range", "cyto_estimation_method", "n_voxels", "n_pix",
    "n_z_slices", "voxel_xy_um", "voxel_z_um", "single_slice",
    "_group", "_rep", "_abund", "_v", "_abin",
}

# Nicer replacements for channel-role tokens embedded in column names. The
# run's actual labels (e.g. MIAT-561 / QKI-640) override these at runtime.
_ROLE_TOKENS = ("rna1", "rna2", "protein", "antibody", "rna", "dapi")


def _spearman(x, y) -> Tuple[float, float, int]:
    from scipy import stats
    x = pd.to_numeric(x, errors="coerce")
    y = pd.to_numeric(y, errors="coerce")
    m = x.notna() & y.notna()
    if int(m.sum()) < 8:
        return (float("nan"), float("nan"), int(m.sum()))
    r, p = stats.spearmanr(x[m], y[m])
    return (float(r), float(p), int(m.sum()))


def _prettify(col: str, chan_labels: Dict[str, str]) -> str:
    """Human label for a metric column, substituting the run's channel labels
    for role tokens (rna1/rna/protein/antibody/dapi)."""
    s = col
    for tok in _ROLE_TOKENS:
        lab = chan_labels.get(tok)
        if lab:
            s = re.sub(rf"\b{tok}\b", lab, s)
    s = s.replace("_", " ").strip()
    return s[:1].upper() + s[1:] if s else col


def _channel_labels_from_config(rc: dict) -> Dict[str, str]:
    """Pull role -> display-label from run_config channels (best effort)."""
    ch = (rc.get("config_resolved", {}) or {}).get("channels", {}) or {}
    out: Dict[str, str] = {}
    for role, key in (("rna", "rna_label"), ("rna1", "rna_label"),
                      ("rna2", "rna2_label"), ("protein", "antibody_label"),
                      ("antibody", "antibody_label"), ("dapi", "dapi_label")):
        v = ch.get(key)
        if isinstance(v, str) and v and v.lower() not in ("rna1", "rna2",
                                                          "protein", "protein2"):
            out[role] = v
    return out


# ---------------------------------------------------------------------------
# (group, replicate) parsing -- generic
# ---------------------------------------------------------------------------
_REP_SUFFIX = re.compile(r"^(?P<group>.+?)[_\-\s](?P<rep>w\d+|rep\d+|r\d+|\d+)$",
                         re.IGNORECASE)


def _parse_group_replicate(df: pd.DataFrame, condition_col: str
                           ) -> Tuple[pd.Series, pd.Series]:
    """Return (group, replicate) series.

    Default scheme: if a condition looks like ``<group><sep><rep>`` (e.g.
    ``NT_w1``) split it -> group=NT, rep=w1. Otherwise group=condition and
    rep=image (the locked convention where each image is a biological
    replicate). Decided per-row so mixed schemes still work.
    """
    groups = []
    reps = []
    for _, row in df.iterrows():
        cond = str(row[condition_col])
        m = _REP_SUFFIX.match(cond)
        if m:
            groups.append(m.group("group"))
            reps.append(m.group("rep"))
        else:
            groups.append(cond)
            reps.append(str(row.get("image", cond)))
    return pd.Series(groups, index=df.index), pd.Series(reps, index=df.index)


def _pick_abundance_col(df: pd.DataFrame, requested: Optional[str]) -> str:
    if requested:
        if requested not in df.columns:
            raise ValueError(
                f"--abundance-col {requested!r} not found in nuclei_metrics.csv. "
                f"Available spot/intensity columns include: "
                f"{[c for c in df.columns if 'spot_count' in c or 'intensity' in c][:8]}"
            )
        return requested
    for cand in ("nuclear_spot_count", "rna_spot_count", "sum_rna_intensity",
                 "nuclear_total_intensity_rna1"):
        if cand in df.columns:
            return cand
    raise ValueError(
        "Could not auto-pick an abundance axis (no nuclear_spot_count / "
        "rna_spot_count / sum_rna_intensity in nuclei_metrics.csv). Pass "
        "--abundance-col explicitly."
    )


def _pick_two_groups(groups_present: List[str], group_a: Optional[str],
                     group_b: Optional[str], counts: pd.Series
                     ) -> Tuple[Optional[str], Optional[str]]:
    """Resolve (control, perturbation) group labels for the matched comparison."""
    from ._superplot import is_control_like
    if group_a and group_b:
        return group_a, group_b
    gp = list(groups_present)
    if len(gp) == 0:
        return None, None
    if len(gp) == 1:
        return gp[0], None
    ctrl = [g for g in gp if is_control_like(g)]
    pert = [g for g in gp if not is_control_like(g)]
    a = group_a or (ctrl[0] if ctrl else None)
    b = group_b or None
    if a is None or b is None:
        # rank by nucleus count; control-like preferred as A
        ranked = sorted(gp, key=lambda g: -int(counts.get(g, 0)))
        if a is None:
            a = ctrl[0] if ctrl else ranked[0]
        remaining = [g for g in ranked if g != a]
        if b is None:
            b = (pert[0] if pert and pert[0] != a else
                 (remaining[0] if remaining else None))
    return a, b


def _make_bins(abund: pd.Series) -> Tuple[List[float], List[str], List[float]]:
    """Return (bin_edges, bin_labels, bin_centers) for the abundance axis.

    Integer-ish small-range axes (spot counts) get the canonical count bins
    (1-2/3-5/6-10/11-20/21-40/40+); anything else gets 5 quantile bins."""
    a = pd.to_numeric(abund, errors="coerce").dropna()
    a = a[a > 0]
    looks_count = (a.max() <= 200) and np.allclose(a, a.round())
    if looks_count and len(a):
        edges = [0, 2, 5, 10, 20, 40, 10 ** 9]
        labels = ["1-2", "3-5", "6-10", "11-20", "21-40", "40+"]
        centers = [1.5, 4, 8, 15, 30, 47]
        return edges, labels, centers
    # quantile bins
    try:
        qs = np.unique(np.nanquantile(a, [0, .2, .4, .6, .8, 1.0]))
    except Exception:
        qs = np.array([a.min(), a.max()]) if len(a) else np.array([0.0, 1.0])
    if len(qs) < 3:
        qs = np.linspace(a.min() if len(a) else 0.0, a.max() if len(a) else 1.0, 4)
    edges = list(qs)
    labels = [f"{edges[i]:.3g}-{edges[i+1]:.3g}" for i in range(len(edges) - 1)]
    centers = [0.5 * (edges[i] + edges[i + 1]) for i in range(len(edges) - 1)]
    return edges, labels, centers


# ---------------------------------------------------------------------------
# CORE computation -- returns a results dict (pure; no I/O)
# ---------------------------------------------------------------------------
def compute_singlecell(
    nm: pd.DataFrame,
    *,
    abundance_col: Optional[str] = None,
    condition_col: str = "condition",
    group_a: Optional[str] = None,
    group_b: Optional[str] = None,
    exclude_secondary: bool = True,
    metrics: Optional[List[str]] = None,
    chan_labels: Optional[Dict[str, str]] = None,
) -> Dict[str, Any]:
    """Compute the full single-cell analysis from a per-nucleus table."""
    from scipy import stats
    chan_labels = chan_labels or {}
    nm = nm.copy()

    # drop secondary-only nuclei (generic: the boolean column, else "sec" label)
    if exclude_secondary:
        if "secondary_only" in nm.columns:
            nm = nm[~nm["secondary_only"].astype(bool)]
        else:
            nm = nm[~nm[condition_col].astype(str).str.contains("sec", case=False)]
    nm = nm.reset_index(drop=True)
    if nm.empty:
        raise ValueError("no non-secondary nuclei to analyze in nuclei_metrics.csv")

    g, rep = _parse_group_replicate(nm, condition_col)
    nm["_group"] = g.values
    nm["_rep"] = rep.values
    abundance_col = _pick_abundance_col(nm, abundance_col)
    nm["_abund"] = pd.to_numeric(nm[abundance_col], errors="coerce")

    counts = nm["_group"].value_counts()
    groups_present = list(counts.index)
    gA, gB = _pick_two_groups(groups_present, group_a, group_b, counts)

    # derived association-count column when rotation assoc fraction is present
    has_rot = ("protein_rotation_assoc_fraction_at_rna1_spots" in nm.columns
               and "rotation_null_usable" in nm.columns)
    if has_rot:
        nm["assoc_count_rotation"] = (
            pd.to_numeric(nm["protein_rotation_assoc_fraction_at_rna1_spots"],
                          errors="coerce") * nm["_abund"])

    edges, blab, bctr = _make_bins(nm["_abund"])
    nm["_abin"] = pd.cut(nm["_abund"], bins=edges, labels=blab)

    # metric column selection (auto-discover numeric, minus structural)
    if metrics:
        metric_cols = [c for c in metrics if c in nm.columns]
    else:
        metric_cols = []
        for c in nm.columns:
            if c in _STRUCTURAL_COLS or c == abundance_col:
                continue
            if pd.api.types.is_bool_dtype(nm[c]):
                continue  # boolean flags (e.g. rotation_null_usable) are not metrics
            if pd.api.types.is_numeric_dtype(nm[c]) or c == "assoc_count_rotation":
                s = pd.to_numeric(nm[c], errors="coerce")
                if s.notna().sum() >= 8 and s.nunique(dropna=True) > 1:
                    metric_cols.append(c)
    # always include the abundance axis itself (self dose skipped)
    if abundance_col not in metric_cols:
        metric_cols = [abundance_col] + metric_cols

    rot_gated = {"protein_rotation_enrichment_at_rna1_spots",
                 "protein_rotation_assoc_fraction_at_rna1_spots",
                 "protein_rotation_null_z_at_rna1_spots",
                 "protein_rotation_null_p_at_rna1_spots",
                 "assoc_count_rotation"}
    pct_cols = {c for c in metric_cols if "fraction" in c.lower()
                or c.endswith("_frac") or "nc_ratio" not in c and False}

    results: Dict[str, Any] = {}
    for col in metric_cols:
        d = nm.copy()
        needs_rot = col in rot_gated
        if needs_rot and "rotation_null_usable" in d.columns:
            d = d[d["rotation_null_usable"] == True]  # noqa: E712
        d["_v"] = pd.to_numeric(d[col], errors="coerce")
        d = d[d["_v"].notna()]
        if d.empty:
            continue
        is_self = (col == abundance_col)
        rec: Dict[str, Any] = {
            "label": _prettify(col, chan_labels), "column": col,
            "needs_rot": needs_rot, "pct": col in pct_cols,
            "n_used": int(len(d)),
        }
        # (1) dose response
        rec["rho_pool"] = None if is_self else _spearman(d["_abund"], d["_v"])
        rec["rho_A"] = (None if (is_self or gA is None)
                        else _spearman(d[d._group == gA]["_abund"], d[d._group == gA]["_v"]))
        rec["rho_B"] = (None if (is_self or gB is None)
                        else _spearman(d[d._group == gB]["_abund"], d[d._group == gB]["_v"]))
        # binned means pooled/A/B
        binned = []
        for lab in blab:
            row = {"bin": lab}
            for gkey, sub in (("pool", d),
                              ("A", d[d._group == gA] if gA else d.iloc[0:0]),
                              ("B", d[d._group == gB] if gB else d.iloc[0:0])):
                b = sub[sub._abin == lab]
                row[f"{gkey}_n"] = int(len(b))
                row[f"{gkey}_mean"] = float(b._v.mean()) if len(b) else float("nan")
                row[f"{gkey}_sem"] = float(b._v.sem()) if len(b) > 1 else float("nan")
            binned.append(row)
        rec["binned"] = binned
        # (2) matched-abundance A vs B per bin
        matched = []
        for lab in blab:
            a_ = d[(d._abin == lab) & (d._group == gA)]._v.dropna() if gA else pd.Series([], dtype=float)
            b_ = d[(d._abin == lab) & (d._group == gB)]._v.dropna() if gB else pd.Series([], dtype=float)
            p = float("nan")
            if len(a_) >= 5 and len(b_) >= 5:
                try:
                    p = float(stats.mannwhitneyu(a_, b_).pvalue)
                except Exception:
                    p = float("nan")
            matched.append({"bin": lab, "A_n": int(len(a_)), "B_n": int(len(b_)),
                            "A_mean": float(a_.mean()) if len(a_) else float("nan"),
                            "B_mean": float(b_.mean()) if len(b_) else float("nan"),
                            "p_mwu": p})
        rec["matched"] = matched
        # (3) perturbation-group heterogeneity by own abundance tertile
        tert = []
        kruskal_p = float("nan")
        if gB is not None:
            pert = d[(d._group == gB) & (d._abund >= 1)].copy()
            if len(pert) >= 9:
                try:
                    pert["_tert"] = pd.qcut(pert._abund, 3,
                                            labels=["low", "mid", "high"],
                                            duplicates="drop")
                except Exception:
                    pert["_tert"] = np.nan
                for t in ("low", "mid", "high"):
                    bb = pert[pert._tert == t]
                    if len(bb):
                        tert.append({"tertile": t, "n": int(len(bb)),
                                     "abund_median": float(bb._abund.median()),
                                     "mean": float(bb._v.mean()),
                                     "sem": float(bb._v.sem()) if len(bb) > 1 else float("nan")})
                grps = [pert[pert._tert == t]._v.dropna() for t in pert._tert.dropna().unique()]
                grps = [x for x in grps if len(x) >= 3]
                if len(grps) >= 2:
                    try:
                        kruskal_p = float(stats.kruskal(*grps).pvalue)
                    except Exception:
                        kruskal_p = float("nan")
        rec["tertiles"] = tert
        rec["kruskal_p"] = kruskal_p
        # (4) distribution + per-replicate Welch
        dist: Dict[str, Any] = {}
        for gkey, gname in (("A", gA), ("B", gB)):
            v = d[d._group == gname]._v.dropna() if gname else pd.Series([], dtype=float)
            dist[gkey] = {
                "group": gname, "n": int(len(v)),
                "mean": float(v.mean()) if len(v) else float("nan"),
                "median": float(v.median()) if len(v) else float("nan"),
                "std": float(v.std()) if len(v) > 1 else float("nan"),
                "cv": float(v.std() / v.mean()) if len(v) > 1 and v.mean() else float("nan"),
                "skew": float(stats.skew(np.asarray(v, dtype=float))) if len(v) > 2 else float("nan"),
            }
        repm = d.groupby("_rep")._v.mean()
        repA = repm[[r for r in repm.index if r in set(d[d._group == gA]._rep)]].values if gA else np.array([])
        repB = repm[[r for r in repm.index if r in set(d[d._group == gB]._rep)]].values if gB else np.array([])
        welch_p = float("nan")
        if len(repA) >= 2 and len(repB) >= 2:
            try:
                welch_p = float(stats.ttest_ind(repA, repB, equal_var=False).pvalue)
            except Exception:
                welch_p = float("nan")
        dist["rep_A_mean"] = float(np.nanmean(repA)) if len(repA) else float("nan")
        dist["rep_B_mean"] = float(np.nanmean(repB)) if len(repB) else float("nan")
        dist["rep_pct_chg"] = (float(100 * (np.nanmean(repB) - np.nanmean(repA)) / np.nanmean(repA))
                               if len(repA) and len(repB) and np.nanmean(repA) else float("nan"))
        dist["rep_welch_p"] = welch_p
        dist["n_reps_A"] = int(len(repA))
        dist["n_reps_B"] = int(len(repB))
        rec["dist"] = dist
        results[col] = rec

    # ---- saturation headline (rotation assoc, when present) ----
    if has_rot:
        u = nm[(nm["rotation_null_usable"] == True)  # noqa: E712
               & nm["assoc_count_rotation"].notna() & (nm["_abund"] >= 1)].copy()
        af = pd.to_numeric(u["protein_rotation_assoc_fraction_at_rna1_spots"], errors="coerce")
        u3 = u[u["_abund"] >= 3]
        af3 = pd.to_numeric(u3["protein_rotation_assoc_fraction_at_rna1_spots"], errors="coerce")
        results["_saturation"] = {
            "n_usable": int(len(u)),
            "rho_assoc_count_vs_abund": _spearman(u["_abund"], u["assoc_count_rotation"]),
            "rho_assoc_frac_vs_abund": _spearman(u["_abund"], af),
            "rho_assoc_frac_vs_abund_ge3": _spearman(u3["_abund"], af3),
            "rho_assoc_count_vs_abund_ge3": _spearman(u3["_abund"], u3["assoc_count_rotation"]),
        }

    results["_meta"] = {
        "abundance_col": abundance_col, "group_A": gA, "group_B": gB,
        "groups_present": groups_present, "n_nuclei": int(len(nm)),
        "n_A": int((nm._group == gA).sum()) if gA else 0,
        "n_B": int((nm._group == gB).sum()) if gB else 0,
        "bin_labels": blab, "bin_centers": bctr,
        "metric_cols": metric_cols, "has_rotation": has_rot,
        "chan_labels": chan_labels,
    }
    return results


# ---------------------------------------------------------------------------
# Excel
# ---------------------------------------------------------------------------
def _write_excel(results: Dict[str, Any], xlsx: Path) -> None:
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    meta = results["_meta"]
    order = [c for c in results if not c.startswith("_")]

    def rho3(x):
        if x is None or x[0] is None or (isinstance(x[0], float) and np.isnan(x[0])):
            return (None, None, 0)
        return (round(x[0], 3), x[1], x[2])

    dr_rows, db_rows, mm_rows, het_rows, di_rows = [], [], [], [], []
    for c in order:
        r = results[c]
        rp, ra, rb = rho3(r.get("rho_pool")), rho3(r.get("rho_A")), rho3(r.get("rho_B"))
        dr_rows.append({"metric": r["label"], "column": c, "n_nuclei_used": r["n_used"],
                        "rho_pooled": rp[0], "p_pooled": rp[1], "n_pooled": rp[2],
                        "rho_A": ra[0], "p_A": ra[1], "rho_B": rb[0], "p_B": rb[1]})
        for gkey in ("pool", "A", "B"):
            row = {"metric": r["label"], "group": gkey}
            for b in r["binned"]:
                row[f"{b['bin']}_n"] = b[f"{gkey}_n"]
                mv = b[f"{gkey}_mean"]
                row[f"{b['bin']}_mean"] = None if (mv is None or np.isnan(mv)) else round(mv, 5)
            db_rows.append(row)
        for m in r["matched"]:
            p = m["p_mwu"]
            sig = "" if (p is None or np.isnan(p)) else ("***" if p < 0.001 else "**" if p < 0.01 else "*" if p < 0.05 else "ns")
            mm_rows.append({"metric": r["label"], "abundance_bin": m["bin"], "n_A": m["A_n"], "n_B": m["B_n"],
                            "A_mean": None if np.isnan(m["A_mean"]) else round(m["A_mean"], 5),
                            "B_mean": None if np.isnan(m["B_mean"]) else round(m["B_mean"], 5),
                            "p_MannWhitney": None if (p is None or np.isnan(p)) else p, "sig": sig})
        for t in r["tertiles"]:
            het_rows.append({"metric": r["label"], "perturbation_tertile": t["tertile"], "n_nuclei": t["n"],
                             "abundance_median": round(t["abund_median"], 2), "metric_mean": round(t["mean"], 5),
                             "Kruskal_p": None if np.isnan(r["kruskal_p"]) else r["kruskal_p"]})
        dd = r["dist"]
        for gkey in ("A", "B"):
            s = dd[gkey]
            di_rows.append({"metric": r["label"], "group_slot": gkey, "group": s["group"], "n_nuclei": s["n"],
                            "mean": None if np.isnan(s["mean"]) else round(s["mean"], 5),
                            "median": None if np.isnan(s["median"]) else round(s["median"], 5),
                            "std": None if np.isnan(s["std"]) else round(s["std"], 5),
                            "CV": None if np.isnan(s["cv"]) else round(s["cv"], 3),
                            "skew": None if np.isnan(s["skew"]) else round(s["skew"], 3),
                            "rep_mean_A": None if np.isnan(dd["rep_A_mean"]) else round(dd["rep_A_mean"], 5),
                            "rep_mean_B": None if np.isnan(dd["rep_B_mean"]) else round(dd["rep_B_mean"], 5),
                            "pct_change_BvsA": None if np.isnan(dd["rep_pct_chg"]) else round(dd["rep_pct_chg"], 1),
                            "rep_Welch_p": dd["rep_welch_p"], "n_reps_A": dd["n_reps_A"], "n_reps_B": dd["n_reps_B"]})

    dr = pd.DataFrame(dr_rows)
    db = pd.DataFrame(db_rows)
    mm = pd.DataFrame(mm_rows)
    het = pd.DataFrame(het_rows)
    di = pd.DataFrame(di_rows)

    sat_df = None
    if "_saturation" in results:
        s = results["_saturation"]

        def st(k):
            v = s[k]
            return (round(v[0], 3) if v[0] == v[0] else None, v[1], v[2])
        sat_df = pd.DataFrame([
            {"test": "associated-partner COUNT vs abundance", "rho": st("rho_assoc_count_vs_abund")[0],
             "p": st("rho_assoc_count_vs_abund")[1], "n": st("rho_assoc_count_vs_abund")[2],
             "reads_as": "strong positive => COUNT scales with abundance (no plateau)"},
            {"test": "associated-partner FRACTION vs abundance", "rho": st("rho_assoc_frac_vs_abund")[0],
             "p": st("rho_assoc_frac_vs_abund")[1], "n": st("rho_assoc_frac_vs_abund")[2],
             "reads_as": "near-zero => FRACTION ~ constant (proportional, not saturating)"},
            {"test": "FRACTION vs abundance (abundance>=3 guard)", "rho": st("rho_assoc_frac_vs_abund_ge3")[0],
             "p": st("rho_assoc_frac_vs_abund_ge3")[1], "n": st("rho_assoc_frac_vs_abund_ge3")[2],
             "reads_as": "flat => proportional association"},
            {"test": "COUNT vs abundance (abundance>=3 guard)", "rho": st("rho_assoc_count_vs_abund_ge3")[0],
             "p": st("rho_assoc_count_vs_abund_ge3")[1], "n": st("rho_assoc_count_vs_abund_ge3")[2],
             "reads_as": "positive => COUNT keeps scaling"},
        ])

    how = pd.DataFrame({"Single-cell (per-nucleus) treatment analysis": [
        f"Abundance axis = {meta['abundance_col']}. Groups compared: A={meta['group_A']} (control-slot) vs B={meta['group_B']} (perturbation-slot).",
        f"{meta['n_nuclei']} nuclei analyzed (A n={meta['n_A']} / B n={meta['n_B']}); secondary-only excluded. Each biological replicate = one well/image.",
        "",
        "SHEET GUIDE:",
        "  Dose_response   : Spearman rho of each metric vs the per-nucleus abundance axis (pooled + per group). |rho| large = the metric tracks abundance within single cells.",
        "  Dose_binned     : mean of each metric within abundance bins (pooled + per group), with per-bin n.",
        "  Matched_abund   : within each abundance bin, group A vs B (Mann-Whitney). p>0.05 across all bins = abundance-LEVEL-driven; a significant bin = a treatment effect beyond abundance level.",
        "  Heterogeneity   : perturbation-group nuclei split by their OWN abundance tertile (low/mid/high); Kruskal-Wallis across tertiles.",
        "  Distribution    : per-group shape (mean/median/std/CV/skew) + per-REPLICATE (biological) means, %change and Welch t.",
        "  Saturation      : (coloc runs) associated-partner COUNT scales with abundance (rho>0) while FRACTION stays flat (rho~0) => proportional association, not saturation.",
        "",
        "STAT KEY: rho = Spearman. p two-sided. sig: *** p<0.001, ** p<0.01, * p<0.05, ns = n.s.",
        "The rotation 'proper background' null is the documented canonical coloc metric (density-corrected). All numbers read directly from nuclei_metrics.csv.",
    ]})

    sheets = [("How_to_read", how), ("Dose_response", dr), ("Dose_binned", db),
              ("Matched_abund", mm), ("Heterogeneity", het), ("Distribution", di)]
    if sat_df is not None:
        sheets.append(("Saturation", sat_df))

    xlsx.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(xlsx, engine="openpyxl") as xw:
        for name, dfx in sheets:
            (dfx if not dfx.empty else pd.DataFrame({"(no rows)": []})).to_excel(
                xw, sheet_name=name, index=False)

    wb = load_workbook(xlsx)
    hdr_fill = PatternFill("solid", fgColor="1F4E79")
    hdr_font = Font(bold=True, color="FFFFFF")
    for ws in wb.worksheets:
        if ws.title == "How_to_read":
            ws.column_dimensions["A"].width = 150
            ws["A1"].font = Font(bold=True, size=12, color="1F4E79")
            ws.sheet_view.showGridLines = False
            continue
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for j in range(1, ws.max_column + 1):
            L = get_column_letter(j)
            header = str(ws.cell(row=1, column=j).value or "")
            ws.column_dimensions[L].width = 42 if header in ("metric",) else (46 if header in ("reads_as", "test") else 13)
            if header.startswith("p_") or header in ("p", "p_MannWhitney", "rep_Welch_p", "Kruskal_p"):
                for rr in range(2, ws.max_row + 1):
                    cell = ws.cell(row=rr, column=j)
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = "0.00E+00"
    wb.save(xlsx)


# ---------------------------------------------------------------------------
# Figures
# ---------------------------------------------------------------------------
def _make_figures(nm_prepared: pd.DataFrame, results: Dict[str, Any],
                  fig_dir: Path, *, seed: int = 0) -> List[str]:
    """Locked-style SuperPlots for the top metrics + Okabe-Ito dose scatters +
    matched lines + (if present) a saturation composite. Returns written paths."""
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    from scipy import stats
    from ._superplot import (superplot_into_axes, CONTROL_COLOR, PERTURB_COLOR,
                             order_conditions_control_first)

    meta = results["_meta"]
    gA, gB = meta["group_A"], meta["group_B"]
    blab, bctr = meta["bin_labels"], meta["bin_centers"]
    fig_dir.mkdir(parents=True, exist_ok=True)
    written: List[str] = []
    cond_order = order_conditions_control_first([g for g in (gA, gB) if g])

    # rank metrics by |pooled dose rho| for the SuperPlot / scatter selection
    ranked = []
    for c, r in results.items():
        if c.startswith("_"):
            continue
        rp = r.get("rho_pool")
        rho = abs(rp[0]) if (rp and rp[0] == rp[0]) else -1.0
        ranked.append((rho, c, r))
    ranked.sort(key=lambda t: -t[0])
    top_cols = [c for _, c, _ in ranked[:8]]

    # ---- SuperPlots (locked drawer) ----
    for i, col in enumerate(top_cols, 1):
        r = results[col]
        sub = nm_prepared[["_group", "_rep", col]].copy()
        sub = sub[sub["_group"].isin([g for g in (gA, gB) if g])]
        sub = sub.rename(columns={"_group": "condition", "_rep": "image"})
        sub[col] = pd.to_numeric(sub[col], errors="coerce")
        sub = sub.dropna(subset=[col])
        if sub.empty:
            continue
        fig, ax = plt.subplots(figsize=(4.4, 4.8))
        ok = superplot_into_axes(ax, sub, col, ylabel=r["label"], unit="nucleus",
                                 pct=r["pct"], condition_order=cond_order)
        wp = r["dist"]["rep_welch_p"]
        ax.set_title(f"{r['label']}\nWelch p={wp:.3g} (replicate means)" if wp == wp else r["label"],
                     fontsize=8.5)
        fig.tight_layout()
        p = fig_dir / f"sp{i:02d}_{re.sub(r'[^A-Za-z0-9]+', '_', col)[:40]}.png"
        fig.savefig(p, dpi=600, bbox_inches="tight")
        plt.close(fig)
        if ok:
            written.append(str(p))

    # ---- dose-response scatters (top 6) ----
    for i, col in enumerate(top_cols[:6], 1):
        r = results[col]
        d = nm_prepared.copy()
        if r["needs_rot"] and "rotation_null_usable" in d.columns:
            d = d[d["rotation_null_usable"] == True]  # noqa: E712
        d["_v"] = pd.to_numeric(d[col], errors="coerce")
        d = d[d["_v"].notna() & (d["_abund"] >= 1)]
        if d.empty:
            continue
        fig, ax = plt.subplots(figsize=(5.0, 4.4))
        for gname, c in ((gA, CONTROL_COLOR), (gB, PERTURB_COLOR)):
            if gname is None:
                continue
            s = d[d._group == gname]
            jx = np.random.RandomState(seed + 7).uniform(-0.25, 0.25, len(s))
            ax.scatter(s._abund + jx, s._v, s=10, alpha=0.28, color=c,
                       edgecolor="none", label=f"{gname} (n={len(s)})")
            mx, my, me = [], [], []
            for lab, ctr in zip(blab, bctr):
                b = s[s._abin == lab]._v if "_abin" in s else pd.Series([], dtype=float)
                if len(b) >= 3:
                    mx.append(ctr)
                    my.append(b.mean())
                    me.append(b.sem())
            if mx:
                ax.errorbar(mx, my, yerr=me, color=c, lw=2.2, marker="o", ms=6,
                            capsize=3, zorder=6)
        rr, pp, nn = _spearman(d._abund, d._v)
        ax.set_xlabel(f"Per-nucleus abundance ({meta['abundance_col']})")
        ax.set_ylabel(r["label"])
        ax.grid(alpha=0.25, ls="--")
        ax.set_axisbelow(True)
        ax.set_title(f"{r['label']}\nSpearman rho={rr:+.3f} (p={pp:.1e}, n={nn})", fontsize=9)
        ax.legend(frameon=False, fontsize=8)
        fig.tight_layout()
        p = fig_dir / f"dr{i:02d}_{re.sub(r'[^A-Za-z0-9]+', '_', col)[:40]}.png"
        fig.savefig(p, dpi=600, bbox_inches="tight")
        plt.close(fig)
        written.append(str(p))

    # ---- saturation composite (coloc runs) ----
    if "_saturation" in results and "assoc_count_rotation" in nm_prepared.columns:
        u = nm_prepared.copy()
        if "rotation_null_usable" in u.columns:
            u = u[u["rotation_null_usable"] == True]  # noqa: E712
        u = u[u["assoc_count_rotation"].notna() & (u["_abund"] >= 1)]
        fraccol = "protein_rotation_assoc_fraction_at_rna1_spots"
        if not u.empty and fraccol in u.columns:
            fig, axes = plt.subplots(1, 2, figsize=(9.4, 4.3))
            for ax, col, yl, mm, tt, ispct in (
                (axes[0], "assoc_count_rotation", "Associated-partner count / nucleus", 1, "COUNT scales with abundance", False),
                (axes[1], fraccol, "Associated-partner fraction", 3, "FRACTION flat vs abundance", True)):
                dd = u[u["_abund"] >= mm].copy()
                dd["_v"] = pd.to_numeric(dd[col], errors="coerce")
                for gname, c in ((gA, CONTROL_COLOR), (gB, PERTURB_COLOR)):
                    if gname is None:
                        continue
                    s = dd[dd._group == gname]
                    jx = np.random.RandomState(seed + 1).uniform(-0.25, 0.25, len(s))
                    ax.scatter(s._abund + jx, s._v, s=9, alpha=0.22, color=c, edgecolor="none", label=gname)
                    mx, my, me = [], [], []
                    for lab, ctr in zip(blab, bctr):
                        b = s[s._abin == lab]._v if "_abin" in s else pd.Series([], dtype=float)
                        if len(b) >= 3:
                            mx.append(ctr)
                            my.append(b.mean())
                            me.append(b.sem())
                    if mx:
                        ax.errorbar(mx, my, yerr=me, color=c, lw=2.2, marker="o", ms=6, capsize=3, zorder=6)
                rr, pp, _ = _spearman(dd._abund, dd._v)
                if ispct:
                    import matplotlib.ticker as mtick
                    ax.yaxis.set_major_formatter(mtick.PercentFormatter(xmax=1.0, decimals=0))
                ax.set_xlabel(f"Per-nucleus abundance ({meta['abundance_col']})")
                ax.set_ylabel(yl)
                ax.grid(alpha=0.25, ls="--")
                ax.set_axisbelow(True)
                ax.set_title(f"{tt}\nrho={rr:+.3f} (p={pp:.1e})", fontsize=9.5)
                ax.legend(frameon=False, fontsize=8)
            fig.suptitle("Single-cell saturation test", fontsize=11)
            fig.tight_layout(rect=[0, 0, 1, 0.96])
            p = fig_dir / "SAT00_saturation_composite.png"
            fig.savefig(p, dpi=600, bbox_inches="tight")
            plt.close(fig)
            written.append(str(p))
    return written


def _write_findings(results: Dict[str, Any], md_path: Path, written_figs: List[str]) -> None:
    meta = results["_meta"]
    lines = ["# Single-cell (per-nucleus) treatment analysis", "",
             f"- Abundance axis: `{meta['abundance_col']}`",
             f"- Groups: A = `{meta['group_A']}` (control-slot) vs B = `{meta['group_B']}` (perturbation-slot)",
             f"- Nuclei analyzed: {meta['n_nuclei']} (A n={meta['n_A']} / B n={meta['n_B']}); secondary-only excluded",
             f"- Metrics evaluated: {len(meta['metric_cols'])}", ""]
    # top dose-response
    ranked = []
    for c, r in results.items():
        if c.startswith("_"):
            continue
        rp = r.get("rho_pool")
        if rp and rp[0] == rp[0]:
            ranked.append((abs(rp[0]), rp[0], rp[1], rp[2], r["label"]))
    ranked.sort(reverse=True)
    lines.append("## Strongest single-cell dose-responses (|Spearman rho|)")
    for a, rho, p, n, lab in ranked[:12]:
        lines.append(f"- rho={rho:+.3f} (p={p:.1e}, n={n}) — {lab}")
    lines.append("")
    if "_saturation" in results:
        s = results["_saturation"]
        lines += ["## Saturation headline (rotation proper-background association)",
                  f"- associated COUNT vs abundance: rho={s['rho_assoc_count_vs_abund'][0]:+.3f} (p={s['rho_assoc_count_vs_abund'][1]:.1e})",
                  f"- associated FRACTION vs abundance: rho={s['rho_assoc_frac_vs_abund'][0]:+.3f} (p={s['rho_assoc_frac_vs_abund'][1]:.1e})",
                  f"- FRACTION vs abundance (>=3 guard): rho={s['rho_assoc_frac_vs_abund_ge3'][0]:+.3f} (p={s['rho_assoc_frac_vs_abund_ge3'][1]:.1e})",
                  "- Reads as: COUNT scaling + FRACTION flat => proportional association (not saturation).", ""]
    lines.append(f"## Figures ({len(written_figs)})")
    for f in written_figs:
        lines.append(f"- {Path(f).name}")
    lines.append("")
    lines.append("_Numbers read directly from nuclei_metrics.csv; the rotation null is the documented canonical coloc metric._")
    md_path.parent.mkdir(parents=True, exist_ok=True)
    md_path.write_text("\n".join(lines), encoding="utf-8")


# ---------------------------------------------------------------------------
# I/O SHELL
# ---------------------------------------------------------------------------
def singlecell_run(
    run_dir,
    *,
    abundance_col: Optional[str] = None,
    condition_col: str = "condition",
    group_a: Optional[str] = None,
    group_b: Optional[str] = None,
    exclude_secondary: bool = True,
    metrics: Optional[List[str]] = None,
    do_excel: bool = True,
    do_figures: bool = True,
    out_subdir: str = "singlecell",
    seed: int = 0,
    verbose: bool = True,
) -> dict:
    """Run the single-cell analysis on a completed run directory.

    Reads ``<run>/nuclei_metrics.csv`` (+ run_config.json for channel labels),
    computes the dose-response / matched-abundance / heterogeneity / distribution
    / saturation results, and writes an Excel + SuperPlots + findings under
    ``<run>/deliverables/<out_subdir>/``. Returns a summary dict.
    """
    run_dir = Path(run_dir)
    rc_path = run_dir / "run_config.json"
    nm_path = run_dir / "nuclei_metrics.csv"
    if not rc_path.exists():
        raise FileNotFoundError(f"no run_config.json in {run_dir}")
    if not nm_path.exists():
        raise FileNotFoundError(f"run_dir missing nuclei_metrics.csv: {run_dir}")
    rc = json.loads(rc_path.read_text(encoding="utf-8"))
    chan_labels = _channel_labels_from_config(rc)

    nm = pd.read_csv(nm_path)
    results = compute_singlecell(
        nm, abundance_col=abundance_col, condition_col=condition_col,
        group_a=group_a, group_b=group_b, exclude_secondary=exclude_secondary,
        metrics=metrics, chan_labels=chan_labels,
    )

    # rebuild the prepared frame (mirrors compute) for figures
    nm_prep = nm.copy()
    if exclude_secondary:
        if "secondary_only" in nm_prep.columns:
            nm_prep = nm_prep[~nm_prep["secondary_only"].astype(bool)]
        else:
            nm_prep = nm_prep[~nm_prep[condition_col].astype(str).str.contains("sec", case=False)]
    nm_prep = nm_prep.reset_index(drop=True)
    g, rep = _parse_group_replicate(nm_prep, condition_col)
    nm_prep["_group"] = g.values
    nm_prep["_rep"] = rep.values
    abund_col = results["_meta"]["abundance_col"]
    nm_prep["_abund"] = pd.to_numeric(nm_prep[abund_col], errors="coerce")
    if "protein_rotation_assoc_fraction_at_rna1_spots" in nm_prep.columns:
        nm_prep["assoc_count_rotation"] = (
            pd.to_numeric(nm_prep["protein_rotation_assoc_fraction_at_rna1_spots"], errors="coerce")
            * nm_prep["_abund"])
    edges, blab, _ = _make_bins(nm_prep["_abund"])
    nm_prep["_abin"] = pd.cut(nm_prep["_abund"], bins=edges, labels=blab)

    out_dir = run_dir / "deliverables" / out_subdir
    out_dir.mkdir(parents=True, exist_ok=True)
    written: Dict[str, str] = {}
    if do_excel:
        xlsx = out_dir / "single_cell_analysis.xlsx"
        _write_excel(results, xlsx)
        written["excel"] = str(xlsx)
    figs: List[str] = []
    if do_figures:
        figs = _make_figures(nm_prep, results, out_dir / "figures", seed=seed)
        written["n_figures"] = str(len(figs))
    md = out_dir / "SINGLE_CELL_FINDINGS.md"
    _write_findings(results, md, figs)
    written["findings"] = str(md)

    meta = results["_meta"]
    if verbose:
        print(f"[singlecell] {meta['n_nuclei']} nuclei | A={meta['group_A']} (n={meta['n_A']}) "
              f"vs B={meta['group_B']} (n={meta['n_B']}) | abundance={meta['abundance_col']} | "
              f"{len(meta['metric_cols'])} metrics | {len(figs)} figures")
    return {"out_dir": str(out_dir), "written": written, "meta": meta,
            "n_metrics": len(meta["metric_cols"]), "n_figures": len(figs)}


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------
def main(argv: Optional[List[str]] = None) -> int:
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass
    ap = argparse.ArgumentParser(
        prog="python -m fishsuite.core.singlecell",
        description="Single-cell (per-nucleus) treatment analysis of a completed "
                    "run (CPU; reads nuclei_metrics.csv).")
    ap.add_argument("--run-dir", required=True, help="completed run output dir")
    ap.add_argument("--abundance-col", default=None,
                    help="per-nucleus abundance axis (default: nuclear_spot_count)")
    ap.add_argument("--condition-col", default="condition")
    ap.add_argument("--group-a", default=None, help="control group label")
    ap.add_argument("--group-b", default=None, help="perturbation group label")
    ap.add_argument("--include-secondary", action="store_true",
                    help="do NOT drop secondary-only nuclei")
    ap.add_argument("--no-figures", action="store_true")
    ap.add_argument("--no-excel", action="store_true")
    ap.add_argument("--out-subdir", default="singlecell")
    ap.add_argument("--seed", type=int, default=0)
    args = ap.parse_args(argv)

    res = singlecell_run(
        args.run_dir, abundance_col=args.abundance_col,
        condition_col=args.condition_col, group_a=args.group_a, group_b=args.group_b,
        exclude_secondary=not args.include_secondary,
        do_excel=not args.no_excel, do_figures=not args.no_figures,
        out_subdir=args.out_subdir, seed=args.seed,
    )
    print("written:", json.dumps(res["written"], indent=2))
    return 0


if __name__ == "__main__":
    sys.exit(main())
